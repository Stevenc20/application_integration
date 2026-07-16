#!/usr/bin/env python3
import sys, json, warnings
warnings.filterwarnings('ignore')  # suppress semua warning openpyxl

import openpyxl

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    filepath = sys.argv[1]

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # ── Sheet SPREEDSHEET ──
        spreedsheet = None
        for name in wb.sheetnames:
            if 'spreedsheet' in name.lower() or 'spreadsheet' in name.lower():
                spreedsheet = wb[name]
                break

        spreed_rows = []
        if spreedsheet:
            for row in spreedsheet.iter_rows(min_row=3, values_only=True):
                if not row[1]: continue
                spreed_rows.append({
                    'job_no':    str(row[1]  or '').strip(),
                    'item_name': str(row[2]  or '').strip(),
                    'proses':    str(row[4]  or '').strip(),
                    'source':    str(row[5]  or '').strip(),
                    'customer':  str(row[6]  or '').strip(),
                    'pcs_day':   float(row[7]  or 0),
                    'stock':     float(row[8]  or 0),
                    'strength':  float(row[9]  or 0),
                    'remarks':   str(row[10] or '').strip(),
                })

        # ── Sheet RUNDOWN STOCK FP ──
        rundown_sheet = None
        for name in wb.sheetnames:
            if 'rundown' in name.lower() and 'fp' in name.lower():
                rundown_sheet = wb[name]
                break

        rundown_rows = []
        if rundown_sheet:
            for row in rundown_sheet.iter_rows(min_row=7, values_only=True):
                if row[0] is None: continue
                try:
                    rundown_rows.append({
                        'no':             int(row[0]  or 0),
                        'job_no':         str(row[1]  or '').strip(),
                        'part_number':    str(row[3]  or '').strip(),
                        'sourching':      str(row[4]  or '').strip(),
                        'qty_palet':      float(row[5] or 0),
                        'type_pallet':    str(row[7]  or '').strip(),
                        'proses':         str(row[9]  or '').strip(),
                        'source':         str(row[11] or '').strip(),
                        'customer':       str(row[12] or '').strip(),
                        'type_of_part':   str(row[13] or '').strip(),
                        'stock_movement': str(row[14] or '').strip(),
                        'cycle_issue':    str(row[15] or '') if row[15] is not None else '',
                        'pcs_day':        float(row[16] or 0),
                        'stock_fg':       float(row[17] or 0),
                        'strength':       float(row[18] or 0),
                        'remarks':        str(row[19] or '').strip(),
                        'stock_sap':      float(row[20] or 0),
                        'stock_diff':     float(row[21] or 0),
                        'accuracy':       float(row[23] or 0),
                        'price_pcs':      float(row[24] or 0),
                        'new_price':      float(row[25] or 0),
                        'loss_gain':      float(row[26] or 0),
                        'pending_gi':     float(row[27] or 0),
                        'min_stock':      float(row[29] or 0) if row[29] else 0,
                        'max_stock':      float(row[30] or 0) if row[30] else 0,
                        'stock_shortage': float(row[33] or 0),
                        'status_order':   int(row[35]) if row[35] is not None else 0,
                    })
                except:
                    continue

        # ── Sheet STOCK PALLET SUBCONT ──
        pallet_sheet = None
        for name in wb.sheetnames:
            if 'pallet' in name.lower() and 'subcont' in name.lower():
                pallet_sheet = wb[name]
                break

        pallet_rows = []
        if pallet_sheet:
            for row in pallet_sheet.iter_rows(min_row=5, values_only=True):
                if len(row) < 10 or row[1] is None:
                    continue
                try:
                    no_val = int(row[1])
                    month_str = None
                    if hasattr(row[2], 'strftime'):
                        month_str = row[2].strftime('%Y-%m-%d')
                    elif row[2]:
                        month_str = str(row[2]).strip()
                    
                    pallet_rows.append({
                        'no': no_val,
                        'month': month_str,
                        'vendor': str(row[3] or '').strip(),
                        'type_pallet': str(row[4] or '').strip(),
                        'type': str(row[5] or '').strip(),
                        'initial_stock': int(row[6] or 0),
                        'pallet_in': int(row[7] or 0),
                        'pallet_out': int(row[8] or 0),
                        'final_stock': int(row[9] or 0),
                    })
                except ValueError:
                    continue

        # ── Sheet SMR VENDOR ──
        smr_sheet = None
        for name in wb.sheetnames:
            if 'smr' in name.lower() and 'vendor' in name.lower():
                smr_sheet = wb[name]
                break

        smr_rows = []
        if smr_sheet:
            for row in smr_sheet.iter_rows(min_row=4, values_only=True):
                if len(row) < 13 or row[2] is None:
                    continue
                try:
                    no_val = int(row[2])
                    
                    def parse_date(d):
                        if d is None:
                            return None
                        import datetime
                        if isinstance(d, datetime.datetime) or isinstance(d, datetime.date):
                            return d.strftime('%Y-%m-%d')
                        return None
                        
                    t_keluar = parse_date(row[9])
                    t_masuk = parse_date(row[10])
                    
                    smr_rows.append({
                        'no': no_val,
                        'month': str(row[3] or '').strip(),
                        'vendor': str(row[4] or '').strip(),
                        'no_smr': str(row[5] or '').strip() if row[5] is not None else '',
                        'part_name': str(row[6] or '').strip(),
                        'qty': int(row[7] or 0),
                        'problem': str(row[8] or '').strip(),
                        'tanggal_keluar': t_keluar,
                        'tanggal_masuk': t_masuk,
                        'qty_pengganti': int(row[11] or 0),
                        'status_barang': str(row[12] or '').strip(),
                    })
                except ValueError:
                    continue

        # ── Sheet SMR CUSTOMER ──
        smr_cust_sheet = None
        for name in wb.sheetnames:
            if 'smr' in name.lower() and 'customer' in name.lower():
                smr_cust_sheet = wb[name]
                break

        smr_cust_rows = []
        if smr_cust_sheet:
            for row in smr_cust_sheet.iter_rows(min_row=4, values_only=True):
                if len(row) < 18 or row[2] is None:
                    continue
                try:
                    no_val = int(row[2])
                    
                    def parse_date(d):
                        if d is None:
                            return None
                        import datetime
                        if isinstance(d, datetime.datetime) or isinstance(d, datetime.date):
                            return d.strftime('%Y-%m-%d')
                        return None

                    t_date = parse_date(row[4])
                    
                    def clean_val(val, val_type=int):
                        if val is None:
                            return 0
                        try:
                            return val_type(val)
                        except:
                            return 0

                    smr_cust_rows.append({
                        'no': no_val,
                        'year': clean_val(row[3], int),
                        'date': t_date,
                        'month': str(row[5] or '').strip(),
                        'quarterly': str(row[6] or '').strip(),
                        'no_smr': str(row[7] or '').strip() if row[7] is not None else '',
                        'job_no': str(row[8] or '').strip(),
                        'part_number': str(row[9] or '').strip(),
                        'part_name': str(row[10] or '').strip(),
                        'qty_smr': clean_val(row[11], int),
                        'total_production': clean_val(row[12], int),
                        'cost_rijection': clean_val(row[13], float),
                        'rijection_rate': clean_val(row[14], float),
                        'customer': str(row[15] or '').strip(),
                        'problem': str(row[16] or '').strip(),
                        'countermeasures': str(row[17] or '').strip(),
                    })
                except ValueError:
                    continue

        # ── Sheet DATA GR ──
        gr_sheet = None
        for name in wb.sheetnames:
            if 'data' in name.lower() and 'gr' in name.lower():
                gr_sheet = wb[name]
                break

        gr_rows = []
        if gr_sheet:
            for row in gr_sheet.iter_rows(min_row=5, values_only=True):
                if len(row) < 15 or row[2] is None or row[2] == 'GR Status':
                    continue
                try:
                    def parse_gr_date(d):
                        if d is None:
                            return None
                        import datetime
                        if isinstance(d, datetime.datetime) or isinstance(d, datetime.date):
                            return d.strftime('%Y-%m-%d %H:%M:%S')
                        return str(d).strip()

                    def parse_dn_date(d):
                        if d is None:
                            return None
                        import datetime
                        if isinstance(d, datetime.datetime) or isinstance(d, datetime.date):
                            return d.strftime('%Y-%m-%d')
                        return str(d).strip()

                    gr_rows.append({
                        'gr_status': str(row[2] or '').strip(),
                        'po_number': str(row[3] or '').strip() if row[3] is not None else '',
                        'job_number': str(row[4] or '').strip(),
                        'material': str(row[5] or '').strip(),
                        'vendor_name': str(row[6] or '').strip(),
                        'qty': int(row[7] or 0),
                        'dn_number': str(row[8] or '').strip(),
                        'kanban_number': str(row[9] or '').strip(),
                        'gr_number_edn': str(row[10] or '').strip(),
                        'dn_date': parse_dn_date(row[11]),
                        'gr_date': parse_gr_date(row[12]),
                        'gr_number_sap': str(row[13] or '').strip() if row[13] is not None else '',
                        'sap_message': str(row[14] or '').strip() if row[14] is not None else '',
                    })
                except Exception as e:
                    continue

        # ── Sheet DATA SCRAP ──
        scrap_sheet = None
        for name in wb.sheetnames:
            if 'data' in name.lower() and 'scrap' in name.lower():
                scrap_sheet = wb[name]
                break

        scrap_rows = []
        if scrap_sheet:
            for row in scrap_sheet.iter_rows(min_row=4, values_only=True):
                if len(row) < 16 or row[2] is None or row[2] == 'No':
                    continue
                try:
                    no_val = int(row[2])
                    
                    def clean_val(val, val_type=int):
                        if val is None:
                            return 0
                        try:
                            return val_type(val)
                        except:
                            return 0

                    scrap_rows.append({
                        'no': no_val,
                        'year': clean_val(row[3], int),
                        'month': str(row[4] or '').strip(),
                        'ba_no': str(row[5] or '').strip(),
                        'job_no': str(row[6] or '').strip(),
                        'sourch_1': str(row[7] or '').strip(),
                        'part_number': str(row[8] or '').strip(),
                        'part_name': str(row[9] or '').strip(),
                        'sourch_2': str(row[10] or '').strip(),
                        'customer': str(row[11] or '').strip(),
                        'qty': clean_val(row[12], int),
                        'value': clean_val(row[13], float),
                        'total_production': clean_val(row[14], int),
                        'reject_rate': clean_val(row[15], float),
                    })
                except Exception as e:
                    continue

        wb.close()

        # Output JSON murni ke stdout
        sys.stdout.write(json.dumps({
            'success': True,
            'spreedsheet': {'count': len(spreed_rows), 'data': spreed_rows},
            'rundown':     {'count': len(rundown_rows), 'data': rundown_rows},
            'pallet':      {'count': len(pallet_rows), 'data': pallet_rows},
            'smr_vendor':  {'count': len(smr_rows), 'data': smr_rows},
            'smr_customer':{'count': len(smr_cust_rows), 'data': smr_cust_rows},
            'data_gr':     {'count': len(gr_rows), 'data': gr_rows},
            'data_scrap':  {'count': len(scrap_rows), 'data': scrap_rows},
        }))

    except Exception as e:
        sys.stdout.write(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()