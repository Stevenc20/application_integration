#!/usr/bin/env python3
import sys
import os
import json
import warnings
import openpyxl
from datetime import datetime

warnings.filterwarnings('ignore')

if len(sys.argv) < 3:
    print(json.dumps({'error': 'Usage: python read_rundown_incoming_monthly.py <excel_file> <price_file>'}))
    sys.exit(1)

data_file = sys.argv[1]
price_file = sys.argv[2]

try:
    # 1. Load price data from ADD.xlsx (optional — skipped if not found)
    price_data = {}
    if price_file and price_file.strip() and os.path.isfile(price_file):
        wb_price = openpyxl.load_workbook(price_file, data_only=True)
        ws_price = wb_price.active
        for row in range(6, ws_price.max_row + 1):
            job_no = ws_price.cell(row, 2).value
            cycle  = ws_price.cell(row, 3).value
            customer = ws_price.cell(row, 4).value
            price  = ws_price.cell(row, 5).value
            if job_no:
                job_no_clean = str(job_no).strip().upper()
                price_data[job_no_clean] = {
                    'customer': str(customer).strip() if customer else '',
                    'cycle_issue': int(cycle) if cycle and str(cycle).replace('.', '').isdigit() else 1,
                    'price': float(price) if price and str(price).replace('.', '').replace(',', '').isdigit() else 0
                }

    # 2. Read all sheets — use dict keyed by (date, job_no) for deduplication.
    #    Sheets are read in order; LATER sheets override EARLIER ones for the same
    #    (date, job_no), so a combined FINISH+SINGLE sheet (like Sheet1) always wins
    #    over a historical-only sheet (like Worksheet).
    #
    # all_sheets_dict[date][job_no] = row_data
    all_sheets_dict = {}

    wb_data = openpyxl.load_workbook(data_file, data_only=True)

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]

        header_row = None
        default_cat = 'SINGLE PART'

        for r in range(1, 15):
            for c in range(1, 10):
                val = ws.cell(r, c).value
                if val:
                    v_str = str(val).upper()
                    if 'FINISH PART' in v_str:
                        default_cat = 'FINISH PART'
                    if 'JOB NO' in v_str:
                        header_row = r
            if header_row:
                break

        if not header_row:
            continue

        # Build column map from header row
        col_map = {}
        for c in range(1, 45):
            h_val = ws.cell(header_row, c).value
            if not h_val:
                continue
            h_name = str(h_val).strip().upper()
            if h_name in ('JOB NO', 'PART NO'):
                col_map['job_no'] = c
            elif h_name in ('DATE', 'TGL', 'TANGGAL'):
                col_map['date'] = c
            elif h_name in ('NO', '#'):
                col_map['no'] = c
            elif ('FINISH' in h_name and ('JOB' in h_name or 'NO' in h_name)) or h_name == 'JOB NO FINISH':
                col_map['job_no_finish'] = c
            elif 'PALLET' in h_name or 'PALL' in h_name:
                col_map['type_pallet'] = c
            elif 'VENDOR' in h_name:
                col_map['vendor'] = c
            elif 'MATERIAL' in h_name or 'FINISH PART' in h_name or 'KATEGORI' in h_name:
                col_map['category'] = c
            elif 'S. AWAL' in h_name or 'STOCK AWAL' in h_name:
                col_map['stock_awal'] = c
            elif 'ASSY' in h_name:
                col_map['assy'] = c
            elif 'DELIVERY' in h_name:
                col_map['delivery'] = c
            elif 'INCOMING' in h_name:
                col_map['incoming'] = c
            elif 'S. AKHIR' in h_name or 'STOK AKHIR' in h_name:
                col_map['stok_akhir'] = c
            elif 'PCS / DAY' in h_name or 'PCS/DAY' in h_name:
                col_map['pcs_day'] = c
            elif 'STRENGHT' in h_name or 'STRENGTH' in h_name:
                col_map['strength'] = c
            elif h_name in ('IAMI',) or 'IAMI' in h_name:
                col_map['iami'] = c
            elif h_name in ('GKD',) or 'GKD' in h_name:
                col_map['gkd'] = c
            elif h_name in ('SAP',) or 'SAP' in h_name:
                col_map['sap'] = c
            elif h_name in ('KAP',) or 'KAP' in h_name:
                col_map['kap'] = c
            elif any(x in h_name for x in ['GMO', 'TMMIN', 'FTI', 'PTI', 'IKAR', 'IKA']):
                col_map['gmo'] = c
            elif 'CUSTOMER' in h_name:
                col_map['customer'] = c
            elif 'PRICE' in h_name:
                col_map['price_pc'] = c      # PRICE/PC, PRICE/PCS, PRICE PCS …
            elif 'MOVEMENT' in h_name:
                col_map['movement'] = c
            elif 'CYCLE' in h_name:
                col_map['cycle_issue'] = c
            elif 'STATUS' in h_name:
                col_map['status_col'] = c

        def safe_f(v):
            if v is None:
                return 0
            try:
                return float(v)
            except Exception:
                return 0

        for row in range(header_row + 1, ws.max_row + 1):
            job_no = ws.cell(row, col_map.get('job_no', 2)).value
            if not job_no:
                continue

            row_date = str(ws.cell(row, col_map['date']).value).strip() if 'date' in col_map else sheet_name
            if not row_date or row_date == 'None':
                row_date = sheet_name

            job_no_str   = str(job_no).strip()
            job_no_clean = job_no_str.upper()

            vendor        = ws.cell(row, col_map['vendor']).value if 'vendor' in col_map else None
            job_no_finish = ws.cell(row, col_map['job_no_finish']).value if 'job_no_finish' in col_map else ''
            type_pallet   = ws.cell(row, col_map['type_pallet']).value if 'type_pallet' in col_map else ''
            material      = ws.cell(row, col_map['category']).value if 'category' in col_map else ''
            no_val        = ws.cell(row, col_map['no']).value if 'no' in col_map else None

            stock_awal = safe_f(ws.cell(row, col_map['stock_awal']).value) if 'stock_awal' in col_map else 0
            assy       = safe_f(ws.cell(row, col_map['assy']).value) if 'assy' in col_map else 0
            incoming   = safe_f(ws.cell(row, col_map['incoming']).value) if 'incoming' in col_map else 0
            stok_akhir = safe_f(ws.cell(row, col_map['stok_akhir']).value) if 'stok_akhir' in col_map else 0
            pcs_day    = safe_f(ws.cell(row, col_map['pcs_day']).value) if 'pcs_day' in col_map else 0
            strength   = safe_f(ws.cell(row, col_map['strength']).value) if 'strength' in col_map else 0
            iami       = safe_f(ws.cell(row, col_map['iami']).value) if 'iami' in col_map else 0
            gkd        = safe_f(ws.cell(row, col_map['gkd']).value) if 'gkd' in col_map else 0
            sap        = safe_f(ws.cell(row, col_map['sap']).value) if 'sap' in col_map else 0
            kap        = safe_f(ws.cell(row, col_map['kap']).value) if 'kap' in col_map else 0
            gmo        = safe_f(ws.cell(row, col_map['gmo']).value) if 'gmo' in col_map else 0
            delivery   = ws.cell(row, col_map['delivery']).value if 'delivery' in col_map else ''

            # Price: read from Excel first, fallback to ADD.xlsx
            price_from_excel = safe_f(ws.cell(row, col_map['price_pc']).value) if 'price_pc' in col_map else 0
            price_info = price_data.get(job_no_clean, {'customer': '', 'cycle_issue': 1, 'price': 0})
            price_pc   = price_from_excel if price_from_excel > 0 else price_info['price']

            # Customer: Excel column first, fallback to ADD.xlsx
            customer_val = ws.cell(row, col_map['customer']).value if 'customer' in col_map else None
            customer     = str(customer_val).strip() if customer_val else price_info['customer']

            all_price = stok_akhir * price_pc

            # Category
            category = str(material).strip().upper() if material else default_cat
            category = 'FINISH PART' if 'FINISH' in category else 'SINGLE PART'

            if pcs_day > 0 and strength == 0:
                strength = round(stok_akhir / pcs_day, 2)

            status = 'STANDAR'
            if strength <= 2:
                status = 'CRITICAL'
            elif strength > 5:
                status = 'OVER'

            # Movement from Excel, else derive
            movement_val = ws.cell(row, col_map['movement']).value if 'movement' in col_map else None
            movement = str(movement_val).strip() if movement_val else ('FAST MOVING' if strength > 0.5 else 'SLOW MOVING')

            # Cycle issue from Excel, else ADD.xlsx
            cycle_raw = ws.cell(row, col_map['cycle_issue']).value if 'cycle_issue' in col_map else None
            try:
                cycle_issue = int(float(cycle_raw)) if cycle_raw is not None else price_info['cycle_issue']
            except Exception:
                cycle_issue = price_info['cycle_issue']

            # --- Deduplication: later sheet overrides earlier sheet ---
            if row_date not in all_sheets_dict:
                all_sheets_dict[row_date] = {}

            all_sheets_dict[row_date][job_no_str] = {
                'no': no_val,
                'job_no': job_no_str,
                'job_no_finish': str(job_no_finish).strip() if job_no_finish else '',
                'type_pallet': str(type_pallet).strip() if type_pallet else '',
                'vendor': str(vendor).strip() if vendor else '',
                'category': category,
                'customer': customer,
                'price_pc': price_pc,
                'status': status,
                'movement': movement,
                'cycle_issue': cycle_issue,
                'stock_awal': stock_awal,
                'assy': assy,
                'iami': iami,
                'gkd': gkd,
                'sap': sap,
                'kap': kap,
                'gmo': gmo,
                'delivery': str(delivery).strip() if delivery else '',
                'incoming': incoming,
                'stok_akhir': stok_akhir,
                'all_price': all_price,
                'pcs_day': pcs_day,
                'strength': strength,
            }

    # 3. Convert dict back to list, assign sequential 'no' if needed
    all_sheets = {}
    for date, rows_dict in all_sheets_dict.items():
        rows_list = list(rows_dict.values())
        # Re-number 'no' sequentially so there are no gaps
        for idx, r in enumerate(rows_list, start=1):
            if r['no'] is None:
                r['no'] = idx
        all_sheets[date] = rows_list

    print(json.dumps({'success': True, 'total_sheets': len(all_sheets), 'sheets': all_sheets}, ensure_ascii=False))

except Exception as e:
    print(json.dumps({'error': str(e)}))
