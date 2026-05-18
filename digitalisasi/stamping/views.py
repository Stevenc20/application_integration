from django.shortcuts import render, redirect, get_object_or_404
from . import models
from django.contrib import messages
from datetime import datetime
from django.utils import timezone
from django.db import transaction
from dateutil import parser
from django.utils.dateparse import parse_datetime
from django.db.models import Sum
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.http import JsonResponse
import json
import datetime
from django.db.models import F
import math
from itertools import zip_longest
from django.http import JsonResponse
from .models import detail_dt, downtime
from django.http import HttpResponse
import openpyxl
from io import BytesIO
from django.template.loader import get_template
from weasyprint import HTML, CSS 
from django.conf import settings 
import os 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
import copy
from .forms import LoginForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout

def custom_login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            # 'authenticate' sekarang sudah dikenal
            user = authenticate(request, username=username, password=password) 
            
            if user is not None:
                # 'login' juga sudah dikenal
                login(request, user) 
                
                # --- Logika redirect berdasarkan role ---
                if user.groups.filter(name='supervisor').exists():
                    return redirect('dashboard') 
                elif user.groups.filter(name='admin').exists():
                     # Ganti 'productionline' dengan URL data master Anda jika perlu
                    return redirect('productionline')
                elif user.groups.filter(name='foreman').exists():
                    # Foreman diarahkan ke dashboard (sesuai rencana)
                    return redirect('dashboard')
                elif user.groups.filter(name='leader').exists():
                    # Leader diarahkan ke input harian (sesuai rencana)
                    return redirect('input_harian') 
                else:
                    # Default jika tidak ada role yang cocok
                    return redirect('dashboard')
            
            else:
                messages.error(request, "Username atau password salah.")
        else:
            # Jika form tidak valid (misal: field kosong)
            error_msg = "Form tidak valid. Silakan periksa kembali isian Anda."
            # Anda bisa menambahkan detail error dari form jika perlu:
            # error_msg += str(form.errors) 
            messages.error(request, error_msg)
            
    else: # Jika request.method == 'GET'
        form = LoginForm()
        
    return render(request, 'registration/login.html', {'form': form})


def format_time(minutes):
    if minutes is None or minutes == 0:
        return "0.00 Menit"
    
    # Selalu tampilkan dalam format menit dengan 2 angka desimal
    return f"{minutes:.2f} Menit"


# FUNGSI UTAMA UNTUK DASHBOARD
@login_required
def dashboard(request):
    all_breaks = models.BreakTime.objects.all()
    lines = models.productionline.objects.values_list("namaline", flat=True).distinct()
    data = {}
    today = timezone.now().date()

    for line in lines:
        # --- 1. AMBIL DATA PER LINE ---
        jobs_on_schedule = models.detailjob.objects.filter(
            id_job__id_productionline__namaline=line,
            id_job__date=today
        ).order_by('urutan').select_related(
            'id_itemproduksi', 'id_job__id_productionline', 'id_job'
        ).prefetch_related('machine_used', 'dandoris', 'qchecks', 'problems')

        if not jobs_on_schedule.exists():
            continue

        # --- 2. PERHITUNGAN WAKTU & LOOP ---
        current_running_detail_job = None
        actual_job_item_index = 0
        current_shift = jobs_on_schedule.first().id_job.id_productionline.shift
        shift_start = timezone.make_aware(datetime.datetime.combine(today, datetime.time(7, 40) if current_shift == 1 else datetime.time(21, 0)))
        plan_running_time = shift_start
        actual_running_time = shift_start

        total_actual_gsph = 0
        total_actual_stroke = 0
        total_actual_tpt_minutes_finished_jobs = 0

        for index, job in enumerate(jobs_on_schedule):
            job_plan_start_time = plan_running_time
            job_actual_start_time = actual_running_time
            press_time_minutes = (job.actual_qty * float(job.id_itemproduksi.cycle_time)) / 60 if job.id_itemproduksi.cycle_time and float(job.id_itemproduksi.cycle_time) > 0 else 0
            dandori_time = job.total_dandori_minutes
            iq_check_time = job.total_qcheck_minutes
            downtime_time = job.total_downtime_minutes
            plan_tpt_minutes = press_time_minutes + dandori_time + iq_check_time
            item_finish_plan = calculate_finish_time(job_plan_start_time, plan_tpt_minutes, all_breaks, current_shift)
            actual_tpt_minutes = plan_tpt_minutes + downtime_time
            item_finish_actual = calculate_finish_time(job_actual_start_time, actual_tpt_minutes, all_breaks, current_shift)
            
            job_gsph = 0
            num_machines = job.machine_used.count()
            if actual_tpt_minutes > 0 and num_machines > 0:
                job_gsph = ((job.actual_qty * num_machines) / actual_tpt_minutes) * 60
            
            job_stroke = job.actual_qty * num_machines
            
            # Hanya tambahkan GSPH dan Stroke dari job yang sudah selesai
            if item_finish_actual < timezone.now():
                total_actual_stroke += job_stroke
                total_actual_tpt_minutes_finished_jobs += actual_tpt_minutes          
            
            plan_running_time = item_finish_plan
            actual_running_time = item_finish_actual

            if job_plan_start_time <= timezone.now() < (item_finish_plan or (job_plan_start_time + datetime.timedelta(days=1))):
                current_running_detail_job = job
                actual_job_item_index = index + 1

        if total_actual_tpt_minutes_finished_jobs > 0:
            total_actual_gsph = (total_actual_stroke / total_actual_tpt_minutes_finished_jobs) * 60
        else:
            total_actual_gsph = 0
        
        # --- 3. PERHITUNGAN KPI LENGKAP ---
        plans_in_line_today = models.productionplan.objects.filter(id_job__id_productionline__namaline=line, id_job__date=today)
        total_plan_job_sum = jobs_on_schedule.aggregate(Sum("plan_qty"))["plan_qty__sum"] or 0
        total_actual_job_sum = jobs_on_schedule.aggregate(Sum("actual_qty"))["actual_qty__sum"] or 0
        total_repair = sum(j.actual_repair for j in jobs_on_schedule)
        total_reject = sum(j.actual_reject for j in jobs_on_schedule)
        plan_job_items = jobs_on_schedule.count()
        actual_job_count = jobs_on_schedule.filter(actual_qty__gt=0).count()
        actual_job_items_str = f"{actual_job_count} / {plan_job_items}"
        total_plan_gsph = plans_in_line_today.aggregate(Sum("gsph_plan"))["gsph_plan__sum"] or 0
        total_plan_stroke = plans_in_line_today.aggregate(Sum("stroke_plan"))["stroke_plan__sum"] or 0

        current_gsph_actual, current_stroke_actual, current_repair_qty, current_reject_qty, current_repair_percent, current_reject_percent, actual_qty_str = (0, 0, 0, 0, 0, 0, "0 / 0")
        
        if current_running_detail_job:
            job = current_running_detail_job

            # HITUNG ULANG TPT KHUSUS UNTUK CURRENT JOB
            current_press_time = (job.actual_qty * float(job.id_itemproduksi.cycle_time)) / 60 if job.id_itemproduksi.cycle_time and float(job.id_itemproduksi.cycle_time) > 0 else 0
            current_dandori_time = job.total_dandori_minutes
            current_qcheck_time = job.total_qcheck_minutes
            current_downtime_time = job.total_downtime_minutes
            current_actual_tpt = current_press_time + current_dandori_time + current_qcheck_time + current_downtime_time

            num_machines = job.machine_used.count()
            
            # Gunakan TPT yang baru dihitung
            if current_actual_tpt > 0 and num_machines > 0:
                current_gsph_actual = ((job.actual_qty * num_machines) / current_actual_tpt) * 60
            
            current_stroke_actual = job.actual_qty * num_machines
            current_repair_qty = job.actual_repair
            current_reject_qty = job.actual_reject
            actual_qty_current = job.actual_qty
            current_repair_percent = (current_repair_qty / actual_qty_current) * 100 if actual_qty_current > 0 else 0
            current_reject_percent = (current_reject_qty / actual_qty_current) * 100 if actual_qty_current > 0 else 0
            plan_qty_current = job.plan_qty
            actual_qty_str = f"{actual_qty_current} / {plan_qty_current}" if plan_qty_current > 0 else "0 / 0"

        plan_repair_percent, plan_reject_percent = (0, 0)
        if current_running_detail_job:
            try:
                plan_terkait = current_running_detail_job.id_job.productionplan
                plan_repair_percent = plan_terkait.repair_plan
                plan_reject_percent = plan_terkait.reject_plan
            except models.productionplan.DoesNotExist: pass
        plan_repair_qty = total_plan_job_sum * (plan_repair_percent / 100) if plan_repair_percent else 0
        plan_reject_qty = total_plan_job_sum * (plan_reject_percent / 100) if plan_reject_percent else 0
        actual_repair_percent = (total_repair / total_plan_job_sum) * 100 if total_plan_job_sum > 0 else 0
        actual_reject_percent = (total_reject / total_plan_job_sum) * 100 if total_plan_job_sum > 0 else 0
        total_plan_ot = plans_in_line_today.aggregate(Sum("ot_plan"))["ot_plan__sum"] or 0
        
        downtime_for_current_item = {code: 0 for code, name in models.detail_dt._meta.get_field('jenisdowntime').choices}
        idle_for_current_item = 0
        if current_running_detail_job:
            idle_for_current_item = current_running_detail_job.total_idle_minutes
            dts_current_item = models.detail_dt.objects.filter(id_downtime__id_detailjob=current_running_detail_job)
            for dt in dts_current_item:
                if dt.jenisdowntime in downtime_for_current_item: downtime_for_current_item[dt.jenisdowntime] += dt.duration_downtime or 0
        total_downtime_current_item = sum(downtime_for_current_item.values())
        
        total_idle_day = sum(j.total_idle_minutes for j in jobs_on_schedule)
        dts_day = models.detail_dt.objects.filter(id_downtime__id_detailjob__in=jobs_on_schedule)
        downtime_choices_day = {code: 0 for code, name in models.detail_dt._meta.get_field('jenisdowntime').choices}
        for dt in dts_day:
            if dt.jenisdowntime in downtime_choices_day: downtime_choices_day[dt.jenisdowntime] += dt.duration_downtime or 0
        total_downtime_day = sum(downtime_choices_day.values())



        # --- 4. MASUKKAN DATA KE DICTIONARY ---
        data[line] = {
            "JOB": {"plan": plan_job_items, "actual": actual_job_items_str, "current": current_running_detail_job.id_itemproduksi.job_number if current_running_detail_job else "Selesai"},
            "QTY": {"plan": total_plan_job_sum, "actual": f"{total_actual_job_sum} / {total_plan_job_sum}", "current": actual_qty_str},
            "GSPH": {"plan": total_plan_gsph, "actual": round(total_actual_gsph), "current": round(current_gsph_actual)},
            "STROKE": {"plan": total_plan_stroke, "actual": round(total_actual_stroke), "current": current_stroke_actual},
            "REPAIR": {"plan_qty": plan_repair_qty, "plan_percent": plan_repair_percent, "actual_qty": total_repair, "actual_percent": actual_repair_percent, "current_qty": current_repair_qty, "current": current_repair_percent},
            "REJECT": {"plan_qty": plan_reject_qty, "plan_percent": plan_reject_percent, "actual_qty": total_reject, "actual_percent": actual_reject_percent, "current_qty": current_reject_qty, "current": current_reject_percent},
            "IDLE_T": {"plan": "-", "actual": total_idle_day, "current": idle_for_current_item},
            "OVERTIME": {"plan": total_plan_ot, "actual": total_downtime_day, "current": total_downtime_current_item},
            "TOTAL_DT": {"plan": "-", "actual": total_downtime_day, "current": total_downtime_current_item},
            "PROD_T": {"plan": "-", "actual": downtime_choices_day.get('prod_t', 0), "current": downtime_for_current_item.get('prod_t', 0)},
            "MAT_T": {"plan": "-", "actual": downtime_choices_day.get('mat_t', 0), "current": downtime_for_current_item.get('mat_t', 0)},
            "DIES_T": {"plan": "-", "actual": downtime_choices_day.get('dies_t', 0), "current": downtime_for_current_item.get('dies_t', 0)},
            "MACH_T": {"plan": "-", "actual": downtime_choices_day.get('mach_t', 0), "current": downtime_for_current_item.get('mach_t', 0)},
            "LOG_T": {"plan": "-", "actual": downtime_choices_day.get('log_t', 0), "current": downtime_for_current_item.get('log_t', 0)},
        }
        
    return render(request, "dashboard/achievement.html", {"data": data})

#dashboard detail per line
def calculate_finish_time(start_time, duration_minutes, all_breaks, current_shift):
    if not duration_minutes or duration_minutes <= 0:
        return start_time
    
    remaining_duration = datetime.timedelta(minutes=float(duration_minutes))
    temp_running_time = start_time
    
    while remaining_duration.total_seconds() > 0:
        weekday = temp_running_time.weekday()
        if weekday == 6:  # Skip Minggu
            next_day = temp_running_time.date() + datetime.timedelta(days=1)
            shift_start_time_obj = datetime.time(7, 40) if current_shift == 1 else datetime.time(21, 0)
            temp_running_time = timezone.make_aware(datetime.datetime.combine(next_day, shift_start_time_obj))
            continue

        breaks = sorted([b for b in all_breaks if b.shift == current_shift and (b.hari is None or b.hari == weekday)], key=lambda x: x.waktu_mulai)
        next_break_start, next_break_end = None, None
        for br in breaks:
            p_start = timezone.make_aware(datetime.datetime.combine(temp_running_time.date(), br.waktu_mulai))
            if p_start >= temp_running_time:
                next_break_start = p_start
                next_break_end = timezone.make_aware(datetime.datetime.combine(temp_running_time.date(), br.waktu_selesai))
                if next_break_end <= next_break_start:
                    next_break_end += datetime.timedelta(days=1)
                break
        
        shift_end_time_obj = datetime.time(21, 0) if current_shift == 1 else datetime.time(7, 40)
        shift_ends_at = timezone.make_aware(datetime.datetime.combine(temp_running_time.date(), shift_end_time_obj))
        if temp_running_time >= shift_ends_at:
             shift_ends_at += datetime.timedelta(days=1)
        
        work_ends = next_break_start or shift_ends_at
        workable_time = work_ends - temp_running_time
        
        if remaining_duration.total_seconds() > 0 and workable_time.total_seconds() <= 0:
            temp_running_time = next_break_end or shift_ends_at
            continue

        if remaining_duration <= workable_time:
            temp_running_time += remaining_duration
            remaining_duration = datetime.timedelta(seconds=0)
        else:
            remaining_duration -= workable_time
            temp_running_time = next_break_end or shift_ends_at
            if temp_running_time == shift_ends_at:
                 next_day_start_time = datetime.time(7, 40) if current_shift == 1 else datetime.time(21, 0)
                 temp_running_time = timezone.make_aware(datetime.datetime.combine(shift_ends_at.date(), next_day_start_time))
    
    return temp_running_time

@login_required
def dashboard_detail(request, line_name):
    # --- 1. Inisialisasi & Ambil Data ---
    formatted_line_name = line_name.replace('-', ' ').title()
    all_breaks = models.BreakTime.objects.all()
    today = timezone.now().date()
    
    jobs_on_schedule = models.detailjob.objects.filter(
        id_job__id_productionline__namaline=formatted_line_name,
        id_job__date=today
    ).order_by('urutan').select_related(
        'id_itemproduksi', 'id_job__id_productionline', 'id_job'
    ).prefetch_related(
        'machine_used', 'dandoris', 'qchecks', 'problems'
    )

    if not jobs_on_schedule.exists():
        return render(request, "dashboard/dashboard_detail.html", {'line_name': formatted_line_name, 'table_rows': []})

    # --- 2. Siapkan Data Rincian Item & Hitung Waktu ---
    detail_jobs_list = []
    current_running_detail_job = None
    
    current_shift = jobs_on_schedule.first().id_job.id_productionline.shift
    shift_start = timezone.make_aware(datetime.datetime.combine(today, datetime.time(7, 40) if current_shift == 1 else datetime.time(21, 0)))
    
    plan_running_time = shift_start
    actual_running_time = shift_start
    
    total_actual_gsph = 0
    total_actual_stroke = 0
    total_actual_tpt_minutes_finished_jobs = 0

    for index, job in enumerate(jobs_on_schedule):
        job_plan_start_time = plan_running_time
        job_actual_start_time = actual_running_time
        press_time_minutes = (job.actual_qty * float(job.id_itemproduksi.cycle_time)) / 60 if job.id_itemproduksi.cycle_time and float(job.id_itemproduksi.cycle_time) > 0 else 0
        dandori_time = job.total_dandori_minutes
        iq_check_time = job.total_qcheck_minutes
        downtime_time = job.total_downtime_minutes
        plan_tpt_minutes = press_time_minutes + dandori_time + iq_check_time
        item_finish_plan = calculate_finish_time(job_plan_start_time, plan_tpt_minutes, all_breaks, current_shift)
        actual_tpt_minutes = plan_tpt_minutes + downtime_time
        item_finish_actual = calculate_finish_time(job_actual_start_time, actual_tpt_minutes, all_breaks, current_shift)
        
        job_gsph = 0
        num_machines = job.machine_used.count()
        if actual_tpt_minutes > 0:
            job_gsph = ((job.actual_qty * num_machines) / actual_tpt_minutes) * 60
        
        job_stroke = job.actual_qty * num_machines
        
        # Hanya tambahkan GSPH dan Stroke dari job yang sudah selesai
        if item_finish_actual < timezone.now():
            #total_actual_gsph += job_gsph
            total_actual_stroke += job_stroke
            total_actual_tpt_minutes_finished_jobs += actual_tpt_minutes
        
        plan_running_time = item_finish_plan
        actual_running_time = item_finish_actual

        if job_plan_start_time <= timezone.now() < (item_finish_plan or (job_plan_start_time + datetime.timedelta(days=1))):
            current_running_detail_job = job
            
        detail_jobs_list.append({
            'job_number': job.id_itemproduksi.job_number, 'machine_used': [m.code for m in job.machine_used.all()],
            'plan_qty': job.plan_qty, 'actual_good': job.actual_qty, 'actual_repair': job.actual_repair, 
            'actual_reject': job.actual_reject, 'press_time': press_time_minutes, 'dandori_time': dandori_time,
            'iq_check_time': iq_check_time, 'downtime_time': downtime_time, 'tpt': actual_tpt_minutes,
            'finish_time_plan': item_finish_plan, 'finish_time_actual': item_finish_actual
        })

        if total_actual_tpt_minutes_finished_jobs > 0:
                total_actual_gsph = (total_actual_stroke / total_actual_tpt_minutes_finished_jobs) * 60
        else:
            total_actual_gsph = 0
    # --- 3. PERHITUNGAN KPI LENGKAP ---
    plans_in_line_today = models.productionplan.objects.filter(id_job__id_productionline__namaline=formatted_line_name, id_job__date=today)
    total_plan_job_sum = jobs_on_schedule.aggregate(Sum("plan_qty"))["plan_qty__sum"] or 0
    total_actual_job_sum = jobs_on_schedule.aggregate(Sum("actual_qty"))["actual_qty__sum"] or 0
    total_repair = sum(j.actual_repair for j in jobs_on_schedule)
    total_reject = sum(j.actual_reject for j in jobs_on_schedule)
    plan_job_items = jobs_on_schedule.count()
    actual_job_count = jobs_on_schedule.filter(actual_qty__gt=0).count()
    actual_job_items_str = f"{actual_job_count} / {plan_job_items}"
    total_plan_gsph = plans_in_line_today.aggregate(Sum("gsph_plan"))["gsph_plan__sum"] or 0
    total_plan_stroke = plans_in_line_today.aggregate(Sum("stroke_plan"))["stroke_plan__sum"] or 0

    # --- PERHITUNGAN UNTUK KOLOM CURRENT ---
    current_gsph_actual, current_stroke_actual, current_repair_qty, current_reject_qty, current_repair_percent, current_reject_percent, actual_qty_str = (0, 0, 0, 0, 0, 0, "0 / 0")
    if current_running_detail_job:
        job = current_running_detail_job
        current_press_time = (job.actual_qty * float(job.id_itemproduksi.cycle_time)) / 60 if job.id_itemproduksi.cycle_time and float(job.id_itemproduksi.cycle_time) > 0 else 0
        current_dandori_time = job.total_dandori_minutes
        current_qcheck_time = job.total_qcheck_minutes
        current_downtime_time = job.total_downtime_minutes
        current_actual_tpt = current_press_time + current_dandori_time + current_qcheck_time + current_downtime_time
        num_machines = job.machine_used.count()
        if current_actual_tpt > 0:
            current_gsph_actual = ((job.actual_qty * num_machines) / current_actual_tpt) * 60
        current_stroke_actual = job.actual_qty * num_machines
        current_repair_qty, current_reject_qty, actual_qty_current = (job.actual_repair, job.actual_reject, job.actual_qty)
        current_repair_percent = (current_repair_qty / actual_qty_current) * 100 if actual_qty_current > 0 else 0
        current_reject_percent = (current_reject_qty / actual_qty_current) * 100 if actual_qty_current > 0 else 0
        plan_qty_current = job.plan_qty
        actual_qty_str = f"{actual_qty_current} / {plan_qty_current}" if plan_qty_current > 0 else "0 / 0"

    # --- Perhitungan KPI Lainnya ---
    plan_repair_percent, plan_reject_percent = (0, 0)
    if current_running_detail_job:
        try:
            plan_terkait = current_running_detail_job.id_job.productionplan
            plan_repair_percent = plan_terkait.repair_plan
            plan_reject_percent = plan_terkait.reject_plan
        except models.productionplan.DoesNotExist: pass
    plan_repair_qty = total_plan_job_sum * (plan_repair_percent / 100) if plan_repair_percent else 0
    plan_reject_qty = total_plan_job_sum * (plan_reject_percent / 100) if plan_reject_percent else 0
    actual_repair_percent = (total_repair / total_plan_job_sum) * 100 if total_plan_job_sum > 0 else 0
    actual_reject_percent = (total_reject / total_plan_job_sum) * 100 if total_plan_job_sum > 0 else 0
    total_plan_ot = plans_in_line_today.aggregate(Sum("ot_plan"))["ot_plan__sum"] or 0
    downtime_for_current_item = {code: 0 for code, name in models.detail_dt._meta.get_field('jenisdowntime').choices}
    idle_for_current_item = 0
    if current_running_detail_job:
        idle_for_current_item = current_running_detail_job.total_idle_minutes
        dts_current_item = models.detail_dt.objects.filter(id_downtime__id_detailjob=current_running_detail_job)
        for dt in dts_current_item:
            if dt.jenisdowntime in downtime_for_current_item: downtime_for_current_item[dt.jenisdowntime] += dt.duration_downtime or 0
    total_downtime_current_item = sum(downtime_for_current_item.values())
    total_idle_day = sum(j.total_idle_minutes for j in jobs_on_schedule)
    dts_day = models.detail_dt.objects.filter(id_downtime__id_detailjob__in=jobs_on_schedule)
    downtime_choices_day = {code: 0 for code, name in models.detail_dt._meta.get_field('jenisdowntime').choices}
    for dt in dts_day:
        if dt.jenisdowntime in downtime_choices_day: downtime_choices_day[dt.jenisdowntime] += dt.duration_downtime or 0
    total_downtime_day = sum(downtime_choices_day.values())
    
    kpi_data_dict = {
        "JOB": {"plan": plan_job_items, "actual": actual_job_items_str, "current": current_running_detail_job.id_itemproduksi.job_number if current_running_detail_job else "Selesai"},
        "QTY": {"plan": total_plan_job_sum, "actual": f"{total_actual_job_sum} / {total_plan_job_sum}", "current": actual_qty_str},
        "GSPH": {"plan": total_plan_gsph, "actual": round(total_actual_gsph), "current": round(current_gsph_actual)},
        "STROKE": {"plan": total_plan_stroke, "actual": round(total_actual_stroke), "current": current_stroke_actual},
        "REPAIR": {"plan_qty": plan_repair_qty, "plan_percent": plan_repair_percent, "actual_qty": total_repair, "actual_percent": actual_repair_percent, "current_qty": current_repair_qty, "current": current_repair_percent},
        "REJECT": {"plan_qty": plan_reject_qty, "plan_percent": plan_reject_percent, "actual_qty": total_reject, "actual_percent": actual_reject_percent, "current_qty": current_reject_qty, "current": current_reject_percent},
        "IDLE_T": {"plan": "-", "actual": total_idle_day, "current": idle_for_current_item},
        "OVERTIME": {"plan": total_plan_ot, "actual": total_downtime_day, "current": total_downtime_current_item},
        "TOTAL_DT": {"plan": "-", "actual": total_downtime_day, "current": total_downtime_current_item},
        "PROD_T": {"plan": "-", "actual": downtime_choices_day.get('prod_t', 0), "current": downtime_for_current_item.get('prod_t', 0)},
        "MAT_T": {"plan": "-", "actual": downtime_choices_day.get('mat_t', 0), "current": downtime_for_current_item.get('mat_t', 0)},
        "DIES_T": {"plan": "-", "actual": downtime_choices_day.get('dies_t', 0), "current": downtime_for_current_item.get('dies_t', 0)},
        "MACH_T": {"plan": "-", "actual": downtime_choices_day.get('mach_t', 0), "current": downtime_for_current_item.get('mach_t', 0)},
        "LOG_T": {"plan": "-", "actual": downtime_choices_day.get('log_t', 0), "current": downtime_for_current_item.get('log_t', 0)},
    }
    
    table_rows = []
    kpi_list = list(kpi_data_dict.items())
    for (kpi, item) in zip_longest(kpi_list, detail_jobs_list, fillvalue=None):
        table_rows.append({
            'kpi_desc': kpi[0] if kpi else None,
            'kpi_data': kpi[1] if kpi else {},
            'item_data': item,
        })
        
    context = {
        'line_name': formatted_line_name,
        'today': today,
        'table_rows': table_rows,
    }
    
    return render(request, "dashboard/dashboard_detail.html", context)

#input harian
@login_required
def input_harian(request):
    # Ambil tanggal dari GET, defaultnya None jika tidak ada ATAU kosong
    selected_date_str = request.GET.get('tanggal') 

    # --- PERBAIKAN DI SINI ---
    # Jika tanggal kosong atau tidak ada, gunakan default hari ini
    if not selected_date_str: # Ini memeriksa string kosong '' atau None
        selected_date = timezone.now().date()
        selected_date_str = selected_date.strftime('%Y-%m-%d')
    else:
        # Jika ada tanggal, coba parse. Jika gagal, baru fallback ke hari ini
        try:
            # Gunakan datetime.datetime.strptime di sini
            selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError): 
            selected_date = timezone.now().date()
            selected_date_str = selected_date.strftime('%Y-%m-%d')
            messages.warning(request, "Format tanggal tidak valid atau kosong, menampilkan data hari ini.")
    # --- AKHIR PERBAIKAN ---

    selected_filter = request.GET.get('line') # Ini ID line

    if request.method == 'POST':
        # Tidak perlu message success di sini jika update via AJAX atau form terpisah
        return redirect(request.get_full_path())

    # Ambil semua line UNIK untuk dropdown filter (nama + shift)
    all_lines = models.productionline.objects.all().order_by('namaline', 'shift')

    detail_jobs_on_date = models.detailjob.objects.filter(id_job__date=selected_date).order_by(
        'id_job__id_productionline__namaline',
        'id_job__id_productionline__shift',
        'urutan' # Tambahkan urutan
    ).select_related('id_itemproduksi', 'id_job__id_productionline') # Optimasi query

    selected_line_name = None
    selected_shift = None
    selected_line_obj = None # Tambahkan ini untuk menyimpan objek line

    if selected_filter:
        try:
            # Cari berdasarkan ID
            selected_line_obj = models.productionline.objects.get(id=selected_filter)
            selected_line_name = selected_line_obj.namaline
            selected_shift = selected_line_obj.shift

            # Filter detail job berdasarkan objek line yang ditemukan
            detail_jobs_on_date = detail_jobs_on_date.filter(
                id_job__id_productionline=selected_line_obj
            )
        except (models.productionline.DoesNotExist, ValueError):
            messages.error(request, "Filter Line tidak valid.")
            # Reset filter jika tidak valid
            selected_filter = None
            selected_line_obj = None
            selected_line_name = None
            selected_shift = None
            # Tampilkan semua job untuk tanggal tersebut jika filter error
            detail_jobs_on_date = models.detailjob.objects.filter(id_job__date=selected_date).order_by(
                'id_job__id_productionline__namaline',
                'id_job__id_productionline__shift',
                'urutan'
            ).select_related('id_itemproduksi', 'id_job__id_productionline')


    context = {
        'all_lines': all_lines, # Untuk dropdown filter
        'selected_date': selected_date,
        'selected_date_str': selected_date_str,
        'detail_jobs_on_date': detail_jobs_on_date, # Data utama
        'selected_filter': selected_filter, # Kirim ID line yang dipilih
        'selected_line_name': selected_line_name, # Kirim nama untuk tampilan
        'selected_shift': selected_shift, # Kirim shift untuk tampilan
    }
    return render(request, 'inputharian/input_harian.html', context)

@login_required 
@require_POST
def update_single_item(request, id):
    if request.method == 'POST':
        job_to_update = get_object_or_404(models.detailjob, id_detailjob=id)
        
        actual_qty = request.POST.get('actual_qty')
        repair_staging = request.POST.get('repair_staging')
        reject_staging = request.POST.get('reject_staging')

        if actual_qty is not None and actual_qty.isdigit():
            job_to_update.actual_qty = int(actual_qty)
        
        if repair_staging is not None and repair_staging.isdigit():
            job_to_update.repair_staging = int(repair_staging)

        if reject_staging is not None and reject_staging.isdigit():
            job_to_update.reject_staging = int(reject_staging)
        
        job_to_update.save()
        
        messages.success(request, f"Item {job_to_update.id_itemproduksi.job_number} berhasil diperbarui.")

    return redirect(request.META.get('HTTP_REFERER', 'input_harian'))

# CRUD Karyawan ---------------------------------------------
@login_required
def karyawan(request):
    karyawanobj = models.karyawan.objects.all()
    return render(request, 'karyawan/karyawan.html',{
        'karyawanobj' : karyawanobj
    })

@login_required
def create_karyawan(request):
    if request.method == 'GET':
        return render(request, 'karyawan/create_karyawan.html')
    else:
        nama_karyawan = request.POST['nama_karyawan']
        nrp_karyawan = request.POST['nrp_karyawan'] 
        jabatan = request.POST['jabatan'] 
        models.karyawan(
            nama_karyawan = nama_karyawan,
            nrp_karyawan = nrp_karyawan,
            jabatan = jabatan
        ).save()
        return redirect('karyawan')
    
@login_required
def update_karyawan(request, id):
    karyawanobj = models.karyawan.objects.get(id_karyawan = id)
    if request.method == 'GET':
        return render(request, 'karyawan/update_karyawan.html', {
            'karyawanobj': karyawanobj
        })
    else:
        karyawanobj.nama_karyawan  = request.POST['nama_karyawan']
        karyawanobj.nrp_karyawan = request.POST['nrp_karyawan'] 
        karyawanobj.jabatan = request.POST['jabatan'] 
        karyawanobj.save()
        return redirect('karyawan')

@login_required    
def delete_karyawan(request, id):
    karyawanobj = models.karyawan.objects.get(id_karyawan = id)
    karyawanobj.delete()
    return redirect('karyawan')

# Production Line--------------------------------------------
@login_required
def productionline(request): 
    productionlineobj = models.productionline.objects.all()
    return render(request, 'productionline/productionline.html',{
        'productionlineobj' : productionlineobj
    })

@login_required
def create_productionline(request):
    if request.method == 'POST':
        namaline = request.POST.get('namaline')
        shift = request.POST.get('shift')
        kapasitasline = request.POST.get('kapasitasline')

        # Validasi untuk memastikan semua field diisi
        if not all([namaline, shift, kapasitasline]):
            messages.error(request, "Semua field wajib diisi.")
            return redirect('create_productionline')

        # Validasi duplikat (tidak membedakan huruf besar/kecil)
        if models.productionline.objects.filter(namaline__iexact=namaline, shift=shift).exists():
            messages.error(request, f"Production Line '{namaline.title()} - Shift {shift}' sudah ada.")
            return redirect('create_productionline')

        # Jika aman, simpan data baru
        models.productionline.objects.create(
            namaline=namaline,
            shift=shift,
            kapasitasline=kapasitasline
        )
        return redirect('productionline')
    
    else: # method GET
        return render(request, 'productionline/create_productionline.html')

@login_required
def update_productionline(request, id):
    line_to_update = models.productionline.objects.get(id=id)

    if request.method == 'POST':
        # Ambil data baru dari form
        line_to_update.namaline = request.POST.get('namaline')
        line_to_update.shift = request.POST.get('shift')
        line_to_update.kapasitasline = request.POST.get('kapasitasline')
        line_to_update.save() # Simpan perubahan
        return redirect('productionline')

    else: 
        context = {
            'line': line_to_update
        }
        return render(request, 'productionline/update_productionline.html', context)

@login_required
def delete_productionline(request, id):
    productionlineobj = models.productionline.objects.get(id=id)
    productionlineobj.delete()
    return redirect("productionline")

@login_required
def rekap_productionline(request,id):
    dataproductionline= models.productionline.objects.get(id_productionline = id)
    filterjob = models.job.objects.filter(id_productionline = dataproductionline)
    return render(request,'productionline/rekap_productionline.html',{
        'dataproductionline' : dataproductionline,
        'filterjob': filterjob,
        'total_job' : filterjob.count()
    })

# CRUD Item Produksi-----------------------------
@login_required
def itemproduksi(request):
    itemproduksiobj = models.itemproduksi.objects.prefetch_related('production_lines').all()
    
    return render(request, 'itemproduksi/itemproduksi.html', {
        'itemproduksiobj': itemproduksiobj
    })

# Fungsi untuk membuat Item Produksi baru
@login_required
def create_itemproduksi(request):
    if request.method == 'POST':
        job_number = request.POST.get('job_number')
        part_number = request.POST.get('part_number')
        customer = request.POST.get('customer')
        cycle_time = request.POST.get('cycle_time') or 0.0
        
        production_line_ids = request.POST.getlist('production_lines')
        valid_ids = [pid for pid in production_line_ids if pid]

        if not valid_ids:
            messages.error(request, "Anda harus memilih minimal satu Production Line.")
            # Kirim data kembali jika ada error
            all_lines = models.productionline.objects.all().order_by('namaline', 'shift')
            context = {'productionlineobj': all_lines, 'form_data': request.POST}
            return render(request, 'itemproduksi/create_itemproduksi.html', context)
        
        new_item = models.itemproduksi.objects.create(
            job_number=job_number,
            part_number=part_number,
            customer=customer,
            cycle_time=cycle_time
        )
        new_item.production_lines.set(valid_ids)
        
        messages.success(request, "Item Produksi berhasil ditambahkan.")
        return redirect('itemproduksi')

    else: # method GET
        # Ini bagian penting: Kirim daftar production line ke template
        all_lines = models.productionline.objects.all().order_by('namaline', 'shift')
        context = {
            'productionlineobj': all_lines
        }
        return render(request, 'itemproduksi/create_itemproduksi.html', context)

# Fungsi untuk mengupdate Item Produksi
@login_required
def update_itemproduksi(request, id):
    item_to_update = get_object_or_404(models.itemproduksi, id_itemproduksi=id)

    if request.method == 'POST':
        # --- DEBUGGING: Cetak data yang diterima ---
        print("--- Menerima data POST ---")
        print(request.POST)
        # -----------------------------------------

        item_to_update.job_number = request.POST.get('job_number')
        item_to_update.part_number = request.POST.get('part_number')
        item_to_update.customer = request.POST.get('customer')
        item_to_update.cycle_time = request.POST.get('cycle_time') or 0.0
        
        production_line_ids = request.POST.getlist('production_lines')
        valid_production_line_ids = [pid for pid in production_line_ids if pid]

        # --- DEBUGGING: Cetak ID yang valid ---
        print(f"ID Line yang valid: {valid_production_line_ids}")
        # ------------------------------------

        if not valid_production_line_ids:
            print("--- Validasi GAGAL: Tidak ada line yang dipilih ---") # DEBUG
            messages.error(request, "Anda harus memilih minimal satu Production Line.")
            
            all_lines = models.productionline.objects.all().order_by('namaline', 'shift')
            context = {'item': item_to_update, 'productionlineobj': all_lines}
            return render(request, 'itemproduksi/update_itemproduksi.html', context)

        print("--- Validasi BERHASIL: Menyimpan data... ---") # DEBUG
        item_to_update.save()
        item_to_update.production_lines.set(valid_production_line_ids)
        
        messages.success(request, "Item Produksi berhasil diupdate.")
        return redirect('itemproduksi')

    else: # method GET
        all_lines = models.productionline.objects.all().order_by('namaline', 'shift')
        context = {
            'item': item_to_update,
            'productionlineobj': all_lines
        }
        return render(request, 'itemproduksi/update_itemproduksi.html', context)

# Fungsi untuk menghapus Item Produksi
@login_required
def delete_itemproduksi(request, id):
    item_to_delete = get_object_or_404(models.itemproduksi, id_itemproduksi=id)
    item_to_delete.delete()
    messages.success(request, "Item Produksi berhasil dihapus.")
    return redirect("itemproduksi")

# Create Job + Production Plan + Detail Job
@login_required
def create_jobdetails(request):
    if request.method == "GET":
        dataproductionline = models.productionline.objects.all().order_by('namaline', 'shift')
        datakaryawan = models.karyawan.objects.all()
        machines_qs = models.machine.objects.all()
        
        dataitemproduksi_qs = models.itemproduksi.objects.annotate(id=F('id_itemproduksi'))
        
        items_list_for_json = list(dataitemproduksi_qs.values('id', 'part_number', 'cycle_time'))
        machines_list_for_json = list(machines_qs.values('id', 'code'))

        context = {
            'dataproductionline': dataproductionline, 
            'datakaryawan': datakaryawan,
            'machines': machines_qs, 
            'dataitemproduksi': dataitemproduksi_qs,
            'items_json_str': json.dumps(items_list_for_json),
            'machines_json_str': json.dumps(machines_list_for_json),
        }
        return render(request, 'job/create_job.html', context)
    
    else: 
        try:
            with transaction.atomic():
                id_productionline = request.POST.get('id_productionline')
                id_karyawan = request.POST.get('id_karyawan')
                date = request.POST['date']

                if not id_productionline or not id_karyawan or not date:
                    messages.error(request, "Production Line, Karyawan, dan Tanggal wajib diisi.")
                    return redirect('create_job')

                # ### PERUBAHAN DI SINI ###
                # 1. Ambil data persen dari form
                #plan_repair_perc = request.POST.get('plan_repair_percentage', 0.0)
                #plan_reject_perc = request.POST.get('plan_reject_percentage', 0.0)

                new_job = models.job.objects.create(
                    id_productionline_id=id_productionline,
                    id_karyawan_id=id_karyawan,
                    date=date,
                    # 2. Simpan data persen ke database
                )
                
                gsph_plan_val = request.POST.get('gsph_plan') or 0
                stroke_plan_val = request.POST.get('stroke_plan') or 0
                repair_plan_val = request.POST.get('repair_plan') or 0
                reject_plan_val = request.POST.get('reject_plan') or 0
                idle_plan_val = request.POST.get('idle_plan') or 0
                mp_val = request.POST.get('mp') or 0
                ot_plan_val = request.POST.get('ot_plan') or 0

                models.productionplan.objects.create(
                    id_job=new_job,
                    gsph_plan=gsph_plan_val,
                    stroke_plan=stroke_plan_val,
                    repair_plan=repair_plan_val,
                    reject_plan=reject_plan_val,
                    idle_plan=idle_plan_val,
                    mp=mp_val,
                    ot_plan=ot_plan_val
                )

                id_item_produksi_list = request.POST.getlist('id_item_produksi')
                plan_qty_list = request.POST.getlist('plan_qty')
                
                for i, id_item in enumerate(id_item_produksi_list):
                    item_produksi_obj = models.itemproduksi.objects.get(id_itemproduksi=id_item)
                    
                    new_detailjob = models.detailjob.objects.create(
                        id_job=new_job,
                        id_itemproduksi=item_produksi_obj,
                        plan_qty=plan_qty_list[i],
                        plan_ct=item_produksi_obj.cycle_time
                    )
                    
                    machine_ids = request.POST.getlist(f'machine_used_{i}')
                    machines_selected = models.machine.objects.filter(id__in=machine_ids)
                    new_detailjob.machine_used.set(machines_selected)
                
                #messages.success(request, "Job berhasil ditambahkan!")
                return redirect('job')

        except Exception as e:
            messages.error(request, f"Terjadi error: {e}")
            return redirect('create_job')
    
# RUD Job
@login_required
def job(request):
    # Ambil tanggal dari filter, jika tidak ada, default ke hari ini
    selected_date_str = request.GET.get('tanggal', timezone.now().strftime('%Y-%m-%d'))
    selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()

    # Filter job berdasarkan tanggal yang dipilih
    filterjob = models.job.objects.filter(date=selected_date).order_by('id_productionline__namaline')

    context = {
        'filterjob': filterjob,
        'selected_date_str': selected_date_str
    }
    return render(request, 'job/job.html', context)

@login_required
def update_job(request, id):
    jobobj = models.job.objects.get(id_job = id)
    if request.method == 'GET':
        dataproductionline = models.productionline.objects.all()
        datakaryawan = models.karyawan.objects.all()
        tanggal = jobobj.date.strftime('%Y-%m-%d')
        return render(request, 'job/update_job.html', {
            'jobobj' : jobobj,
            'dataproductionline' : dataproductionline,
            'datakaryawan' : datakaryawan,
            'tanggal' : tanggal
        })
    else:
        id_karyawan = request.POST.get('id_karyawan')
        id_productionline = request.POST.get('id_productionline')
    
    if not id_productionline or not id_karyawan:
        messages.error(request, "Production Line dan Karyawan tidak boleh kosong.")
        # Redirect kembali ke halaman edit dengan data yang sudah ada
        return redirect('update_job', id=jobobj.id_job)

    jobobj.id_karyawan = models.karyawan.objects.get(id_karyawan = id_karyawan)
    jobobj.id_productionline = models.productionline.objects.get(id = id_productionline)
    jobobj.date = request.POST.get('date')
    jobobj.save()
    messages.success(request, "Job berhasil diupdate!")
    return redirect('job')

@login_required
def delete_job(request, id):
    jobobj = models.job.objects.get(id_job = id)
    jobobj.delete()
    return redirect('job')

# RU Production Plan 
@login_required
def rekap_productionplan(request,id):
    datajob= models.job.objects.get(id_job = id)
    filterpl = models.productionplan.objects.get(id_job= datajob)
    return render(request,'productionplan/rekap_productionplan.html',{
        'datajob' : datajob,
        'filterpl': filterpl
    })

@login_required
def update_productionplan(request, id):
    productionplanobj = models.productionplan.objects.get(id_production_plan=id)
    if request.method == 'POST':
        productionplanobj.gsph_plan = request.POST.get('gsph_plan', 0)
        productionplanobj.stroke_plan = request.POST.get('stroke_plan', 0)
        productionplanobj.repair_plan = request.POST.get('repair_plan', 0)
        productionplanobj.reject_plan = request.POST.get('reject_plan', 0)
        productionplanobj.idle_plan = request.POST.get('idle_plan', 0)
        productionplanobj.mp = request.POST.get('mp', 0)
        productionplanobj.ot_plan = request.POST.get('ot_plan', 0)
        
        productionplanobj.gsph_actual = request.POST.get('gsph_actual', 0)
        productionplanobj.stroke_actual = request.POST.get('stroke_actual', 0)
        productionplanobj.ot_actual = request.POST.get('ot_actual', 0)
        
        productionplanobj.save()
        messages.success(request, "Production Plan berhasil diupdate.")
        
        job_id = productionplanobj.id_job.id_job
        return redirect('rekap_productionplan', id=job_id)

    context = { 'productionplanobj': productionplanobj }
    return render(request, 'productionplan/update_productionplan.html', context)


# CRUD Detail Job
@login_required
def rekap_detailjob(request,id):
    datajob = models.job.objects.get(id_job=id)
    filterdetailj = models.detailjob.objects.filter(id_job=datajob)
    return render(request, 'detailjob/rekap_detailjob.html', {
        'datajob': datajob,
        'filterdetailj': filterdetailj
    })

@login_required
def create_detailjob(request, id):
    getjob = models.job.objects.get(id_job=id)
    dataitemproduksi = models.itemproduksi.objects.all()
    machines = models.machine.objects.all()
    if request.method == "GET":
        return render(request, 'detailjob/create_detailjob.html', {
            'dataitemproduksi': dataitemproduksi,
            'getjob': getjob,
            'machines': machines,
        })
    else:
        id_item_produksi_list = request.POST.getlist('id_item_produksi')
        plan_qty_list = request.POST.getlist('plan_qty')
        plan_CT_list = request.POST.getlist('plan_CT')
        for i, (id_item, plan_qty, plan_ct) in enumerate(zip(id_item_produksi_list, plan_qty_list, plan_CT_list)):
            item = models.itemproduksi.objects.get(id_itemproduksi=id_item)
            new_detailjob = models.detailjob.objects.create(
                id_job=getjob,
                id_itemproduksi=item,
                plan_qty=plan_qty,
                plan_ct=plan_ct
            )
            machine_ids = request.POST.getlist(f'machine_used_{i}')
            machines_selected = models.machine.objects.filter(id__in=machine_ids)
            new_detailjob.machine_used.set(machines_selected)
        messages.success(request, f"Detail job berhasil ditambahkan untuk Job {id}!")
        return redirect("rekap_detailjob", id=id)

@login_required
def update_detailjob(request, id):
    detail_job_to_update = get_object_or_404(models.detailjob, id_detailjob=id)

    if request.method == 'POST':
        item_produksi_id = request.POST.get('id_itemproduksi')
        machine_ids = request.POST.getlist('machine_used')

        detail_job_to_update.id_itemproduksi = get_object_or_404(
            models.itemproduksi, id_itemproduksi=item_produksi_id
        )

        detail_job_to_update.plan_qty = request.POST.get('plan_qty')
        detail_job_to_update.urutan = request.POST.get('urutan')

        # Plan CT selalu ikut cycle_time dari item produksi
        detail_job_to_update.plan_ct = detail_job_to_update.id_itemproduksi.cycle_time

        detail_job_to_update.save()
        detail_job_to_update.machine_used.set(models.machine.objects.filter(id__in=machine_ids))

        messages.success(request, "Detail Job berhasil diupdate!")
        return redirect("rekap_detailjob", id=detail_job_to_update.id_job.id_job)

    else:
        all_items = models.itemproduksi.objects.all()
        all_machines = models.machine.objects.all()

        context = {
            "detail_job": detail_job_to_update,
            "all_items": all_items,
            "all_machines": all_machines,
        }
        return render(request, "detailjob/update_detailjob.html", context)

@login_required
def delete_detailjob(request, id):
    getdetailj = models.detailjob.objects.get(id_detailjob = id)
    job_id = getdetailj.id_job.id_job
    getdetailj.delete()
    return redirect("rekap_detailjob", id=job_id)

@login_required
@require_POST
def update_actual_qty(request):
    if request.method == 'POST':
        # Kumpulkan semua ID detailjob yang unik dari form yang di-submit
        detail_job_ids = set()
        for key in request.POST:
            if key.startswith('actual_qty_'):
                detail_job_ids.add(key.split('_')[-1])

        # Loop melalui setiap ID yang ditemukan dan update datanya
        for job_id in detail_job_ids:
            try:
                job_to_update = models.detailjob.objects.get(id_detailjob=job_id)
                
                # Ambil nilai 'actual_qty', 'repair_staging', dan 'reject_staging'
                actual_qty_val = request.POST.get(f'actual_qty_{job_id}')
                repair_staging_val = request.POST.get(f'repair_staging_{job_id}')
                reject_staging_val = request.POST.get(f'reject_staging_{job_id}')

                # Update field 'actual_qty'
                if actual_qty_val and actual_qty_val.isdigit():
                    job_to_update.actual_qty = int(actual_qty_val)
                
                # Update field 'repair_staging'
                if repair_staging_val and repair_staging_val.isdigit():
                    job_to_update.repair_staging = int(repair_staging_val)

                # Update field 'reject_staging'
                if reject_staging_val and reject_staging_val.isdigit():
                    job_to_update.reject_staging = int(reject_staging_val)

                # Simpan semua perubahan ke database
                job_to_update.save()
                

            except (models.detailjob.DoesNotExist, ValueError):
                pass
    
    # Arahkan kembali pengguna ke halaman sebelumnya
    return redirect(request.META.get('HTTP_REFERER', 'input_harian'))

@login_required
@require_POST
def update_actual_qty_single(request, id):
    # Dapatkan objek detailjob yang mau diupdate
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    job_id = detail_job.id_job.id_job

    if request.method == "POST":
        try:
            new_qty = request.POST.get('actual_qty')
            if new_qty is not None and new_qty.isdigit():
                detail_job.actual_qty = int(new_qty)
                detail_job.save()
                messages.success(request, f"Actual Qty berhasil diupdate.")
            else:
                messages.error(request, "Input tidak valid.")
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            
    # Arahkan kembali ke halaman rekap detail job
    return redirect('rekap_detailjob', id=job_id)

# CRUD Downtime
@login_required
def pilih_job_untuk_downtime(request):
    if request.method == 'POST':
        # Langkah 2: Proses pilihan pengguna
        detailjob_id = request.POST.get('detailjob_id')
        if detailjob_id:
            # Arahkan ke halaman rekap downtime untuk job yang dipilih
            return redirect('rekap_downtime', id=detailjob_id)
        # Jika tidak ada pilihan, kembali ke halaman ini
        return redirect('pilih_job_downtime')

    # Langkah 1: Tampilkan halaman pilihan
    selected_date_str = request.GET.get('tanggal', timezone.now().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        selected_date = timezone.now().date()
        selected_date_str = selected_date.strftime('%Y-%m-%d')
    
    # Ambil semua detail job yang ada pada tanggal yang dipilih
    jobs_on_date = models.detailjob.objects.filter(id_job__date=selected_date).order_by('id_job__id_productionline')

    context = {
        'semua_detailjob': jobs_on_date,
        'selected_date_str': selected_date_str,
    }
    return render(request, 'downtime/pilih_job_downtime.html', context)

@login_required
def rekap_downtime(request, id):
    # Gunakan get_object_or_404 untuk keamanan
    datadetailjob = get_object_or_404(models.detailjob, id_detailjob=id)
    
    # Ambil semua detail downtime yang relevan
    filterdt = models.detail_dt.objects.filter(
        id_downtime__id_detailjob=datadetailjob
    ).order_by('-start_downtime')

    # Ambil pilihan jenis downtime untuk dropdown di form
    jenisdowntime_choices = models.detail_dt._meta.get_field('jenisdowntime').choices
    
    # --- Logika untuk Tabel Rekapitulasi ---
    classified_dt = {name: {'count': 0, 'total_minutes': 0} for code, name in jenisdowntime_choices}
    
    for dt in filterdt:
        jenis_nama = dt.get_jenisdowntime_display()
        if jenis_nama in classified_dt:
            classified_dt[jenis_nama]['count'] += 1
            classified_dt[jenis_nama]['total_minutes'] += dt.duration_downtime or 0

    total_all_minutes = sum(d['total_minutes'] for d in classified_dt.values())
    total_all_count = sum(d['count'] for d in classified_dt.values())

    context = {
        'datadetailjob': datadetailjob,
        'classified_dt': classified_dt,
        'filterdt': filterdt,
        'jenisdowntime_choices': jenisdowntime_choices,
        'total_all_minutes': total_all_minutes,
        'total_all_count': total_all_count,

    }
    
    return render(request, 'downtime/rekap_downtime.html', context)

@login_required
def downtime(request):
    # Ambil tanggal dari filter, jika tidak ada, default ke hari ini
    selected_date_str = request.GET.get('tanggal', timezone.now().strftime('%Y-%m-%d'))
    selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()

    start_of_day = timezone.make_aware(datetime.datetime.combine(selected_date, datetime.time.min))
    end_of_day = timezone.make_aware(datetime.datetime.combine(selected_date, datetime.time.max))

    # Filter data downtime berdasarkan RENTANG WAKTU
    filterdt = models.detail_dt.objects.filter(
        start_downtime__range=(start_of_day, end_of_day)
    ).order_by('-start_downtime')

    context = {
        'filterdt': filterdt,
        'selected_date_str': selected_date_str
    }
    return render(request, 'downtime/downtime.html', context)

@require_POST
@login_required
def create_downtime(request):
    try:
        data = json.loads(request.body)
        detail_job_id = data.get('id_detailjob')
        detail_job_obj = get_object_or_404(models.detailjob, id_detailjob=detail_job_id)

        downtime_header, created = models.downtime.objects.get_or_create(id_detailjob=detail_job_obj)

        def parse_utc_to_local(val):
            if not val: return None
            if val.endswith('Z'):
                val = val.replace('Z', '+00:00')
            dt_utc = datetime.datetime.fromisoformat(val)
            return dt_utc.astimezone(timezone.get_current_timezone())

        start_dt = parse_utc_to_local(data.get('start_downtime')) or timezone.now()
        finish_dt = parse_utc_to_local(data.get('finish_downtime'))

        # Simpan objek yang baru dibuat ke dalam variabel
        new_downtime_obj = models.detail_dt.objects.create(
            id_downtime=downtime_header,
            jenisdowntime=data.get('jenisdowntime'),
            problem_dt=data.get('problem_dt'),
            penyebab_dt=data.get('penyebab_dt'),
            action_dt=data.get('action_dt'),
            pic_downtime=data.get('pic_downtime'),
            start_downtime=start_dt,
            finish_downtime=finish_dt
        )

        # Siapkan data untuk dikirim kembali sebagai JSON
        new_dt_data = {
            'id': new_downtime_obj.id_detaildt,
            'jenis': new_downtime_obj.get_jenisdowntime_display(),
            'problem_dt': new_downtime_obj.problem_dt or "-",
            'penyebab_dt': new_downtime_obj.penyebab_dt or "-",
            'action_dt': new_downtime_obj.action_dt or "-",
            'pic_downtime': new_downtime_obj.pic_downtime or "-",
            'start': new_downtime_obj.start_downtime.strftime('%d %b, %H:%M'),
            'finish': new_downtime_obj.finish_downtime.strftime('%d %b, %H:%M') if new_downtime_obj.finish_downtime else '-',
            'durasi': f"{new_downtime_obj.duration_downtime:.2f}" if new_downtime_obj.duration_downtime is not None else '-',
        }

        return JsonResponse({
            'status': 'success', 
            'message': 'Downtime berhasil disimpan.',
            'new_downtime': new_dt_data
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_POST        
def stop_downtime(request, id):
    dt = models.detail_dt.objects.get(id_detaildt=id)
    dt.finish_downtime = timezone.now()
    dt.save()
    return redirect('rekap_downtime', id=dt.id_downtime.id_detailjob.id_detailjob)

@login_required
@require_POST
def delete_downtime(request, id):
    get_dt = models.detail_dt.objects.get(id_detaildt=id)
    job_id = get_dt.id_downtime.id_detailjob.id_detailjob
    get_dt.delete()
    return redirect('rekap_downtime', id=job_id)

@login_required
def update_downtime(request, id):
    get_detaildt = models.detail_dt.objects.get(id_detaildt=id)
    jenisdowntime_choices = models.detail_dt._meta.get_field('jenisdowntime').choices

    if request.method == 'POST':
        get_detaildt.jenisdowntime = request.POST.get('jenisdowntime')
        get_detaildt.problem_dt = request.POST.get('problem_dt')
        get_detaildt.penyebab_dt = request.POST.get('penyebab_dt')
        get_detaildt.action_dt = request.POST.get('action_dt')
        get_detaildt.pic_downtime = request.POST.get('pic_downtime')
        get_detaildt.save()
        return redirect('rekap_downtime', id=get_detaildt.id_downtime.id_detailjob.id_detailjob)

    return render(request, 'downtime/update_downtime.html', {
        'get_detaildt': get_detaildt,
        'jenisdowntime_choices': jenisdowntime_choices
    })

@login_required
def dandori_main(request):
    all_lines_from_db = models.productionline.objects.all().order_by('namaline', 'shift')
    
    top_dropdown_lines = []
    seen_line_names = set()
    for line in all_lines_from_db:
        if line.namaline not in seen_line_names:
            top_dropdown_lines.append(line)
            seen_line_names.add(line.namaline)

    history_dropdown_lines = []
    seen_combinations = set()
    for line in all_lines_from_db:
        identifier = (line.namaline, line.shift)
        if identifier not in seen_combinations:
            history_dropdown_lines.append(line)
            seen_combinations.add(identifier)
    
    selected_date_str = request.GET.get('history_date', timezone.now().strftime('%Y-%m-%d'))
    selected_line_id = request.GET.get('history_line')
    
    dandori_history = models.dandori.objects.select_related(
        'id_detailjob__id_job__id_productionline',
        'id_detailjob__id_itemproduksi'
    ).all()

    # Filter berdasarkan tanggal
    try:
        selected_date = timezone.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        dandori_history = dandori_history.filter(id_detailjob__id_job__date=selected_date)
    except (ValueError, TypeError):
        today = timezone.now().date()
        dandori_history = dandori_history.filter(id_detailjob__id_job__date=today)
        selected_date_str = today.strftime('%Y-%m-%d')
        
    # Filter berdasarkan Line/Shift yang dipilih dari dropdown
    if selected_line_id:
        try:
            selected_line_obj = models.productionline.objects.get(id=selected_line_id)
            dandori_history = dandori_history.filter(
                id_detailjob__id_job__id_productionline__namaline=selected_line_obj.namaline,
                id_detailjob__id_job__id_productionline__shift=selected_line_obj.shift
            )
        except models.productionline.DoesNotExist:
            pass

    # Hitung total durasi
    total_duration_history = sum(dd.duration for dd in dandori_history if dd.duration)


    context = {
        'top_dropdown_lines': top_dropdown_lines,
        'history_dropdown_lines': history_dropdown_lines,
        
        'dandori_history': dandori_history.order_by('-start_time'),
        'total_duration_history': total_duration_history,
        
        'selected_history_date': selected_date_str,
        'selected_history_line': int(selected_line_id) if selected_line_id else None,
    }
    
    return render(request, 'dandori/dandori_main.html', context)

@login_required
def list_dandori(request, id):
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    existing_dandoris = models.dandori.objects.filter(id_detailjob=detail_job)
    dandori_status = []
    for type_code, type_display in models.dandori.DANDORI_TYPES:
        record = existing_dandoris.filter(jenis_dandori=type_code).first()
        dandori_status.append({
            'type_code': type_code,
            'type_display': type_display,
            'record': record,
        })
    total_duration = sum(d.duration for d in existing_dandoris if d.duration)
    context = {
        'detail_job': detail_job,
        'dandori_status': dandori_status,
        'total_duration': total_duration,
    }
    return render(request, 'dandori/list_dandori.html', context)

@login_required
@require_POST
def start_dandori(request, id, dandori_type):
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    existing_dandori = models.dandori.objects.filter(
        id_detailjob=detail_job, 
        jenis_dandori=dandori_type, 
        finish_time__isnull=True
    ).exists()
    if not existing_dandori:
        models.dandori.objects.create(
            id_detailjob=detail_job,
            jenis_dandori=dandori_type,
            start_time=timezone.now()
        )
        messages.success(request, f"Timer untuk '{dandori_type}' telah dimulai.")
    else:
        messages.warning(request, f"Timer untuk '{dandori_type}' sudah berjalan.")
    return redirect('list_dandori', id=id)

@login_required
@require_POST
def stop_dandori(request, id):
    dandori_record = get_object_or_404(models.dandori, id_dandori=id)
    dandori_record.finish_time = timezone.now()
    dandori_record.save()
    messages.info(request, "Timer telah dihentikan.")
    return redirect('list_dandori', id=dandori_record.id_detailjob.id_detailjob)

@login_required
@require_POST
def restart_dandori(request, id):
    dandori_record = get_object_or_404(models.dandori, id_dandori=id)
    detail_job_id = dandori_record.id_detailjob.id_detailjob
    dandori_record.delete()
    messages.info(request, "Timer telah di-reset. Anda bisa memulainya kembali.")
    return redirect('list_dandori', id=detail_job_id)

# CRUD IdleTime
@login_required
def idletime_main(request):
    all_lines_from_db = models.productionline.objects.all().order_by('namaline', 'shift')
    
    top_dropdown_lines = []
    seen_line_names = set()
    for line in all_lines_from_db:
        if line.namaline not in seen_line_names:
            top_dropdown_lines.append(line)
            seen_line_names.add(line.namaline)

    history_dropdown_lines = []
    seen_combinations = set()
    for line in all_lines_from_db:
        identifier = (line.namaline, line.shift)
        if identifier not in seen_combinations:
            history_dropdown_lines.append(line)
            seen_combinations.add(identifier)
    
    # Logika filter
    selected_date_str = request.GET.get('history_date', timezone.now().strftime('%Y-%m-%d'))
    selected_line_id = request.GET.get('history_line')

    idletime_history = models.detail_idle.objects.select_related(
        'id_idle__id_detailjob__id_itemproduksi',
        'id_idle__id_detailjob__id_job__id_productionline'
    )

    # Filter berdasarkan tanggal
    try:
        selected_date = timezone.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        # REVISI: Filter berdasarkan tanggal di detail_job, bukan start_idle
        idletime_history = idletime_history.filter(id_idle__id_detailjob__id_job__date=selected_date)
    except (ValueError, TypeError):
        today = timezone.now().date()
        idletime_history = idletime_history.filter(id_idle__id_detailjob__id_job__date=today)
        selected_date_str = today.strftime('%Y-%m-%d')

    # Filter berdasarkan Line/Shift yang dipilih dari dropdown
    if selected_line_id:
        try:
            selected_line_obj = models.productionline.objects.get(id=selected_line_id)
            idletime_history = idletime_history.filter(
                id_idle__id_detailjob__id_job__id_productionline__namaline=selected_line_obj.namaline,
                id_idle__id_detailjob__id_job__id_productionline__shift=selected_line_obj.shift
            )
        except models.productionline.DoesNotExist:
            pass

    total_duration_history = sum(item.duration_idle for item in idletime_history if item.duration_idle)

    context = {
        # Kirim DUA daftar berbeda ke template
        'top_dropdown_lines': top_dropdown_lines,
        'history_dropdown_lines': history_dropdown_lines,
        
        # Data utama yang akan ditampilkan
        'idletime_history': idletime_history.order_by('-start_idle'),
        'total_duration_history': total_duration_history,
        
        # Variabel untuk menjaga state filter di form
        'selected_history_date': selected_date_str,
        'selected_history_line': int(selected_line_id) if selected_line_id else None,
    }
    return render(request, 'idletime/idletime_main.html', context)

@login_required
def rekap_idletime(request, id):
    datadetailjob = models.detailjob.objects.get(id_detailjob=id)
    idle, created = models.idletime.objects.get_or_create(id_detailjob=datadetailjob)
    filteridle = models.detail_idle.objects.filter(id_idle=idle)
    total_idle = sum(d.duration_idle for d in filteridle)
    count_idle = filteridle.count()

    return render(request, "idletime/rekap_idletime.html", {
        "datadetailjob": datadetailjob,
        "filteridle": filteridle,
        "total_idle": total_idle,
        "count_idle": count_idle,
    })

@login_required
@require_POST
def create_idletime(request, id):
    datadetailjob = models.detailjob.objects.get(id_detailjob=id)
    idle, created = models.idletime.objects.get_or_create(id_detailjob=datadetailjob)

    if request.method == "POST":
        start_idle_str = request.POST.get("start_idle")
        finish_idle_str = request.POST.get("finish_idle")
        reason_idle = request.POST.get("reason_idle")

        start_idle = parse_datetime(start_idle_str)
        finish_idle = parse_datetime(finish_idle_str) if finish_idle_str else None

        models.detail_idle.objects.create(
            id_idle=idle,
            start_idle=start_idle,
            finish_idle=finish_idle,
            reason_idle=reason_idle
        )
    return redirect("rekap_idletime", id=id)

@login_required
@require_POST
def stop_idletime(request, id):
    idle_detail = models.detail_idle.objects.get(id_detaild = id)
    if not idle_detail.finish_idle:
        idle_detail.finish_idle = timezone.now()
        idle_detail.save()
    return redirect("rekap_idletime", id=idle_detail.id_idle.id_detailjob.id_detailjob)

@login_required
@require_POST
def delete_idletime(request, id):
    idle_detail = models.detail_idle.objects.get(id_detaild=id)
    id_detailjob = idle_detail.id_idle.id_detailjob.id_detailjob
    idle_detail.delete()
    return redirect("rekap_idletime", id=id_detailjob)

@login_required
def update_idletime(request, id):
    getidle = models.detail_idle.objects.get(id_detaild=id)
    reason = request.POST.get('reason_idle')
    if reason is not None:
        getidle.reason_idle = reason
        getidle.save()
        return redirect("rekap_idletime", id=getidle.id_idle.id_detailjob.id_detailjob)

    return render(request, 'idletime/update_idletime.html', {
        'getidle': getidle
    })

@login_required
def break_time_list(request):
    """Menampilkan semua jadwal istirahat."""
    semua_break = models.BreakTime.objects.all().order_by('shift', 'hari', 'waktu_mulai')
    context = {'semua_break': semua_break}
    return render(request, 'breaktime/breaktime_list.html', context)

@login_required
def break_time_create(request):
    """Membuat jadwal istirahat baru."""
    if request.method == 'POST':
        # Proses form submission
        nama = request.POST.get('nama_istirahat')
        mulai = request.POST.get('waktu_mulai')
        selesai = request.POST.get('waktu_selesai')
        shift = request.POST.get('shift')
        hari = request.POST.get('hari')
        
        # Jika 'hari' tidak dipilih (untuk setiap hari), simpan sebagai None
        if not hari:
            hari = None
            
        models.BreakTime.objects.create(
            nama_istirahat=nama,
            waktu_mulai=mulai,
            waktu_selesai=selesai,
            shift=shift,
            hari=hari
        )
        return redirect('break_time_list')
    
    # Tampilkan form kosong
    context = {'choices_hari': models.BreakTime.Hari.choices}
    return render(request, 'breaktime/breaktime_form.html', context)

@login_required
def break_time_update(request, id):
    """Mengupdate jadwal istirahat yang sudah ada."""
    break_obj = get_object_or_404(models.BreakTime, id=id)
    if request.method == 'POST':
        break_obj.nama_istirahat = request.POST.get('nama_istirahat')
        break_obj.waktu_mulai = request.POST.get('waktu_mulai')
        break_obj.waktu_selesai = request.POST.get('waktu_selesai')
        break_obj.shift = request.POST.get('shift')
        
        hari = request.POST.get('hari')
        break_obj.hari = hari if hari else None
        
        break_obj.save()
        return redirect('break_time_list')
        
    context = {
        'break_obj': break_obj,
        'choices_hari': models.BreakTime.Hari.choices
    }
    return render(request, 'breaktime/breaktime_form.html', context)

@login_required
@require_POST
def break_time_delete(request, id):
    """Menghapus jadwal istirahat."""
    break_obj = get_object_or_404(models.BreakTime, id=id)
    break_obj.delete()
    return redirect('break_time_list')

def laporan(request):
    return render(request, "stamping/laporan.html")

# CRUD Handwork
@login_required
def handwork_main(request):
    all_lines_from_db = models.productionline.objects.all().order_by('namaline', 'shift')
    
    # A. Buat daftar unik untuk dropdown ATAS (Mulai Pencatatan)
    top_dropdown_lines = []
    seen_line_names = set()
    for line in all_lines_from_db:
        if line.namaline not in seen_line_names:
            top_dropdown_lines.append(line)
            seen_line_names.add(line.namaline)

    # B. Buat daftar unik untuk dropdown BAWAH (History)
    history_dropdown_lines = []
    seen_combinations = set()
    for line in all_lines_from_db:
        identifier = (line.namaline, line.shift)
        if identifier not in seen_combinations:
            history_dropdown_lines.append(line)
            seen_combinations.add(identifier)
    
    # Logika filter
    selected_date_str = request.GET.get('history_date', timezone.now().strftime('%Y-%m-%d'))
    selected_line_id = request.GET.get('history_line')

    handwork_history = models.detailhandwork.objects.select_related(
        'id_handwork__id_detailjob__id_itemproduksi',
        'id_handwork__id_detailjob__id_job__id_productionline'
    )

    # Filter berdasarkan tanggal
    try:
        selected_date = timezone.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        # REVISI: Filter berdasarkan tanggal job, bukan created_at agar lebih konsisten
        handwork_history = handwork_history.filter(id_handwork__id_detailjob__id_job__date=selected_date)
    except (ValueError, TypeError):
        today = timezone.now().date()
        handwork_history = handwork_history.filter(id_handwork__id_detailjob__id_job__date=today)
        selected_date_str = today.strftime('%Y-%m-%d')

    # Filter berdasarkan Line/Shift yang dipilih dari dropdown
    if selected_line_id:
        try:
            selected_line_obj = models.productionline.objects.get(id=selected_line_id)
            handwork_history = handwork_history.filter(
                id_handwork__id_detailjob__id_job__id_productionline__namaline=selected_line_obj.namaline,
                id_handwork__id_detailjob__id_job__id_productionline__shift=selected_line_obj.shift
            )
        except models.productionline.DoesNotExist:
            pass

    context = {
        # Kirim DUA daftar berbeda ke template
        'top_dropdown_lines': top_dropdown_lines,
        'history_dropdown_lines': history_dropdown_lines,
        
        # Data utama yang akan ditampilkan
        'handwork_history': handwork_history.order_by('-created_at'),
        
        # Variabel untuk menjaga state filter di form
        'selected_history_date': selected_date_str,
        'selected_history_line': int(selected_line_id) if selected_line_id else None,
    }
    return render(request, 'handwork/handwork_main.html', context)

@login_required
def rekap_handwork(request, id):
    """Menampilkan detail dan form input untuk satu job handwork spesifik."""
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    handwork_header, created = models.handwork.objects.get_or_create(id_detailjob=detail_job)

    if request.method == 'POST':
        problem_hw = request.POST.get('problem_hw')
        status = request.POST.get('status')
        quantity = int(request.POST.get('quantity', 1))
        foto_sebelum = request.FILES.get('foto_sebelum')
        foto_sesudah = request.FILES.get('foto_sesudah')

        is_ok = True if status == 'is_ok' else False

        for _ in range(quantity):
            models.detailhandwork.objects.create(
                id_handwork=handwork_header,
                problem_hw=problem_hw,
                is_ok=is_ok,
                foto_sebelum=foto_sebelum,
                foto_sesudah=foto_sesudah
            )
        
        detail_job.repair_staging -= quantity
        if detail_job.repair_staging < 0:
            detail_job.repair_staging = 0

        detail_job.save()
        
        return redirect('rekap_handwork', id=id)

    handwork_items = models.detailhandwork.objects.filter(id_handwork=handwork_header).order_by('-id_detailhw')
    
    context = {
        'detail_job': detail_job,
        'handwork_items': handwork_items,
        'total_ok': handwork_header.qty_ok,
        'total_reject': handwork_header.qty_reject,
    }
    return render(request, 'handwork/rekap_handwork.html', context)

@login_required
@require_POST
def delete_handwork_item(request, id):
    """Menghapus satu item handwork."""
    item = get_object_or_404(models.detailhandwork, id_detailhw=id)
    detail_job_id = item.id_handwork.id_detailjob.id_detailjob
    item.delete()
    messages.info(request, "Satu item handwork telah dihapus.")
    return redirect('rekap_handwork', id=detail_job_id)

# 1. View untuk Halaman Utama
@login_required
def qcheck_main(request):
    all_lines_from_db = models.productionline.objects.all().order_by('namaline', 'shift')
    
    # A. Buat daftar unik untuk dropdown ATAS (Mulai Pencatatan)
    top_dropdown_lines = []
    seen_line_names = set()
    for line in all_lines_from_db:
        if line.namaline not in seen_line_names:
            top_dropdown_lines.append(line)
            seen_line_names.add(line.namaline)

    # B. Buat daftar unik untuk dropdown BAWAH (History)
    history_dropdown_lines = []
    seen_combinations = set()
    for line in all_lines_from_db:
        identifier = (line.namaline, line.shift)
        if identifier not in seen_combinations:
            history_dropdown_lines.append(line)
            seen_combinations.add(identifier)
    
    # Logika filter
    selected_date_str = request.GET.get('history_date', timezone.now().strftime('%Y-%m-%d'))
    selected_line_id = request.GET.get('history_line')

    qcheck_history = models.qcheck.objects.select_related(
        'id_detailjob__id_itemproduksi',
        'id_detailjob__id_job__id_productionline'
    ).all()

    # Filter berdasarkan tanggal
    try:
        selected_date = timezone.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        qcheck_history = qcheck_history.filter(id_detailjob__id_job__date=selected_date)
    except (ValueError, TypeError):
        today = timezone.now().date()
        qcheck_history = qcheck_history.filter(id_detailjob__id_job__date=today)
        selected_date_str = today.strftime('%Y-%m-%d')

    # Filter berdasarkan Line/Shift yang dipilih dari dropdown
    if selected_line_id:
        try:
            selected_line_obj = models.productionline.objects.get(id=selected_line_id)
            qcheck_history = qcheck_history.filter(
                id_detailjob__id_job__id_productionline__namaline=selected_line_obj.namaline,
                id_detailjob__id_job__id_productionline__shift=selected_line_obj.shift
            )
        except models.productionline.DoesNotExist:
            pass

    total_duration_history = sum(qc.duration for qc in qcheck_history if qc.duration)
    
    context = {
        # Kirim DUA daftar berbeda ke template
        'top_dropdown_lines': top_dropdown_lines,
        'history_dropdown_lines': history_dropdown_lines,
        
        # Data utama yang akan ditampilkan
        'qcheck_history': qcheck_history.order_by('-start_time'),
        'total_duration_history': total_duration_history,
        
        # Variabel untuk menjaga state filter di form
        'selected_history_date': selected_date_str,
        'selected_history_line': int(selected_line_id) if selected_line_id else None,
    }
    return render(request, 'qcheck/qcheck_main.html', context)

# 2. View untuk Halaman "Pilih Item" (setelah filter Line & Shift)
@login_required
def qcheck_select_item(request):
    line_id = request.GET.get('line')
    shift = request.GET.get('shift')
    
    if not line_id or not shift:
        messages.error(request, "Harap pilih Line dan Shift.")
        return redirect('qcheck_main')

    # Cari job yang sesuai untuk hari ini
    job = models.job.objects.filter(
        date=timezone.now().date(),
        id_productionline=line_id 
    ).first()

    items = []
    if job:
        items = models.detailjob.objects.filter(id_job=job)

    selected_line = get_object_or_404(models.productionline, id=line_id)

    context = {
        'items': items,
        'selected_line': selected_line,
        'selected_shift': shift,
    }
    return render(request, 'qcheck/qcheck_select_item.html', context)


# 3. View untuk Halaman Start/Stop
@login_required
def list_qcheck(request, id):
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    
    qcheck_status = []
    total_duration = 0

    for code, display in models.qcheck.QC_TYPES:
        record = models.qcheck.objects.filter(id_detailjob=detail_job, jenis_qcheck=code).first()
        
        qcheck_status.append({
            'type_code': code,
            'type_display': display,
            'record': record,
        })
        if record and record.duration:
            total_duration += record.duration

    return render(request, "qcheck/list_qcheck.html", {
        "detail_job": detail_job,
        "qcheck_status": qcheck_status,
        "total_duration": total_duration,
    })

# 4. View untuk tombol "Start"
@login_required
@require_POST
def start_qcheck(request, id, qcheck_type):
    detail_job = get_object_or_404(models.detailjob, id_detailjob=id)
    
    existing_qc = models.qcheck.objects.filter(id_detailjob=detail_job, jenis_qcheck=qcheck_type).first()

    if not existing_qc:
        models.qcheck.objects.create(
            id_detailjob=detail_job,
            jenis_qcheck=qcheck_type,
            hasil_qcheck="OK",
            start_time=timezone.now()
        )
        messages.success(request, f"Timer untuk '{qcheck_type}' berhasil dimulai.")
    else:
        messages.warning(request, "Aktivitas ini sudah pernah dimulai.")
        
    return redirect('list_qcheck', id=id)

@login_required
@require_POST
def restart_qcheck(request, id):
    qc = get_object_or_404(models.qcheck, id_qcheck=id)
    detail_job_id = qc.id_detailjob.id_detailjob
    qc.delete()
    messages.info(request, "Timer telah di-reset. Anda bisa memulainya kembali.")
    return redirect('list_qcheck', id=detail_job_id)

# 5. View untuk tombol "Stop"
@login_required
@require_POST
def control_qcheck(request, id, action):
    qc = get_object_or_404(models.qcheck, id_qcheck=id)
    if action == 'stop' and not qc.finish_time:
        qc.finish_time = timezone.now()
        qc.save()
        messages.success(request, f"Q Check '{qc.get_jenis_qcheck_display()}' telah dihentikan.")
    return redirect('list_qcheck', id=qc.id_detailjob.id_detailjob)

# 6. View untuk Edit
@login_required
def edit_qcheck(request, id):
    qc = get_object_or_404(models.qcheck, id_qcheck=id)
    if request.method == "POST":
        qc.jenis_qcheck = request.POST.get("jenis_qcheck")
        qc.hasil_qcheck = request.POST.get("hasil_qcheck")
        qc.keterangan = request.POST.get("keterangan")
        qc.save()
        messages.success(request, "Q Check berhasil diperbarui.")
        return redirect("list_qcheck", id=qc.id_detailjob.id_detailjob)

    return render(request, "qcheck/edit_qcheck.html", {
        "qc": qc,
        "qc_types": models.qcheck.QC_TYPES
    })

# 7. View untuk Delete
@login_required
@require_POST
def delete_qcheck(request, id):
    qc = get_object_or_404(models.qcheck, id_qcheck=id)
    job_id = qc.id_detailjob.id_detailjob
    qc.delete()
    messages.success(request, "Q Check berhasil dihapus.")
    return redirect("list_qcheck", id=job_id)

def api_get_jobs_by_line_shift(request):
    line_id = request.GET.get('line')
    shift = request.GET.get('shift')
    today = timezone.now().date()

    jobs = models.detailjob.objects.filter(
        id_job__date=today,
        id_job__id_productionline__id=line_id,
        id_job__id_productionline__shift=shift
    ).select_related('id_itemproduksi').order_by('urutan')

    job_list = [{'id': job.id_detailjob, 'name': job.id_itemproduksi.job_number} for job in jobs]

    return JsonResponse({'jobs': job_list})

def get_cycle_time(request, item_id):
    try:
        item = models.itemproduksi.objects.get(pk=item_id)
        return JsonResponse({'cycle_time': item.cycle_time})
    except models.itemproduksi.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)
    
#fungsi laporan
@login_required 
def _get_laporan_data(request):
    # --- 1. AMBIL FILTER DARI URL ---
    
    # Ambil SEMUA line untuk dropdown, tapi juga nama unik untuk filter
    all_lines_qs = models.productionline.objects.all().order_by('namaline', 'shift')
    
    # Untuk dropdown Line (Nama Unik, cth: ['Stamping A', 'Stamping B'])
    line_names_unique = models.productionline.objects.values_list('namaline', flat=True).distinct().order_by('namaline')
    
    # Untuk dropdown Shift (Nama Unik, cth: [1, 2])
    shift_numbers_unique = all_lines_qs.values_list('shift', flat=True).distinct().order_by('shift')

    if not all_lines_qs:
        # Jika tidak ada line sama sekali di database
        return {'error': "Tidak ada Production Line yang terdaftar di database."}

    # Dapatkan Line pertama sebagai default jika filter tidak diset
    first_line_default = all_lines_qs.first()
    
    # 1. Ambil filter 'line' 
    default_line_slug = first_line_default.get_slug().split('-shift-')[0]
    selected_line_slug = request.GET.get('line', default_line_slug) 
    selected_line_name = selected_line_slug.replace('-', ' ').title() # Jadi 'Stamping A'
    
    # 2. Ambil filter 'shift'
    selected_shift_str = request.GET.get('shift', str(first_line_default.shift))
    try:
        selected_shift = int(selected_shift_str)
    except ValueError:
        selected_shift = first_line_default.shift 

    # 3. Ambil filter 'tanggal'
    selected_date_str = request.GET.get('tanggal', timezone.now().strftime('%Y-%m-%d'))
    # --- Tambahkan validasi tanggal di sini ---
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        # Jika format salah atau string kosong, default ke hari ini
        selected_date = timezone.now().date()
        selected_date_str = selected_date.strftime('%Y-%m-%d')

    try:
        # --- Mencari berdasarkan NAMA dan SHIFT ---
        production_line_obj = models.productionline.objects.get(
            namaline=selected_line_name,
            shift=selected_shift
        )
        current_shift_num = production_line_obj.shift
    
    except models.productionline.DoesNotExist:
        return {'error': f"Production line '{selected_line_name}' (Shift {selected_shift}) tidak ditemukan."}
    except models.productionline.MultipleObjectsReturned:
        return {'error': f"Error: Data duplikat ditemukan untuk '{selected_line_name}' (Shift {selected_shift}). Hubungi admin."}

    # --- 2. AMBIL DATA DARI DATABASE ---
    jobs_on_schedule = models.detailjob.objects.filter(
        id_job__id_productionline=production_line_obj, 
        id_job__date=selected_date
    ).order_by('urutan').select_related('id_itemproduksi', 'id_job').prefetch_related('machine_used')
    
    # Ambil data terkait SEKALI saja di luar loop
    all_downtimes_for_jobs = models.detail_dt.objects.filter(id_downtime__id_detailjob__in=jobs_on_schedule).select_related('id_downtime') # select_related
    all_dandoris_for_jobs = models.dandori.objects.filter(id_detailjob__in=jobs_on_schedule)
    all_qchecks_for_jobs = models.qcheck.objects.filter(id_detailjob__in=jobs_on_schedule)
    all_idles_for_jobs = models.detail_idle.objects.filter(id_idle__id_detailjob__in=jobs_on_schedule).select_related('id_idle') # Ambil data idle

    if not jobs_on_schedule.exists():
        empty_totals = {
            'plan_qty': 0, 'actual_good': 0, 'actual_repair': 0, 'actual_reject': 0, 'avg_plan_ct': 0,
            'total_schedule_start': None, 'total_schedule_finish': None, 'press_time': 0,
            'total_uchi_dies': 0, 'total_uchi_variant': 0, 'total_uchi_qcheck': 0, 'total_uchi': 0,
            'total_report_tpt': 0, 'idle_time': 0, 'total_panel_record_ct': 0, 'total_break_time': 0,
            'total_work_time': 0, 'gsph': 0, 'downtime_dies': 0, 'downtime_mach': 0,
            'downtime_matl': 0, 'downtime_pallet': 0, 'downtime_prod': 0, 'downtime_ubp': 0,
            'downtime_total': 0, 'tpt_plan': 0, 'tpt_act': 0, 'pass_rate': 0, 'repair_rate': 0,
            'reject_rate': 0, 'oee': 0,
        }
        return {
            'shop_line': selected_line_name, 'tanggal': selected_date, 'shift_kerja': f"Shift {current_shift_num}",
            'jobs_data': [], 'summary': {}, 'totals': empty_totals, 
            'all_line_names': line_names_unique,
            'all_shift_numbers': shift_numbers_unique,
            'selected_line_name': selected_line_name,
            'selected_shift': selected_shift,
            'selected_date_str': selected_date_str,
            'all_lines_qs': all_lines_qs, 
            'selected_line_slug': selected_line_slug,
        }

    # --- 3. PROSES SETIAP JOB & LAKUKAN PERHITUNGAN ---
    all_breaks = models.BreakTime.objects.all()
    shift_start = timezone.make_aware(datetime.datetime.combine(selected_date, datetime.time(7, 40) if current_shift_num == 1 else datetime.time(21, 0)))
    jobs_data = []
    plan_running_time = shift_start
    actual_running_time = shift_start
    list_of_cycle_times = []

    total_actual_stroke_finished = 0
    total_actual_tpt_minutes_finished = 0.0 # Gunakan float
    total_plan_stroke = 0
    total_plan_tpt_minutes = 0.0 # Gunakan float


    # --- PERBAIKAN INISIALISASI TOTALS ---
    totals = {
        'plan_qty': 0, 'actual_good': 0, 'actual_repair': 0, 'actual_reject': 0, 'press_time': 0.0, # Gunakan float
        'total_uchi': 0.0, 'idle_time': 0.0, 'downtime_total': 0.0, 'total_uchi_dies': 0.0,
        'total_uchi_variant': 0.0, 'total_uchi_qcheck': 0.0, 'total_report_tpt': 0.0,
        'total_panel_record_ct': 0.0, 'total_break_time': 0.0, 'total_work_time': 0.0, 
        'downtime_prod': 0.0, 'downtime_dies': 0.0, 'downtime_mach': 0.0, 'downtime_matl': 0.0,
        'downtime_log': 0.0, # (downtime_log)
        'downtime_ubp': 0.0, 
        'tpt_plan': 0.0, 'tpt_act': 0.0, 
    }
    # --- AKHIR PERBAIKAN INISIALISASI ---


    for job in jobs_on_schedule:
        # Gunakan nilai 0 jika None saat perhitungan
        actual_qty_safe = job.actual_qty or 0
        actual_repair_safe = job.actual_repair or 0
        actual_reject_safe = job.actual_reject or 0
        cycle_time_safe = float(job.id_itemproduksi.cycle_time or 0.0)
        plan_qty_safe = job.plan_qty or 0
        total_idle_minutes_safe = job.total_idle_minutes or 0.0 # Ambil dari property


        if cycle_time_safe > 0:
            list_of_cycle_times.append(cycle_time_safe)
        
        # press_time dihitung dari ACTUAL QTY
        press_time = (actual_qty_safe * cycle_time_safe) / 60.0 if cycle_time_safe else 0.0
        
        # Ambil total dari property model detailjob (lebih efisien)
        dandori_time = job.total_dandori_minutes or 0.0
        iq_check_time = job.total_qcheck_minutes or 0.0
        downtime_time = job.total_downtime_minutes or 0.0

        # plan_tpt dihitung dari press_time (berbasis actual_qty) + dandori + qcheck
        plan_tpt = press_time + dandori_time + iq_check_time
        # actual_tpt adalah plan_tpt + downtime
        actual_tpt = plan_tpt + downtime_time
        
        item_finish_plan = calculate_finish_time(plan_running_time, plan_tpt, all_breaks, current_shift_num)
        item_finish_actual = calculate_finish_time(actual_running_time, actual_tpt, all_breaks, current_shift_num)
        
        # Rincian Uchi (gunakan data yang sudah di-prefetch jika memungkinkan)
        dandoris_for_this_job = [d for d in all_dandoris_for_jobs if d.id_detailjob_id == job.id_detailjob]
        uchi_dies_change = sum(d.duration or 0.0 for d in dandoris_for_this_job if d.jenis_dandori == 'waktu_dandori')
        uchi_variant_change = sum(d.duration or 0.0 for d in dandoris_for_this_job if d.jenis_dandori == 'variant_change')
        qchecks_for_this_job = [q for q in all_qchecks_for_jobs if q.id_detailjob_id == job.id_detailjob]
        uchi_qcheck = sum(q.duration or 0.0 for q in qchecks_for_this_job)
        total_uchi = uchi_dies_change + uchi_variant_change + uchi_qcheck
        report_tpt = press_time + total_uchi

        # Hitung durasi break untuk job ini
        job_break_duration = 0.0
        job_start_actual = actual_running_time
        job_finish_actual = item_finish_actual
        # Pastikan waktu start dan finish valid sebelum menghitung break
        if job_start_actual and job_finish_actual and job_finish_actual > job_start_actual:
            for br in all_breaks:
                if br.shift == current_shift_num:
                    try: # Tambah try-except untuk parse waktu
                         break_start_dt = timezone.make_aware(datetime.datetime.combine(selected_date, br.waktu_mulai))
                         break_end_dt = timezone.make_aware(datetime.datetime.combine(selected_date, br.waktu_selesai))
                         if break_end_dt <= break_start_dt: break_end_dt += datetime.timedelta(days=1)
                         
                         overlap_start = max(job_start_actual, break_start_dt)
                         overlap_end = min(job_finish_actual, break_end_dt)
                         if overlap_end > overlap_start: 
                              job_break_duration += (overlap_end - overlap_start).total_seconds() / 60.0
                    except Exception as e_br:
                         # Log error jika waktu break tidak valid
                         print(f"Error calculating break overlap: {e_br}")
                         pass 

        # work_time_duration: Total waktu dari start sampai finish aktual
        # Ini berbeda dengan TPT Aktual. TPT = waktu kerja produktif + stop terencana + stop tak terencana
        # Work Time = TPT Aktual + Idle Time + Break Time
        work_time_duration = actual_tpt + total_idle_minutes_safe + job_break_duration

        # Ambil data idle
        idle_instance_for_this_job = next((idle for idle in all_idles_for_jobs if idle.id_idle.id_detailjob_id == job.id_detailjob), None)
        idle_type_display = idle_instance_for_this_job.reason_idle if idle_instance_for_this_job and idle_instance_for_this_job.reason_idle else "-"

        # Ambil data downtime
        dts_for_this_job = [dt for dt in all_downtimes_for_jobs if dt.id_downtime.id_detailjob_id == job.id_detailjob]
        
        # --- PERUBAHAN: Hitung dt_breakdown dengan loop biasa ---
        dt_breakdown = {'prod_t': 0.0, 'dies_t': 0.0, 'mach_t': 0.0, 'mat_t': 0.0, 'log_t': 0.0, 'ubp_t': 0.0} 
        jenis_map = {'prod_t': 'prod_t', 'dies_t': 'dies_t', 'mach_t': 'mach_t', 'mat_t': 'mat_t', 'log_t': 'log_t', 'ubp_t': 'ubp_t'} 

        for d in dts_for_this_job:
            try:
                jenis = d.jenisdowntime
                duration = d.duration_downtime or 0.0
                for key_map, value_map in jenis_map.items():
                     if jenis and value_map and jenis.lower() == value_map.lower(): 
                         if key_map in dt_breakdown:
                              dt_breakdown[key_map] += duration
                         break 
            except AttributeError as e:
                # Error ini seharusnya sudah hilang, tapi jaga-jaga
                print(f"!!! ERROR SAAT LOOP BIASA UNTUK BREAKDOWN: {e} pada objek {d}")
                pass 
        # --- AKHIR PERUBAHAN ---

        # Hitung Quality Rate & OEE
        total_pieces = actual_qty_safe + actual_repair_safe + actual_reject_safe
        pass_rate = (actual_qty_safe / total_pieces * 100.0) if total_pieces > 0 else 0.0
        repair_rate = (actual_repair_safe / total_pieces * 100.0) if total_pieces > 0 else 0.0
        reject_rate = (actual_reject_safe / total_pieces * 100.0) if total_pieces > 0 else 0.0
        
        oee = 0.0
        quality_oee = (actual_qty_safe / total_pieces) if total_pieces > 0 else 0.0
        
        # OEE Calculation components:
        # Availability = Run Time / Planned Production Time
        # Performance = (Total Pieces * Ideal Cycle Time) / Run Time
        # Quality = Good Pieces / Total Pieces
        
        # Planned Production Time (asumsi total durasi shift dikurangi break standar?) 
        # Atau gunakan work_time_duration? Definisi bisa bervariasi.
        # Mari gunakan work_time_duration sebagai basis waktu kerja.
        
        working_time_oee = work_time_duration # Waktu total dari start sampai finish aktual job ini
        planned_stops_oee = total_uchi + total_idle_minutes_safe + job_break_duration # Stop terencana + idle + break
        unplanned_stops_oee = downtime_time # Hanya downtime
        
        # Run time adalah waktu kerja total dikurangi semua stop (terencana & tak terencana)
        run_time_oee = working_time_oee - planned_stops_oee - unplanned_stops_oee 
        # Pastikan run_time tidak negatif
        run_time_oee = max(run_time_oee, 0.0) 

        availability_oee = (run_time_oee / working_time_oee) if working_time_oee > 0 else 0.0
        
        ideal_press_time_oee = (total_pieces * cycle_time_safe) / 60.0 if cycle_time_safe > 0 else 0.0
        
        performance_oee = (ideal_press_time_oee / run_time_oee) if run_time_oee > 0 else 0.0
        
        # OEE = Availability * Performance * Quality
        if availability_oee > 0 and performance_oee > 0 and quality_oee >= 0: # Quality bisa 0
             # Batasi nilai A, P, Q maksimal 1 (100%) untuk perhitungan OEE
             availability_oee = min(availability_oee, 1.0)
             performance_oee = min(performance_oee, 1.0)
             quality_oee = min(quality_oee, 1.0)
             oee = availability_oee * performance_oee * quality_oee * 100.0

        
        # Hitung Stroke & GSPH
        num_machines = job.machine_used.count()
        job_plan_stroke = plan_qty_safe * num_machines      
        job_actual_stroke = actual_qty_safe * num_machines  
        
        # Akumulasi untuk Plan GSPH (semua job)
        total_plan_stroke += job_plan_stroke
        total_plan_tpt_minutes += plan_tpt 

        # Akumulasi untuk Actual GSPH (hanya job yang sudah selesai)
        # Pastikan item_finish_actual tidak None
        if item_finish_actual and item_finish_actual < timezone.now():
            total_actual_stroke_finished += job_actual_stroke
            total_actual_tpt_minutes_finished += actual_tpt 
            
        # Hitung GSPH per item (berdasarkan TPT aktual)
        if actual_tpt > 0 and num_machines > 0:
            gsph_actual_per_item = ((actual_qty_safe * num_machines) / actual_tpt) * 60.0
        else:
            gsph_actual_per_item = 0.0
        
        
        jobs_data.append({
            'job_no': job.id_itemproduksi.job_number, 
            'plan_qty': plan_qty_safe, 
            'plan_ct': job.id_itemproduksi.cycle_time, # Tampilkan CT asli dari item
            'actual_good': actual_qty_safe, 
            'actual_repair': actual_repair_safe, 
            'actual_reject': actual_reject_safe,
            'total_pieces': total_pieces, 
            'schedule_start': plan_running_time, 
            'schedule_finish': item_finish_plan,
            'actual_start': actual_running_time, 
            'actual_finish': item_finish_actual, 
            'press_time': press_time,
            'uchi_dies_change': uchi_dies_change, 
            'uchi_variant_change': uchi_variant_change, 
            'uchi_qcheck': uchi_qcheck,
            'total_uchi': total_uchi, 
            'report_tpt': report_tpt, 
            'idle_time': total_idle_minutes_safe, # Gunakan total idle
            'dt_breakdown': dt_breakdown, # Kirim rincian downtime
            'dt_total': downtime_time, 
            'tpt_plan': plan_tpt, 
            'tpt_act': actual_tpt, 
            'pass_rate': pass_rate,
            'repair_rate': repair_rate, 
            'reject_rate': reject_rate, 
            'oee': oee, 
            'gsph': gsph_actual_per_item, 
            'panel_record_ct': cycle_time_safe or 0, # Gunakan cycle time yg sudah divalidasi
            'break_time_duration': job_break_duration,
            'work_time_duration': work_time_duration, 
            'idle_time_type': idle_type_display,
            'note': job.notes if hasattr(job, 'notes') else '' # Tambahkan note jika ada
        })
        
        # Update running time (handle None)
        plan_running_time = item_finish_plan if item_finish_plan else plan_running_time
        actual_running_time = item_finish_actual if item_finish_actual else actual_running_time
        
        # --- PERBAIKAN AKUMULASI TOTALS ---
        totals['plan_qty'] += plan_qty_safe
        totals['actual_good'] += actual_qty_safe
        totals['actual_repair'] += actual_repair_safe
        totals['actual_reject'] += actual_reject_safe
        totals['press_time'] += press_time
        totals['total_uchi'] += total_uchi
        totals['idle_time'] += total_idle_minutes_safe 
        totals['total_report_tpt'] += report_tpt
        totals['downtime_total'] += downtime_time 
        totals['downtime_prod'] += dt_breakdown.get('prod_t', 0.0) 
        totals['downtime_dies'] += dt_breakdown.get('dies_t', 0.0)
        totals['downtime_mach'] += dt_breakdown.get('mach_t', 0.0)
        totals['downtime_matl'] += dt_breakdown.get('mat_t', 0.0)
        # Ambil dari 'log_t' tapi simpan ke 'downtime_pallet'
        totals['downtime_log'] += dt_breakdown.get('log_t', 0.0)
        totals['downtime_ubp'] += dt_breakdown.get('ubp_t', 0.0) 
        totals['total_uchi_dies'] += uchi_dies_change
        totals['total_uchi_variant'] += uchi_variant_change
        totals['total_uchi_qcheck'] += uchi_qcheck
        totals['total_panel_record_ct'] += cycle_time_safe or 0.0 
        totals['total_break_time'] += job_break_duration
        totals['total_work_time'] += work_time_duration 
        totals['tpt_plan'] += plan_tpt 
        totals['tpt_act'] += actual_tpt 
        # --- AKHIR PERBAIKAN AKUMULASI ---


    # --- 4. HITUNG NILAI TOTAL AKHIR (SEBELUM SUMMARY) ---
    total_all_actual_pieces_final = totals['actual_good'] + totals['actual_repair'] + totals['actual_reject']

    if total_all_actual_pieces_final > 0:
        totals['pass_rate'] = (totals['actual_good'] / total_all_actual_pieces_final) * 100.0
        totals['repair_rate'] = (totals['actual_repair'] / total_all_actual_pieces_final) * 100.0
        totals['reject_rate'] = (totals['actual_reject'] / total_all_actual_pieces_final) * 100.0
    else:
        totals['pass_rate'], totals['repair_rate'], totals['reject_rate'] = 0.0, 0.0, 0.0

    # Rata-rata OEE
    all_oee_values = [job['oee'] for job in jobs_data if job.get('oee', 0) > 0]
    totals['oee'] = sum(all_oee_values) / len(all_oee_values) if all_oee_values else 0.0

    # Hitung GSPH Final
    if total_plan_tpt_minutes > 0:
        gsph_plan_final = (total_plan_stroke / total_plan_tpt_minutes) * 60.0
    else:
        gsph_plan_final = 0.0

    if total_actual_tpt_minutes_finished > 0:
        gsph_act_final = (total_actual_stroke_finished / total_actual_tpt_minutes_finished) * 60.0
    else:
        gsph_act_final = 0.0

    totals['gsph'] = gsph_act_final
        
    # --- 5. HITUNG SUMMARY ---
    totals['avg_plan_ct'] = sum(list_of_cycle_times) / len(list_of_cycle_times) if list_of_cycle_times else 0.0
    totals['total_schedule_start'] = shift_start
    # Handle jika plan_running_time None (misal tidak ada job)
    totals['total_schedule_finish'] = plan_running_time if plan_running_time else shift_start 
    
    summary = {
        'item_plan': jobs_on_schedule.count(), 
        'item_act': jobs_on_schedule.filter(actual_qty__gt=0).count(),
        'qty_plan': totals['plan_qty'], 
        'qty_act': totals['actual_good'],
        'tpt_plan': totals['tpt_plan'], # Gunakan total TPT Plan
        'tpt_act': totals['tpt_act'],   # Gunakan total TPT Actual
        
        'gsph_plan': gsph_plan_final,
        'gsph_act': gsph_act_final,
        
        'pass_rate_plan': 100.00, 'pass_rate_act': totals['pass_rate'],
        # Ambil plan rate dari productionplan jika ada, jika tidak default 2%
        'reject_rate_plan': plan_reject_percent if 'plan_reject_percent' in locals() else 2.00, 
        'reject_rate_act': totals['reject_rate'],
        'repair_rate_plan': plan_repair_percent if 'plan_repair_percent' in locals() else 2.00, 
        'repair_rate_act': totals['repair_rate'],
    }
    
    # Hitung persentase summary (handle pembagian dengan nol)
    summary['item_percent'] = (summary['item_act'] / summary['item_plan'] * 100.0) if summary['item_plan'] > 0 else 0.0
    summary['qty_percent'] = (summary['qty_act'] / summary['qty_plan'] * 100.0) if summary['qty_plan'] > 0 else 0.0
    summary['tpt_percent'] = (summary['tpt_act'] / summary['tpt_plan'] * 100.0) if summary['tpt_plan'] > 0 else 0.0
    summary['gsph_percent'] = (summary['gsph_act'] / summary['gsph_plan'] * 100.0) if summary['gsph_plan'] > 0 else 0.0
    
    # --- 6. KEMBALIKAN SEMUA DATA ---
    return {
        'shop_line': production_line_obj.namaline, 
        'tanggal': selected_date, 
        'shift_kerja': f"Shift {production_line_obj.shift}",
        'jobs_data': jobs_data, 
        'summary': summary, 
        'totals': totals, 
        
        # Data untuk filter dropdown di template laporan
        'all_line_names': line_names_unique,     
        'all_shift_numbers': shift_numbers_unique,
        'selected_line_name': selected_line_name,  
        'selected_shift': selected_shift,          
        'selected_date_str': selected_date_str,      

        # Mungkin tidak perlu lagi jika template sudah disesuaikan
        'all_lines_qs': all_lines_qs, 
        'selected_line_slug': selected_line_slug, 
    }

# VIEW UNTUK MENAMPILKAN HALAMAN WEB
@login_required
def laporan_harian(request):
    context = _get_laporan_data(request)
    if 'error' in context:
        return HttpResponse(context['error'])
    return render(request, 'laporan/laporan.html', context)

# VIEW UNTUK EKSPOR EXCEL
@login_required
def export_laporan_excel(request):
    context = _get_laporan_data(request)
    if 'error' in context:
        return HttpResponse(context['error'])

    shop_line = context.get('shop_line', '-')
    tanggal = context.get('tanggal')
    shift_kerja = context.get('shift_kerja', '-')
    jobs_data = context.get('jobs_data', [])
    summary = context.get('summary', {})
    totals = context.get('totals', {})

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    line_name_slug = context.get('selected_line_name', 'data').replace(' ', '-')
    shift_num = context.get('selected_shift', 'S')
    tanggal_str = tanggal.strftime("%Y-%m-%d")
    
    filename = f'Laporan-Harian-{line_name_slug}-Shift-{shift_num}-{tanggal_str}.xlsx'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Harian"

    # === STYLE ===
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    def auto_fit_columns(worksheet):
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # === BAGIAN 0: INFO LAPORAN ===
    sheet.merge_cells('A1:AI1')
    main_title = sheet['A1']
    main_title.value = "LAPORAN KERJA HARIAN STAMPING"
    main_title.font = Font(bold=True, size=16)
    main_title.alignment = copy.copy(center_align)
    sheet['A3'] = "SHOP LINE"; sheet['B3'] = shop_line
    sheet['A4'] = "TANGGAL"; sheet['B4'] = tanggal.strftime("%d/%m/%Y")
    sheet['A5'] = "SHIFT KERJA"; sheet['B5'] = shift_kerja

    sheet['AG3'] = "Supervisor"
    sheet['AH3'] = "Foreman"
    sheet['AI3'] = "T.Leader"

    sheet.row_dimensions[4].height = 60 

    for col_letter in ['AG', 'AH', 'AI']:
            sheet.column_dimensions[col_letter].width = 18

            header_cell = sheet[f'{col_letter}3']
            header_cell.font = Font(bold=True, size=12) 
            header_cell.alignment = copy.copy(center_align) 
            header_cell.border = copy.copy(thin_border)     

            box_cell = sheet[f'{col_letter}4']
            box_cell.border = copy.copy(thin_border)        
            box_cell.alignment = copy.copy(center_align)    

# === BAGIAN 1: SUMMARY ACHIEVEMENT ===
    current_row = 7
    sheet.merge_cells(f'A{current_row}:D{current_row}')
    summary_title = sheet.cell(row=current_row, column=1, value="SUMMARY ACHIEVEMENT")
    summary_title.font = header_font; summary_title.alignment = center_align; summary_title.fill = header_fill
    for col_idx in range(1, 5):
        sheet.cell(row=current_row, column=col_idx).border = thin_border

    current_row += 1
    summary_headers = ["", "Plan", "Act", "%"]
    sheet.append(summary_headers)

    for col_idx in range(1, 5):
        cell = sheet.cell(row=current_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    summary_data_rows = [
        ("ITEM PROCESS", summary.get('item_plan', 0), summary.get('item_act', 0), f"{summary.get('item_percent', 0):.1f}%"),
        ("QTY PROCESS", summary.get('qty_plan', 0), summary.get('qty_act', 0), f"{summary.get('qty_percent', 0):.1f}%"),
        ("TPT PROCESS", f"{summary.get('tpt_plan', 0):.0f}", f"{summary.get('tpt_act', 0):.0f}", f"{summary.get('tpt_percent', 0):.1f}%"),
        ("GSPH", f"{summary.get('gsph_plan', 0):.0f}", f"{summary.get('gsph_act', 0):.0f}", f"{summary.get('gsph_percent', 0):.1f}%"),
        ("PASS RATE", f"{summary.get('pass_rate_plan', 0):.2f}%", f"{summary.get('pass_rate_act', 0):.2f}%", ""),
        ("REJECT RATE", f"{summary.get('reject_rate_plan', 0):.2f}%", f"{summary.get('reject_rate_act', 0):.2f}%", ""),
        ("REPAIR RATE", f"{summary.get('repair_rate_plan', 0):.2f}%", f"{summary.get('repair_rate_act', 0):.2f}%", ""),
    ]

    for row_data in summary_data_rows:
        sheet.append(row_data)
        max_row = sheet.max_row
        for col_idx in range(1, 5): 
            cell = sheet.cell(row=max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 1 else center_align

# === BAGIAN 2: PRODUCTION SCHEDULE ===
    current_row = sheet.max_row + 2
    sheet.merge_cells(f'A{current_row}:S{current_row}')
    ps_title = sheet.cell(row=current_row, column=1, value="PRODUCTION SCHEDULE")
    ps_title.font = header_font; ps_title.alignment = center_align; ps_title.fill = header_fill
    for col_idx in range(1, 20):
        sheet.cell(row=current_row, column=col_idx).border = thin_border

    current_row += 1
    h1 = ["NO", "Jobs No", "QTY (PCS)", None, "PLAN CYCLE TIME (detik)", "SCHEDULE", None, "UCHI DANDORI (MENIT)", None, None, None, None, "TPT (MENIT)", "IDLE TIME (MENIT)", None, "BREAK TIME", "WORK TIME", "GSPH", "NOTE"]
    h2 = [None, None, "PLAN", "ACT", None, "START", "FINISH", "PRESS TIME", "DIES CHANGE", "VARIANT", "EARLY CHECK", "TOTAL UCHI", None, "TYPE", "TIME", "(MENIT)", None, None, None]
    sheet.append(h1)
    sheet.append(h2)

    sheet.merge_cells(f'A{current_row}:A{current_row+1}'); sheet.merge_cells(f'B{current_row}:B{current_row+1}')
    sheet.merge_cells(f'C{current_row}:D{current_row}'); sheet.merge_cells(f'E{current_row}:E{current_row+1}')
    sheet.merge_cells(f'F{current_row}:G{current_row}'); sheet.merge_cells(f'H{current_row}:L{current_row}')
    sheet.merge_cells(f'M{current_row}:M{current_row+1}'); sheet.merge_cells(f'N{current_row}:O{current_row}')
    sheet.merge_cells(f'P{current_row}:P{current_row+1}'); sheet.merge_cells(f'Q{current_row}:Q{current_row+1}')
    sheet.merge_cells(f'R{current_row}:R{current_row+1}'); sheet.merge_cells(f'S{current_row}:S{current_row+1}')

    for row in sheet.iter_rows(min_row=current_row, max_row=current_row+1, min_col=1, max_col=19):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border

    for i, job in enumerate(jobs_data, 1):
        start_time = job.get('schedule_start').strftime("%H:%M") if job.get('schedule_start') else "-"
        finish_time = job.get('schedule_finish').strftime("%H:%M") if job.get('schedule_finish') else "-"
        row_data = [
            i, job.get('job_no', '-'), job.get('plan_qty', 0), job.get('actual_good', 0),
            job.get('plan_ct', 0), start_time, finish_time, job.get('press_time', 0),
            job.get('uchi_dies_change', 0), job.get('uchi_variant_change', 0),
            job.get('uchi_qcheck', 0), job.get('total_uchi', 0), job.get('report_tpt', 0),
            job.get('idle_time_type', '-'), job.get('idle_time', 0),
            job.get('break_time_duration', 0), 
            job.get('work_time_duration', 0), job.get('gsph', 0), job.get('note', '')
        ]
        sheet.append(row_data)
        max_row = sheet.max_row
        for col_idx in range(1, 20):
            cell = sheet.cell(row=max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align

    total_row_data = [
        "TOTAL PLAN", None,
        totals.get('plan_qty', 0),
        totals.get('actual_good', 0),
        None, None, None, 
        totals.get('press_time', 0),
        totals.get('total_uchi_dies', 0),
        totals.get('total_uchi_variant', 0),
        totals.get('total_uchi_qcheck', 0),
        totals.get('total_uchi', 0),
        totals.get('total_report_tpt', 0), 
        None,
        totals.get('idle_time', 0),
        totals.get('total_break_time', 0), 
        totals.get('total_work_time', 0),
        f"{totals.get('gsph', 0):.2f}",
        None
    ]
    sheet.append(total_row_data)
    sheet.merge_cells(f'A{sheet.max_row}:B{sheet.max_row}')
    max_row = sheet.max_row
    for col_idx in range(1, 20):
        cell = sheet.cell(row=max_row, column=col_idx)
        cell.font = header_font; cell.border = thin_border; cell.fill = total_fill; cell.alignment = center_align
    
# === BAGIAN 3: ACTUAL PROCESS ===
    current_row = sheet.max_row + 3
    sheet.merge_cells(f'A{current_row}:AI{current_row}')
    ap_title = sheet.cell(row=current_row, column=1, value="ACTUAL PROCESS")
    ap_title.font = header_font
    ap_title.alignment = center_align
    ap_title.fill = header_fill
    for col_idx in range(1, 36):
        sheet.cell(row=current_row, column=col_idx).border = thin_border

    current_row += 1
    h1_ap = ["NO", "Jobs No", "PLAN QTY", "ACTUAL QTY", None, None, None, "PLAN SCHEDULE", None, "ACTUAL PROCESS", None, "CT Actual (detik)", "UCHI DANDORI (MENIT)", None, None, None, None, "DOWN TIME (MENIT)", None, None, None, None, None, "TPT (MENIT)", None, "IDLE TIME", None, "BREAK TIME", None, "Work Time", "QUALITY RATE", None, None, "OEE", "GSPH"]
    h2_ap = [None, None, "Plan", "Good", "Repair", "Reject", "Total", "Start", "Finish", "Start", "Finish", None, "Press Time", "Dies Change", "Variant", "Early Check", "Total", "Dies", "Mach", "Matl", "Log", "Prod/ Handling", "Total", "Plan", "Act", "Type", "Time", "Type", "Time", None, "Pass Rate", "Repair Rate", "Reject Rate", None, None]

    sheet.append(h1_ap)
    sheet.append(h2_ap)

    header_start_row = current_row
    sheet.merge_cells(f'A{header_start_row}:A{header_start_row+1}'); sheet.merge_cells(f'B{header_start_row}:B{header_start_row+1}')
    sheet.merge_cells(f'C{header_start_row}:C{header_start_row+1}'); sheet.merge_cells(f'D{header_start_row}:G{header_start_row}')
    sheet.merge_cells(f'H{header_start_row}:I{header_start_row}'); sheet.merge_cells(f'J{header_start_row}:K{header_start_row}')
    sheet.merge_cells(f'L{header_start_row}:L{header_start_row+1}'); sheet.merge_cells(f'M{header_start_row}:Q{header_start_row}')
    sheet.merge_cells(f'R{header_start_row}:W{header_start_row}'); sheet.merge_cells(f'X{header_start_row}:Y{header_start_row}')
    sheet.merge_cells(f'Z{header_start_row}:AA{header_start_row}'); sheet.merge_cells(f'AB{header_start_row}:AC{header_start_row}')
    sheet.merge_cells(f'AD{header_start_row}:AD{header_start_row+1}'); sheet.merge_cells(f'AE{header_start_row}:AG{header_start_row}')
    sheet.merge_cells(f'AH{header_start_row}:AH{header_start_row+1}'); sheet.merge_cells(f'AI{header_start_row}:AI{header_start_row+1}')

    for row in sheet.iter_rows(min_row=header_start_row, max_row=header_start_row+1, min_col=1, max_col=35):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border

    for i, job in enumerate(jobs_data, 1):
        dt = job.get('dt_breakdown', {})
        row_data = [
            i, job.get('job_no', '-'), job.get('plan_qty', 0), job.get('actual_good', 0),
            job.get('actual_repair', 0), job.get('actual_reject', 0), job.get('total_pieces', 0),
            job.get('schedule_start').strftime('%H:%M') if job.get('schedule_start') else '-',
            job.get('schedule_finish').strftime('%H:%M') if job.get('schedule_finish') else '-',
            job.get('actual_start').strftime('%H:%M') if job.get('actual_start') else '-',
            job.get('actual_finish').strftime('%H:%M') if job.get('actual_finish') else '-',
            f"{job.get('panel_record_ct', 0):.1f}", f"{job.get('press_time', 0):.1f}", f"{job.get('uchi_dies_change', 0):.2f}",
            f"{job.get('uchi_variant_change', 0):.2f}", f"{job.get('uchi_qcheck', 0):.2f}", f"{job.get('total_uchi', 0):.2f}",
            f"{dt.get('dies_t', 0):.2f}", f"{dt.get('mach_t', 0):.2f}", f"{dt.get('mat_t', 0):.2f}",
            f"{dt.get('pallet_t', 0):.2f}", f"{dt.get('prod_t', 0):.2f}", f"{job.get('dt_total', 0):.1f}",
            f"{job.get('tpt_plan', 0):.1f}", f"{job.get('tpt_act', 0):.1f}", job.get('idle_time_type', '-'),
            f"{job.get('idle_time', 0):.2f}", "Istirahat" if job.get('break_time_duration', 0) > 0 else "-",
            f"{job.get('break_time_duration', 0):.1f}", f"{job.get('work_time_duration', 0):.1f}",
            f"{job.get('pass_rate', 0):.2f}%", f"{job.get('repair_rate', 0):.2f}%", f"{job.get('reject_rate', 0):.2f}%",
            f"{job.get('oee', 0):.2f}%", f"{job.get('gsph', 0):.0f}"
        ]
        sheet.append(row_data)
        max_row = sheet.max_row
        for col_idx in range(1, 36): 
            cell = sheet.cell(row=max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align

    total_ap_data = [
        "TOTAL", None, 
        totals.get('plan_qty', 0), 
        totals.get('actual_good', 0),
        totals.get('actual_repair', 0), 
        totals.get('actual_reject', 0),
        totals.get('actual_good', 0) + totals.get('actual_repair', 0) + totals.get('actual_reject', 0),
        None, None, None, None, 
        f"{totals.get('total_panel_record_ct', 0):.1f}", 
        f"{totals.get('press_time', 0):.1f}",
        f"{totals.get('total_uchi_dies', 0):.1f}", 
        f"{totals.get('total_uchi_variant', 0):.1f}", 
        f"{totals.get('total_uchi_qcheck', 0):.1f}",
        f"{totals.get('total_uchi', 0):.1f}", 
        f"{totals.get('downtime_dies', 0):.1f}", 
        f"{totals.get('downtime_mach', 0):.1f}",
        f"{totals.get('downtime_matl', 0):.1f}", 
        f"{totals.get('downtime_pallet', 0):.1f}", 
        f"{totals.get('downtime_prod', 0):.1f}",
        f"{totals.get('downtime_total', 0):.1f}", 
        f"{totals.get('tpt_plan', 0):.1f}",
        f"{totals.get('tpt_act', 0):.1f}",
        None, 
        f"{totals.get('idle_time', 0):.1f}",
        None, 
        f"{totals.get('total_break_time', 0):.1f}",
        f"{totals.get('total_work_time', 0):.1f}",
        f"{totals.get('pass_rate', 0):.2f}%",
        f"{totals.get('repair_rate', 0):.2f}%",
        f"{totals.get('reject_rate', 0):.2f}%",
        f"{totals.get('oee', 0):.2f}%",
        f"{totals.get('gsph', 0):.2f}"
    ]
    sheet.append(total_ap_data)
    sheet.merge_cells(f'A{sheet.max_row}:B{sheet.max_row}')
    sheet.merge_cells(f'H{sheet.max_row}:K{sheet.max_row}')

    max_row = sheet.max_row
    for col_idx in range(1, 36):
        cell = sheet.cell(row=max_row, column=col_idx)
        cell.font = header_font; cell.border = thin_border; cell.fill = total_fill; cell.alignment = center_align

    auto_fit_columns(sheet)
    workbook.save(response)
    return response

# VIEW UNTUK EKSPOR PDF
@login_required
def export_laporan_pdf(request):
    context = _get_laporan_data(request)
    if 'error' in context:
        return HttpResponse(context['error'])

    # Render template HTML menjadi string
    html_string = get_template('laporan/laporan_pdf.html').render(context)
    
    # Buat objek HTML dari WeasyPrint
    html = HTML(string=html_string, base_url=request.build_absolute_uri())

    # Render PDF
    pdf_file = html.write_pdf()

    # Buat response HTTP
    response = HttpResponse(pdf_file, content_type='application/pdf')
    
    # 1. Ambil data untuk nama file
    selected_line_slug = context.get('selected_line_slug', 'data')
    selected_shift = context.get('selected_shift', 'S') 
    selected_date_str = context.get('tanggal').strftime('%Y-%m-%d')

    # 2. Tambahkan shift ke nama file
    filename = f'Laporan-Harian-{selected_line_slug}-Shift-{selected_shift}-{selected_date_str}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

