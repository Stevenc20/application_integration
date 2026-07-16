#!/usr/bin/env python3
"""
read_schedule_stamping.py
Parses Schedule Stamping Excel files.

Structure per sheet:
  - One sheet may contain PRESS A, PRESS B, PRESS C, PRESS D as vertical sections.
  - Each section has its own mini-header block (Hari, Tgl, Jam rows) followed
    by a data-header row (NO. | JOB MASTER | ... | JOB NO. | ... | START | FINISH)
    and then data rows until the next PRESS marker or end of sheet.
  - Rev sheet takes priority over non-Rev counterpart.
"""
import sys
import json
import warnings
import os
import re
import glob
from datetime import datetime, time as dt_time, date as dt_date
import openpyxl

warnings.filterwarnings('ignore')

# ── helpers ───────────────────────────────────────────────────────────────────

def safe_f(v):
    if v is None: return 0.0
    try:
        if isinstance(v, str):
            v = v.replace(',', '.').strip()
            if v in ('', '-', '#N/A', 'N/A', '#REF!', '#VALUE!'): return 0.0
        return float(v)
    except:
        return 0.0

def safe_i(v):
    if v is None: return 0
    try:
        return int(float(str(v).strip()))
    except:
        return 0

def fmt_time(v):
    """Convert a cell value (time, datetime, string) to HH:MM string or None."""
    if v is None:
        return None
    if isinstance(v, dt_time):
        return v.strftime('%H:%M')
    if isinstance(v, datetime):
        return v.strftime('%H:%M')
    if isinstance(v, dt_date) and not isinstance(v, datetime):
        return None
    if isinstance(v, str):
        s = v.strip()
        if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', s):
            return s[:5]
        return None
    # openpyxl sometimes gives a float (fraction of day)
    if isinstance(v, float):
        total_min = round(v * 24 * 60)
        h = (total_min // 60) % 24
        m = total_min % 60
        return f"{h:02d}:{m:02d}"
    return None

def extract_date(text):
    if not text:
        return None
    months = ['JANUARI','FEBRUARI','MARET','APRIL','MEI','JUNI',
              'JULI','AGUSTUS','SEPTEMBER','OKTOBER','NOVEMBER','DESEMBER']
    text_up = text.upper()
    for m in months:
        match = re.search(rf'(\d{{1,2}})[\s\-/]+{m}[\s\-/]+(\d{{4}})', text_up)
        if match:
            return f"{match.group(1).zfill(2)} {m} {match.group(2)}"
    # Try numeric date e.g. 07-05-2026 or 07/05/2026
    match = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', text_up)
    if match:
        d, mo, y = match.group(1), int(match.group(2)), match.group(3)
        return f"{d} {months[mo-1]} {y}"
    return None

# ── main section parser ────────────────────────────────────────────────────────

def is_press_marker(row, col2_val):
    """True if this row marks the beginning of a new Press section."""
    if col2_val is None:
        return False
    s = str(col2_val).strip().upper()
    return bool(re.match(r'^PRESS\s+[A-Z]$', s))

HEADER_KEYWORDS = frozenset({'NO.', 'NO', 'JOB MASTER'})

def is_header_row(row, col2_val):
    """True if this row is the data header row (contains NO./NO or JOB MASTER in any cell)."""
    for cell in row:
        if cell is not None and str(cell).strip().upper() in HEADER_KEYWORDS:
            return True
    return False

def row_value(row, idx, default=None):
    """Safe index into a row tuple."""
    if idx is None or idx >= len(row):
        return default
    return row[idx] if row[idx] is not None else default

def parse_sheet(ws, sheet_name):
    """
    Parse a single worksheet.
    Returns a dict: { 'SheetName|||PRESS X': {shift_name, press_name, hari, tgl, jam, revisi, rows:[]} }
    """
    all_rows = list(ws.iter_rows(values_only=True))
    total = len(all_rows)

    result = {}

    # Scan for PRESS section starts (col index 2 contains 'PRESS A' etc.)
    press_starts = []  # list of (row_idx_0based, press_name)
    for r_idx, row in enumerate(all_rows):
        col2 = row[2] if len(row) > 2 else None
        if is_press_marker(row, col2):
            press_name = str(col2).strip().upper()
            press_starts.append((r_idx, press_name))

    if not press_starts:
        # Fallback: treat whole sheet as one block, try to detect press from sheet name
        press_name = 'PRESS A'
        for p in ['PRESS A', 'PRESS B', 'PRESS C', 'PRESS D']:
            if p in sheet_name.upper():
                press_name = p
                break
        press_starts = [(0, press_name)]

    # Determine end row for each press section
    for i, (start_r, press_name) in enumerate(press_starts):
        end_r = press_starts[i + 1][0] if i + 1 < len(press_starts) else total

        section_rows = all_rows[start_r:end_r]
        section = parse_press_section(section_rows, start_r, sheet_name, press_name)
        if section and section['rows']:
            key = f"{sheet_name}|||{press_name}"
            result[key] = section

    return result

def parse_press_section(section_rows, offset, sheet_name, press_name):
    """
    Parse rows belonging to one press section.
    `offset` is the absolute row index of the first row in section_rows.
    """
    hari = None
    tgl = None
    jam = None
    revisi = None
    last_job_master = None

    # Find the header row (col 2 == 'NO.' or 'NO')
    header_local_idx = None
    col_map = {}

    for local_i, row in enumerate(section_rows):
        col2 = row[2] if len(row) > 2 else None
        if is_header_row(row, col2):
            header_local_idx = local_i
            # Build column map from this header row
            # We track job_no separately: prefer 'JOB NO.' (with dot, the main schedule
            # column at ~col 8) over 'JOB NO' (no dot, the right-side summary column
            # at ~col 37 which is often empty in data rows for PRESS B/C/D).
            job_no_main = None   # JOB NO. (with dot)
            job_no_alt  = None   # JOB NO  (without dot, right-side summary)
            for j, v in enumerate(row):
                if v is None:
                    continue
                v_str = str(v).upper().replace('\n', ' ').strip()
                if v_str in ('NO', 'NO.') and 'row_no' not in col_map:                    col_map['row_no']       = j
                elif 'JOB MASTER' in v_str and 'job_master' not in col_map:                  col_map['job_master']   = j
                elif 'TYPE PLT' in v_str and 'type_plt' not in col_map:                    col_map['type_plt']     = j
                elif ('QTY/PLT' in v_str or 'QTY/ PLT' in v_str or 'QTY/ PLT' in v_str) and 'qty_plt' not in col_map:
                                                              col_map['qty_plt']      = j
                elif 'KEB. MTL' in v_str and 'keb_mtl' not in col_map:                    col_map['keb_mtl']      = j
                elif 'TOTAL PLT' in v_str and 'total_plt' not in col_map:                   col_map['total_plt']    = j
                elif 'JOB NO.' in v_str and job_no_main is None:                     job_no_main             = j
                elif v_str == 'JOB NO' and job_no_alt is None:                      job_no_alt              = j
                elif 'EACH PART' in v_str and 'each_part' not in col_map:                   col_map['each_part']    = j
                elif (v_str == 'PLAN (PCS)' or v_str == 'PLAN') and 'plan' not in col_map:
                                                              col_map['plan']         = j
                elif v_str == 'OK' and 'ok' not in col_map:                          col_map['ok']           = j
                elif v_str == 'REPAIR' and 'repair' not in col_map:                      col_map['repair']       = j
                elif v_str == 'REJECT' and 'reject' not in col_map:                      col_map['reject']       = j
                elif 'TOTAL MESIN' in v_str and 'total_mesin' not in col_map:                 col_map['total_mesin']  = j
                elif ('CT (' in v_str or 'CYCLE TIME' in v_str) and 'ct_detik' not in col_map: col_map['ct_detik']  = j
                elif 'PROCESS TIME' in v_str and 'process_time' not in col_map:                col_map['process_time'] = j
                elif ('REG. ACTIVE' in v_str or 'REG.ACTIVE' in v_str) and 'reg_active' not in col_map:
                                                              col_map['reg_active']   = j
                elif v_str == 'DCT' and 'dct' not in col_map:                         col_map['dct']          = j
                elif v_str == 'MCT' and 'mct' not in col_map:                         col_map['mct']          = j
                elif 'PLAN DCT' in v_str and 'plan_dct' not in col_map:                    col_map['plan_dct']     = j
                elif v_str == 'TPT' and 'tpt' not in col_map:                         col_map['tpt']          = j
                elif 'GSPH' in v_str and 'gsph_item' not in col_map:                        col_map['gsph_item']    = j
                elif v_str == 'START' and 'start_time' not in col_map:                       col_map['start_time']   = j
                elif v_str == 'FINISH' and 'finish_time' not in col_map:                      col_map['finish_time']  = j
                elif 'ACT START' in v_str and 'act_start' not in col_map:                   col_map['act_start']    = j
                elif 'ACT FINISH' in v_str and 'act_finish' not in col_map:                  col_map['act_finish']   = j
                elif 'KETERANGAN' in v_str and 'keterangan' not in col_map:                  col_map['keterangan']   = j
                elif v_str in ('A-1', 'B-1', 'C-1', 'D-1') and 'a1' not in col_map:  col_map['a1']           = j
                elif v_str in ('A-2', 'B-2', 'C-2', 'D-2') and 'a2' not in col_map:  col_map['a2']           = j
                elif v_str in ('A-3', 'B-3', 'C-3', 'D-3') and 'a3' not in col_map:  col_map['a3']           = j
                elif v_str in ('A-4', 'B-4', 'C-4', 'D-4') and 'a4' not in col_map:  col_map['a4']           = j
                elif ('DT (MENIT)' in v_str and 'TOTAL' not in v_str) and 'dt_menit' not in col_map:
                                                              col_map['dt_menit']     = j
                elif 'TOTAL PCS' in v_str and 'total_pcs' not in col_map:                   col_map['total_pcs']    = j
                elif 'TPT TOTAL' in v_str and 'tpt_total' not in col_map:                   col_map['tpt_total']    = j
            # Prioritise JOB NO. (with dot) as the primary job_no column;
            # only fall back to JOB NO (no dot) if the dotted version wasn't found.
            col_map['job_no_main'] = job_no_main
            col_map['job_no_alt']  = job_no_alt
            col_map['job_no']      = job_no_main if job_no_main is not None else job_no_alt
            break

    if header_local_idx is None:
        return None  # no data header found in this section

    # Extract meta info (Hari/Tgl/Jam/Revisi) from rows BEFORE the header
    for row in section_rows[:header_local_idx]:
        for cell in row:
            if cell is None:
                continue
            v_str = str(cell).upper()
            # Detect Hari
            if re.match(r'^HARI\s*:', v_str):
                hari = str(cell).split(':', 1)[-1].strip()
            elif re.match(r'^TGL\s*:', v_str) or re.match(r'^HARI\s*$', v_str):
                pass
        # Also check col 3 for colon-separated values when col 2 is label
        col2 = row[2] if len(row) > 2 else None
        col3 = row[3] if len(row) > 3 else None
        if col2 and col3:
            label = str(col2).strip().upper()
            val   = str(col3).strip()
            if label in ('HARI', 'HARI :', 'HARI:'):
                # e.g. ':   Kamis Pagi' or 'Kamis Pagi'
                hari = val.lstrip(':').strip()
            elif label in ('TGL', 'TGL :', 'TGL:'):
                tgl  = val.lstrip(':').strip()
            elif label in ('JAM', 'JAM :', 'JAM:'):
                jam  = val.lstrip(':').strip()
            elif label in ('REVISI', 'REVISI :', 'REVISI:'):
                revisi = val.lstrip(':').strip()

    # Parse data rows (skip header row itself and one blank row after)
    data_start_local = header_local_idx + 1
    rows_out = []
    row_counter = 0

    # Keywords that indicate a summary/aggregate row (not an actual job or break)
    SUMMARY_KEYWORDS = (
        'TOTAL STROKE', 'TOTAL TPT', 'TOTAL FINISH', 'TOTAL PROD',
        'PLAN STROKE', 'GSPH TOTAL', 'GSPH ITEM', 'TARGET GSPH', 'GSPH',
        'TOTAL', 'TOTAL PCS'
    )
    # A row is a summary row if job_master is one of these summary labels
    # (and has no valid start_time / finish_time)
    def is_summary_row(jm_str, start_t, finish_t):
        if not jm_str:
            return False
        jm_up = jm_str.strip().upper()
        # If it matches a known summary keyword, skip
        for kw in SUMMARY_KEYWORDS:
            if jm_up == kw or jm_up.startswith(kw):
                return True
        # If job_master is exactly 'PLAN' with no time → summary row
        if jm_up == 'PLAN' and start_t is None and finish_t is None:
            return True
        return False

    for local_i in range(data_start_local, len(section_rows)):
        row = section_rows[local_i]
        if not any(v is not None for v in row):
            continue

        # Read job_no: prefer job_no_main (JOB NO. with dot) column;
        # if that is empty/None, fall back to job_no_alt (JOB NO without dot).
        job_no_val_main = row_value(row, col_map.get('job_no_main'))
        job_no_val_alt  = row_value(row, col_map.get('job_no_alt'))
        # Use main if non-empty, else alt
        def _nonempty(v):
            if v is None: return False
            s = str(v).strip()
            return s not in ('', '0', '#N/A', '#REF!', '#VALUE!')
        job_no_val = job_no_val_main if _nonempty(job_no_val_main) else job_no_val_alt

        job_master_val = row_value(row, col_map.get('job_master'))
        plan_val_raw   = row_value(row, col_map.get('plan'))

        # Read start/finish times early so we can use them in summary-row detection
        start_raw_early  = row_value(row, col_map.get('start_time'))
        finish_raw_early = row_value(row, col_map.get('finish_time'))
        start_str_early  = fmt_time(start_raw_early)
        finish_str_early = fmt_time(finish_raw_early)

        # Check if summary/aggregate row
        jm_str = str(job_master_val).strip() if job_master_val is not None else ''
        jn_str = str(job_no_val).strip() if job_no_val is not None else ''
        if is_summary_row(jm_str, start_str_early, finish_str_early) or is_summary_row(jn_str, start_str_early, finish_str_early):
            row_type = 'summary'
            if 'TOTAL FINISH' in jm_str.upper() or 'TOTAL FNISH' in jm_str.upper():
                # Read the total finish row, then break completely
                pass 
            elif is_header_row(row, None):
                # If we hit another header row, we've entered a new table. Stop.
                break
        else:
            # Skip rows with no job identity at all (only for non-summary)
            if not job_no_val and not job_master_val:
                continue

        # Skip obviously invalid rows (row_no column has '#N/A' etc)
        job_no_str = str(job_no_val).strip() if job_no_val is not None else ''
        if job_no_str.startswith('#'):
            continue

        # Determine row type
        row_type = 'job'
        if is_summary_row(jm_str, start_str_early, finish_str_early) or is_summary_row(jn_str, start_str_early, finish_str_early):
            row_type = 'summary'
        else:
            check_str = job_no_str.upper()
            if not check_str and job_master_val:
                check_str = str(job_master_val).upper()
            # Remove spaces to match "BREAK TIME" as well, and support ISHOMA
            check_clean = check_str.replace(' ', '')
            if any(x in check_clean for x in ['ISTIRAHAT', 'CINGKORAK', 'BREAKTIME']) or 'ISHOMA' in check_str:
                row_type = 'break'

        # Read all fields
        def gv(key, default=None):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            v = row[idx]
            return v if v is not None else default

        # Re-use the already-computed start/finish times
        start_str  = start_str_early
        finish_str = finish_str_early

        act_start_raw  = gv('act_start')
        act_finish_raw = gv('act_finish')
        act_start_str  = fmt_time(act_start_raw)
        act_finish_str = fmt_time(act_finish_raw)

        row_no_val = gv('row_no')
        if row_no_val is None or str(row_no_val).strip() == '':
            row_counter += 1
            row_no_val = row_counter
        else:
            try:
                row_no_val = int(float(str(row_no_val)))
                if row_type == 'job':
                    row_counter = row_no_val
            except:
                row_counter += 1
                row_no_val = row_counter

        current_job_master = str(gv('job_master', '') or '').strip()
        if row_type == 'job':
            if not current_job_master and last_job_master:
                current_job_master = last_job_master
            elif current_job_master:
                last_job_master = current_job_master

        item = {
            'row_no':       safe_i(row_no_val),
            'row_type':     row_type,
            'job_master':   current_job_master,
            'type_plt':     str(gv('type_plt', '') or ''),
            'qty_plt':      safe_f(gv('qty_plt')),
            'keb_mtl':      safe_f(gv('keb_mtl')),
            'total_plt':    safe_f(gv('total_plt')),
            'job_no':       job_no_str,
            'each_part':    str(gv('each_part', '') or ''),
            'plan':         safe_f(plan_val_raw),
            'ok':           safe_f(gv('ok')),
            'repair':       safe_f(gv('repair')),
            'reject':       safe_f(gv('reject')),
            'total_mesin':  safe_i(gv('total_mesin')),
            'ct_detik':     safe_f(gv('ct_detik')),
            'process_time': safe_f(gv('process_time')),
            'reg_active':   safe_f(gv('reg_active')),
            'dct':          safe_f(gv('dct')),
            'mct':          safe_f(gv('mct')),
            'plan_dct':     safe_f(gv('plan_dct')),
            'tpt':          safe_f(gv('tpt')),
            'gsph_item':    safe_f(gv('gsph_item')),
            'start_time':   start_str,
            'finish_time':  finish_str,
            'act_start':    act_start_str,
            'act_finish':   act_finish_str,
            'keterangan':   str(gv('keterangan', '') or ''),
            'a1':           safe_f(gv('a1')),
            'a2':           safe_f(gv('a2')),
            'a3':           safe_f(gv('a3')),
            'a4':           safe_f(gv('a4')),
            'dt_menit':     safe_f(gv('dt_menit')),
            'total_pcs':    safe_f(gv('total_pcs')),
            'tpt_total':    safe_f(gv('tpt_total')),
        }
        rows_out.append(item)

        # ── SAFETY BRAKE: Stop parsing section if we hit TOTAL FINISH ──
        if row_type == 'summary' and ('TOTAL FINISH' in jm_str.upper() or 'TOTAL FNISH' in jm_str.upper()):
            break

    # ── VALIDATION ──
    validation_errors = []
    for r in rows_out:
        if r['row_type'] == 'job' and (str(r['job_master']).strip().upper() == 'JOB MASTER' or str(r['job_no']).strip().upper() == 'JOB NO.'):
            validation_errors.append(f"Header row leaked as data: row_no={r['row_no']} job_master='{r['job_master']}' job_no='{r['job_no']}'")
    if len(rows_out) == 0:
        validation_errors.append("Zero data rows parsed — possible header detection failure")

    return {
        'shift_name': sheet_name,
        'press_name': press_name,
        'hari':       hari,
        'tgl':        tgl,
        'jam':        jam,
        'revisi':     revisi,
        'rows':        rows_out,
        '_stats': {
            'section_total_rows': len(section_rows),
            'meta_rows_skipped':  header_local_idx if header_local_idx is not None else 0,
            'data_rows_output':   len(rows_out),
            'validation_errors':  validation_errors,
            'header_local_idx':   header_local_idx,
            'header_content':     [str(c) for c in section_rows[header_local_idx]] if header_local_idx is not None else None,
        },
    }

# ── sheet selection ────────────────────────────────────────────────────────────

def _match_sheet(name):
    """Find actual sheet name matching ``name`` (case-insensitive)."""
    name_upper = name.upper()
    for sn in sheetnames:
        if sn.upper() == name_upper:
            return sn
    return None

def choose_sheets(wb, target_shift_req):
    """
    Return list of sheet names to process, applying Rev-priority logic.
    If a Rev variant exists, the non-Rev base sheet is also included so that
    press sections present only in the non-Rev sheet (e.g. PRESS C, D) are not lost.
    """
    global sheetnames
    sheetnames = wb.sheetnames

    def base_sheets(base):
        """Return sheet names for a base shift. When a Rev variant exists,
        returns [Rev, non-Rev] so non-Rev can supply presses missing from Rev."""
        rev = f"{base} (Rev)"
        rev_variants = [rev]
        for sn in sheetnames:
            su = sn.upper()
            if base.upper() in su and 'REV' in su and sn != base:
                if sn not in rev_variants:
                    rev_variants.append(sn)
        rev_found = None
        for rv in rev_variants:
            actual = _match_sheet(rv)
            if actual:
                rev_found = actual
                break
        base_actual = _match_sheet(base)
        if rev_found:
            sheets = [rev_found]
            if base_actual:
                sheets.append(base_actual)
            return sheets
        if base_actual:
            return [base_actual]
        return []

    if target_shift_req != 'AUTO':
        # Single shift requested
        chosen = base_sheets(target_shift_req)
        if chosen:
            return chosen
        # Fall back to any sheet that matches
        for sn in sheetnames:
            if target_shift_req.upper() in sn.upper() and 'MASTER' not in sn.upper():
                return [sn]
        return []

    # AUTO — collect both shifts
    result = []
    for base in ['Shift Pagi', 'Shift Malam']:
        result.extend(base_sheets(base))

    if not result:
        # Last resort: any sheet with SHIFT in name
        for sn in sheetnames:
            if 'SHIFT' in sn.upper() and 'MASTER' not in sn.upper():
                result.append(sn)

    return result

# ── entry point ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    filepath          = sys.argv[1]
    original_name     = sys.argv[2] if len(sys.argv) > 2 else filepath
    target_shift_req  = sys.argv[3] if len(sys.argv) > 3 else 'AUTO'

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        sheets_to_process = choose_sheets(wb, target_shift_req)

        if not sheets_to_process:
            print(json.dumps({'error': 'Tidak ada sheet Shift Pagi / Shift Malam yang ditemukan di file ini.'}))
            sys.exit(1)

        result_sheets = {}

        for sn in sheets_to_process:
            ws = wb[sn]
            parsed = parse_sheet(ws, sn)
            result_sheets.update(parsed)

        # 1. Try to get date from the parsed 'tgl' cell of the first valid sheet
        upload_date = None
        for key, section in result_sheets.items():
            if section.get('tgl'):
                upload_date = extract_date(section['tgl'])
                if upload_date:
                    break

        # 2. Fallback to extracting from the filename
        if not upload_date:
            upload_date = extract_date(original_name)

        # 3. Fallback to current date
        if not upload_date:
            upload_date = datetime.now().strftime('%d MEI %Y').upper()

        wb.close()

        # Build import log (row counts per sheet / section)
        import_log = {}
        total_excel_rows   = 0
        total_data_rows    = 0
        total_meta_skipped = 0
        for key, section in result_sheets.items():
            sheet_name = key.split('|||')[0]
            if sheet_name not in import_log:
                import_log[sheet_name] = {
                    'sections': {},
                    'sheet_excel_rows': 0,
                }
            stats = section.get('_stats', {})
            sr = stats.get('section_total_rows', 0)
            ms = stats.get('meta_rows_skipped', 0)
            dr = stats.get('data_rows_output', 0)
            import_log[sheet_name]['sections'][section['press_name']] = {
                'section_rows':     sr,
                'meta_rows_before_header': ms,
                'data_rows_output': dr,
                'header_local_idx': stats.get('header_local_idx'),
                'header_content':   stats.get('header_content'),
            }
            import_log[sheet_name]['sheet_excel_rows'] += sr
            total_excel_rows   += sr
            total_meta_skipped += ms
            total_data_rows    += dr
        import_log['summary'] = {
            'total_excel_rows_scanned': total_excel_rows,
            'total_meta_rows_before_header': total_meta_skipped,
            'total_data_rows_output': total_data_rows,
        }

        # Remove _stats from output rows (not needed by PHP)
        for section in result_sheets.values():
            section.pop('_stats', None)

        # Remove empty sections
        final_sheets = {k: v for k, v in result_sheets.items() if v['rows']}

        if not final_sheets:
            print(json.dumps({'error': 'Tidak ada data job yang berhasil dibaca dari file Excel. Pastikan format file sesuai.'}))
            sys.exit(1)

        print(json.dumps({
            'success':     True,
            'upload_date': upload_date,
            'sheets':      final_sheets,
            'log':         import_log,
        }, ensure_ascii=False))

    except Exception as e:
        import traceback
        print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
        sys.exit(1)

if __name__ == '__main__':
    main()
