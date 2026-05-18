#!/usr/bin/env python3
import sys
import json
import warnings
import openpyxl
from datetime import datetime

warnings.filterwarnings('ignore')

if len(sys.argv) < 3:
    print(json.dumps({'error': 'Usage: python read_single_part_monthly.py <excel_file> <price_file>'}))
    sys.exit(1)

data_file = sys.argv[1]
price_file = sys.argv[2]

try:
    # 1. Load price data
    price_data = {}
    wb_price = openpyxl.load_workbook(price_file, data_only=True)
    ws_price = wb_price.active
    for row in range(6, ws_price.max_row + 1):
        job_no = ws_price.cell(row, 2).value
        cycle = ws_price.cell(row, 3).value
        customer = ws_price.cell(row, 4).value
        price = ws_price.cell(row, 5).value
        if job_no:
            job_no_clean = str(job_no).strip().upper()
            price_data[job_no_clean] = {
                'customer': str(customer).strip() if customer else '',
                'cycle_issue': int(cycle) if cycle and str(cycle).isdigit() else 1,
                'price': float(price) if price and str(price).replace('.','').replace(',','').isdigit() else 0
            }
    
    # 2. Read single part data (all sheets)
    wb_data = openpyxl.load_workbook(data_file, data_only=True)
    all_sheets = {}
    
    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        
        header_row = None
        default_cat = 'SINGLE PART'
        
        for r in range(1, 15):
            for c in range(1, 10):
                val = ws.cell(r, c).value
                if val:
                    v_str = str(val).upper()
                    if 'FINISH PART' in v_str:
                        default_cat = 'FINISH PART'
                    if 'JOB NO' in v_str:
                        header_row = r
            if header_row:
                break
        
        if not header_row:
            continue
        
        col_map = {}
        for c in range(1, 45):
            h_val = ws.cell(header_row, c).value
            if h_val:
                h_name = str(h_val).strip().upper()
                if h_name == 'JOB NO': col_map['job_no'] = c
                elif h_name == 'DATE' or h_name == 'TGL' or h_name == 'TANGGAL': col_map['date'] = c
                elif h_name == 'NO' or h_name == '#': col_map['no'] = c
                elif 'FINISH' in h_name and ('JOB' in h_name or 'NO' in h_name): col_map['job_no_finish'] = c
                elif 'PALLET' in h_name or 'PALL' in h_name: col_map['type_pallet'] = c
                elif 'VENDOR' in h_name: col_map['vendor'] = c
                elif 'MATERIAL' in h_name or 'FINISH PART' in h_name or 'KATEGORI' in h_name: col_map['category'] = c
                elif 'S. AWAL' in h_name or 'STOCK AWAL' in h_name: col_map['stock_awal'] = c
                elif 'ASSY' in h_name: col_map['assy'] = c
                elif 'DELIVERY' in h_name: col_map['delivery'] = c
                elif 'INCOMING' in h_name: col_map['incoming'] = c
                elif 'S. AKHIR' in h_name or 'STOK AKHIR' in h_name: col_map['stok_akhir'] = c
                elif 'PCS / DAY' in h_name or 'PCS/DAY' in h_name: col_map['pcs_day'] = c
                elif 'STRENGHT' in h_name or 'STRENGTH' in h_name: col_map['strength'] = c
                elif 'IAMI' in h_name or 'TANI' in h_name: col_map['iami'] = c
                elif 'GKD' in h_name: col_map['gkd'] = c
                elif 'SAP' in h_name: col_map['sap'] = c
                elif 'KAP' in h_name: col_map['kap'] = c
                elif 'GMO' in h_name or 'TMMIN' in h_name or 'FTI' in h_name or 'PTI' in h_name or 'IKAR' in h_name: col_map['gmo'] = c
                elif 'CUSTOMER' in h_name: col_map['customer'] = c

        for row in range(header_row + 1, ws.max_row + 1):
            row_date = str(ws.cell(row, col_map.get('date', 0)).value).strip() if 'date' in col_map else sheet_name
            if not row_date or row_date == 'None' or row_date == '': row_date = sheet_name
            
            job_no = ws.cell(row, col_map.get('job_no', 2)).value
            vendor = ws.cell(row, col_map.get('vendor', 3)).value
            if not job_no or not vendor: continue
            
            job_no_finish = ws.cell(row, col_map['job_no_finish']).value if 'job_no_finish' in col_map else ''
            type_pallet = ws.cell(row, col_map['type_pallet']).value if 'type_pallet' in col_map else ''
            material = ws.cell(row, col_map['category']).value if 'category' in col_map else ''
            
            def safe_f(v):
                if v is None: return 0
                try: return float(v)
                except: return 0

            stock_awal = safe_f(ws.cell(row, col_map['stock_awal']).value) if 'stock_awal' in col_map else 0
            assy = safe_f(ws.cell(row, col_map['assy']).value) if 'assy' in col_map else 0
            incoming = safe_f(ws.cell(row, col_map['incoming']).value) if 'incoming' in col_map else 0
            stok_akhir = safe_f(ws.cell(row, col_map['stok_akhir']).value) if 'stok_akhir' in col_map else 0
            pcs_day = safe_f(ws.cell(row, col_map['pcs_day']).value) if 'pcs_day' in col_map else 0
            strength = safe_f(ws.cell(row, col_map['strength']).value) if 'strength' in col_map else 0
            
            iami = safe_f(ws.cell(row, col_map['iami']).value) if 'iami' in col_map else 0
            gkd = safe_f(ws.cell(row, col_map['gkd']).value) if 'gkd' in col_map else 0
            sap = safe_f(ws.cell(row, col_map['sap']).value) if 'sap' in col_map else 0
            kap = safe_f(ws.cell(row, col_map['kap']).value) if 'kap' in col_map else 0
            gmo = safe_f(ws.cell(row, col_map['gmo']).value) if 'gmo' in col_map else 0
            delivery = ws.cell(row, col_map['delivery']).value if 'delivery' in col_map else ''
            
            job_no_clean = str(job_no).strip().upper()
            price_info = price_data.get(job_no_clean, {'customer': '-', 'cycle_issue': 1, 'price': 0})
            
            customer = ws.cell(row, col_map.get('customer', 0)).value if 'customer' in col_map else price_info['customer']
            if not customer: customer = price_info['customer']
            
            price_pc = price_info['price']
            all_price = stok_akhir * price_pc
            
            category = str(material).strip().upper() if material else default_cat
            if 'FINISH' in category: category = 'FINISH PART'
            else: category = 'SINGLE PART'
            
            if pcs_day > 0 and strength == 0:
                strength = round(stok_akhir / pcs_day, 2)
            
            status = 'STANDAR'
            if strength <= 2: status = 'CRITICAL'
            elif strength > 5: status = 'OVER'
            
            if row_date not in all_sheets: all_sheets[row_date] = []
            
            no_val = ws.cell(row, col_map.get('no', 0)).value if 'no' in col_map else (len(all_sheets[row_date]) + 1)
            
            all_sheets[row_date].append({
                'no': no_val,
                'job_no': str(job_no).strip(),
                'job_no_finish': str(job_no_finish).strip() if job_no_finish else '',
                'type_pallet': str(type_pallet).strip() if type_pallet else '',
                'vendor': str(vendor).strip(),
                'category': category,
                'customer': str(customer).strip(),
                'price_pc': price_pc,
                'status': status,
                'movement': 'FAST MOVING' if strength > 0.5 else 'SLOW MOVING',
                'cycle_issue': price_info['cycle_issue'],
                'stock_awal': stock_awal,
                'assy': assy,
                'iami': iami,
                'gkd': gkd,
                'sap': sap,
                'kap': kap,
                'gmo': gmo,
                'delivery': str(delivery).strip() if delivery else '',
                'incoming': incoming,
                'stok_akhir': stok_akhir,
                'all_price': all_price,
                'pcs_day': pcs_day,
                'strength': strength,
            })
    
    print(json.dumps({'success': True, 'total_sheets': len(all_sheets), 'sheets': all_sheets}, ensure_ascii=False))

except Exception as e:
    print(json.dumps({'error': str(e)}))
