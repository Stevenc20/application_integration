#!/usr/bin/env python3
import sys, json, warnings
warnings.filterwarnings('ignore')  # suppress semua warning openpyxl

import openpyxl

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    filepath = sys.argv[1]

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # ── Sheet SPREEDSHEET ──
        spreedsheet = None
        for name in wb.sheetnames:
            if 'spreedsheet' in name.lower() or 'spreadsheet' in name.lower():
                spreedsheet = wb[name]
                break

        spreed_rows = []
        if spreedsheet:
            for row in spreedsheet.iter_rows(min_row=3, values_only=True):
                if not row[1]: continue
                spreed_rows.append({
                    'job_no':    str(row[1]  or '').strip(),
                    'item_name': str(row[2]  or '').strip(),
                    'proses':    str(row[4]  or '').strip(),
                    'source':    str(row[5]  or '').strip(),
                    'customer':  str(row[6]  or '').strip(),
                    'pcs_day':   float(row[7]  or 0),
                    'stock':     float(row[8]  or 0),
                    'strength':  float(row[9]  or 0),
                    'remarks':   str(row[10] or '').strip(),
                })

        # ── Sheet RUNDOWN STOCK FP ──
        rundown_sheet = None
        for name in wb.sheetnames:
            if 'rundown' in name.lower() and 'fp' in name.lower():
                rundown_sheet = wb[name]
                break

        rundown_rows = []
        if rundown_sheet:
            for row in rundown_sheet.iter_rows(min_row=7, values_only=True):
                if row[0] is None: continue
                try:
                    rundown_rows.append({
                        'no':             int(row[0]  or 0),
                        'job_no':         str(row[1]  or '').strip(),
                        'part_number':    str(row[3]  or '').strip(),
                        'sourching':      str(row[4]  or '').strip(),
                        'qty_palet':      float(row[5] or 0),
                        'type_pallet':    str(row[7]  or '').strip(),
                        'proses':         str(row[9]  or '').strip(),
                        'source':         str(row[11] or '').strip(),
                        'customer':       str(row[12] or '').strip(),
                        'type_of_part':   str(row[13] or '').strip(),
                        'stock_movement': str(row[14] or '').strip(),
                        'cycle_issue':    str(row[15] or '') if row[15] is not None else '',
                        'pcs_day':        float(row[16] or 0),
                        'stock_fg':       float(row[17] or 0),
                        'strength':       float(row[18] or 0),
                        'remarks':        str(row[19] or '').strip(),
                        'stock_sap':      float(row[20] or 0),
                        'stock_diff':     float(row[21] or 0),
                        'accuracy':       float(row[23] or 0),
                        'price_pcs':      float(row[24] or 0),
                        'new_price':      float(row[25] or 0),
                        'loss_gain':      float(row[26] or 0),
                        'pending_gi':     float(row[27] or 0),
                        'min_stock':      float(row[29] or 0) if row[29] else 0,
                        'max_stock':      float(row[30] or 0) if row[30] else 0,
                        'stock_shortage': float(row[33] or 0),
                        'status_order':   int(row[35]) if row[35] is not None else 0,
                    })
                except:
                    continue

        wb.close()

        # Output JSON murni ke stdout
        sys.stdout.write(json.dumps({
            'success': True,
            'spreedsheet': {'count': len(spreed_rows), 'data': spreed_rows},
            'rundown':     {'count': len(rundown_rows), 'data': rundown_rows},
        }))

    except Exception as e:
        sys.stdout.write(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()