#!/usr/bin/env python3
"""
read_schedule_stamping.py
Membaca file Schedule Stamping Excel (.xlsx/.xlsm).
1 file berisi data 4 press (PRESS A, B, C, D) yang disusun berurutan dalam 1 sheet.
Setiap press memiliki blok tersendiri, diawali dengan header "PRESS X" di kolom C.

Usage: python read_schedule_stamping.py <excel_file> [original_filename]
"""
import sys
import json
import warnings
import re
import os
from datetime import datetime, time
import openpyxl

warnings.filterwarnings('ignore')

BREAK_KEYWORDS = {
    'ISTIRAHAT SIANG', 'ISTIRAHAT SORE', 'ISTIRAHAT JUMAT',
    'ISTIRAHAT PAGI', 'CINGKORAK', 'BREAKTIME', 'BREAK TIME', 'BREAKTI',
    'ISTIRAHAT SORE RAMADHAN', 'TOTAL FNISH', 'TOTAL FINISH',
    'ISTIRAHAT', 'JUMAT', 'SORE', 'MALAM', 'PAGI', 'SIANG', 'BREAK'
}

MONTHS_ID = {
    'JAN':'01','FEB':'02','MAR':'03','APR':'04','MEI':'05','MAY':'05',
    'JUN':'06','JUL':'07','AGU':'08','AUG':'08','SEP':'09',
    'OKT':'10','OCT':'10','NOV':'11','DES':'12','DEC':'12',
}

# Keywords that mark the start of a PRESS section header in col C
PRESS_HEADER_KEYWORDS = ['PRESS A', 'PRESS B', 'PRESS C', 'PRESS D',
                         'PRESS-A', 'PRESS-B', 'PRESS-C', 'PRESS-D']

def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '.')
    if s in ('', '-', '#N/A', '#VALUE!', '#REF!', '0'):
        return None
    try:
        return float(s)
    except:
        return None

def safe_int(v):
    f = safe_float(v)
    return int(round(f)) if f is not None else None

def fmt_time(v):
    """Convert time/datetime/float to 'HH:MM' string."""
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime('%H:%M')
    if isinstance(v, datetime):
        return v.strftime('%H:%M')
    if isinstance(v, float):
        total_min = round(v * 24 * 60)
        h, m = divmod(total_min, 60)
        return f"{h % 24:02d}:{m:02d}"
    return None

def get_cell(row, col_0based):
    """Get value from row by 0-based index."""
    if col_0based is None:
        return None
    if col_0based < 0 or col_0based >= len(row):
        return None
    v = row[col_0based]
    if isinstance(v, str) and v.strip() in ('#N/A', '#VALUE!', '#REF!', ''):
        return None
    return v

def normalize_press_name(s):    
    """Normalize press name: 'PRESS-A' -> 'PRESS A'"""
    if not s:
        return None
    s = s.strip().upper().replace('-', ' ')
    m = re.match(r'(PRESS)\s*([A-D])', s)
    if m:
        return f"PRESS {m.group(2)}"
    return s

def is_press_header(row):
    """Check if row contains a PRESS name in ANY column."""
    for cell in row:
        if cell is not None:
            s = str(cell).strip().upper().replace('-', ' ')
            for pk in PRESS_HEADER_KEYWORDS:
                if pk.replace('-', ' ') in s:
                    # Return normalized version of the KEYWORD found
                    return normalize_press_name(pk)
    return False

def is_data_header(row):
    """Check if row is the column header row (NO. | JOB MASTER | ...)"""
    c2 = get_cell(row, 2)
    c3 = get_cell(row, 3)
    if c2 and 'NO' in str(c2).upper():
        return True
    if c3 and 'JOB MASTER' in str(c3).upper():
        return True
    return False

def is_summary_row(row):
    """Check if row marks the end of a data section."""
    keywords = ['PLAN', 'TOTAL STROKE', 'TOTAL TPT', 'TARGET GSPH', 'TOTAL FINISH', 'TOTAL FNISH', 'GSPH', 'TOTAL PCS']
    for cell in row:
        if cell is not None:
            s = str(cell).upper()
            if any(kw in s for kw in keywords):
                return True
    return False

def parse_data_row(row, row_no, col_map):
    """Parse one data row after data header using col_map."""
    job_master_idx = col_map.get('job_master')
    job_no_idx = col_map.get('job_no')
    
    jm_raw = get_cell(row, job_master_idx) if job_master_idx is not None else None
    jn_raw = get_cell(row, job_no_idx) if job_no_idx is not None else None
    
    jm_str = str(jm_raw).strip() if jm_raw is not None else ''
    jn_str = str(jn_raw).strip() if jn_raw is not None else ''

    # 1. Structural Detection
    row_no_raw = get_cell(row, 0)
    is_dash_row = (str(row_no_raw).strip() in ("—", "#N/A"))

    # 2. Keyword Detection
    full_row_text = " ".join([str(c).strip().upper() for c in row if c is not None])
    has_break_kw = any(kw in full_row_text for kw in BREAK_KEYWORDS)
    
    is_break = has_break_kw or (is_dash_row and not jm_str)
    
    if is_break:
        actual_desc = ""
        all_texts = []
        for cell_val in row:
            if cell_val is not None:
                s_val = str(cell_val).strip()
                # Skip numeric-only values and common symbols
                if s_val and s_val not in ("0", "—", "#N/A") and not s_val.replace('.','',1).isdigit():
                    # If it's a known keyword or looks like a label
                    all_texts.append(s_val)
        
        # Priority: Pick the one that actually contains a keyword
        keyword_texts = [t for t in all_texts if any(kw in t.upper() for kw in BREAK_KEYWORDS)]
        if keyword_texts:
            actual_desc = " ".join(keyword_texts)
        elif all_texts:
            actual_desc = " ".join(all_texts)
        
        # Override fields
        if actual_desc:
            jm_str = actual_desc
            jn_str = actual_desc

    if not is_break and not jm_str and not jn_str:
        return None

    def get_val(key, default=None, is_int=False, is_time=False):
        idx = col_map.get(key)
        if idx is None:
            return default
        val = get_cell(row, idx)
        if is_time:
            return fmt_time(val)
        if is_int:
            return safe_int(val)
        return safe_float(val)

    ok = get_val('ok') or 0
    repair = get_val('repair') or 0
    reject = get_val('reject') or 0
    total_pcs = ok + repair + reject

    type_plt_idx = col_map.get('type_plt')
    type_plt = str(get_cell(row, type_plt_idx) or '').strip() if type_plt_idx is not None else None
    
    each_part_idx = col_map.get('each_part')
    each_part = get_cell(row, each_part_idx)
    each_part_str = str(each_part).strip() if each_part and str(each_part).strip() not in ('None','') else None

    keterangan_idx = col_map.get('keterangan')
    keterangan = str(get_cell(row, keterangan_idx) or '').strip() if keterangan_idx is not None else None

    job_no_str = jn_str
    if is_break:
        job_no_str = jm_str

    return {
        'row_no':       row_no if not is_break else '—',
        'row_type':     'break' if is_break else 'job',
        'job_master':   jm_str,
        'job_no':       jn_str,
        'type_plt':     type_plt if not is_break else None,
        'qty_plt':      get_val('qty_plt'),
        'keb_mtl' :     get_val('keb_mtl'),
        'total_plt':    get_val('total_plt'),
        'job_no':       job_no_str or None,
        'each_part':    each_part_str,
        'plan':         get_val('plan'),
        'ok':           ok,
        'repair':       repair,
        'reject':       reject,
        'total_mesin':  get_val('total_mesin', is_int=True),
        'ct_detik':     get_val('ct_detik'),
        'process_time': get_val('process_time'),
        'reg_active':   get_val('reg_active'),
        'dct':          get_val('dct'),
        'mct':          get_val('mct'),
        'plan_dct':     get_val('plan_dct'),
        'tpt':          get_val('tpt'),
        'gsph_item':    get_val('gsph_item'),
        'start_time':   get_val('start_time', is_time=True),
        'finish_time':  get_val('finish_time', is_time=True),
        'act_start':    get_val('act_start', is_time=True),
        'act_finish':   get_val('act_finish', is_time=True),
        'keterangan':   keterangan,
        'a1':           get_val('a1', is_int=True),
        'a2':           get_val('a2', is_int=True),
        'a3':           get_val('a3', is_int=True),
        'a4':           get_val('a4', is_int=True),
        'dt_menit':     get_val('dt_menit', is_int=True),
        'total_pcs':    total_pcs,
        'tpt_total':    get_val('tpt_total'),
    }

def parse_sheet(ws, sheet_name):
    """Parse a single sheet."""
    all_rows = list(ws.iter_rows(min_row=1, max_row=500, max_col=50, values_only=True))

    global_meta = {'hari': None, 'tgl': None, 'jam': None, 'revisi': None}
    for ri in range(0, 30):
        if ri >= len(all_rows):
            break
        row = all_rows[ri]
        col4 = get_cell(row, 3)
        if col4 is None:
            continue
        s = str(col4).strip()
        if ':' in s:
            parts = s.split(':', 1)
            label = parts[0].upper()
            clean = parts[1].strip()
            if 'HARI' in label or any(d in clean.upper() for d in ['SENIN','SELASA','RABU','KAMIS','JUMAT','SABTU','MINGGU']):
                global_meta['hari'] = clean
            elif 'TGL' in label or re.search(r'\d{1,2}[-/]\w+[-/]\d{4}', clean):
                global_meta['tgl'] = clean
            elif 'JAM' in label or 'WIB' in clean.upper() or re.search(r'\d+:\d+', clean):
                global_meta['jam'] = clean
        
        for cell_val in row:
            if cell_val and 'Revisi' in str(cell_val):
                global_meta['revisi'] = str(cell_val).strip()
                break

    press_sections = {p: {
        'hari': global_meta['hari'],
        'tgl': global_meta['tgl'],
        'jam': global_meta['jam'],
        'revisi': global_meta['revisi'],
        'rows': []
    } for p in ['PRESS A', 'PRESS B', 'PRESS C', 'PRESS D']}
    current_press   = "PRESS A"
    in_data_section = False
    row_no          = 1
    current_meta    = dict(global_meta)
    col_map         = {}

    KEYWORD_MAP = {
        'JOB MASTER': 'job_master',
        'TYPE': 'type_plt',
        'QTY': 'qty_plt',
        'TOTAL PLT': 'total_plt',
        'JOB NO': 'job_no',
        'JOB.NO': 'job_no',
        'JOB': 'job_no',
        'PLAN': 'plan',
        'OK': 'ok',
        'MESIN': 'total_mesin',
        'CT': 'ct_detik',
        'PROC': 'process_time',
        'REG': 'reg_active',
        'DCT': 'dct',
        'MCT': 'mct',
        'TPT': 'tpt',
        'GSPH': 'gsph_item',
        'START': 'start_time',
        'FINISH': 'finish_time',
        'KETERANGAN': 'keterangan',
        'TOTAL PCS': 'total_pcs',
    }

    for ri, row in enumerate(all_rows):
        press_label = is_press_header(row)
        if press_label:
            current_press   = press_label
            in_data_section = False
            row_no          = 1
            current_meta    = dict(global_meta)
            for offset in range(1, 10):
                ni = ri + offset
                if ni >= len(all_rows):
                    break
                nrow = all_rows[ni]
                col2 = get_cell(nrow, 2)
                col3 = get_cell(nrow, 3)
                if col2 and str(col2).strip().upper().startswith('H') and col3 and ':' in str(col3):
                    current_meta['hari'] = str(col3).split(':', 1)[1].strip()
                elif col2 and str(col2).strip().upper().startswith('T') and col3 and ':' in str(col3):
                    current_meta['tgl'] = str(col3).split(':', 1)[1].strip()
                elif col2 and str(col2).strip().upper().startswith('J') and col3 and ':' in str(col3):
                    current_meta['jam'] = str(col3).split(':', 1)[1].strip()

            press_sections[current_press] = {
                'hari':  current_meta['hari'],
                'tgl':   current_meta['tgl'],
                'jam':   current_meta['jam'],
                'revisi': current_meta['revisi'],
                'rows':  press_sections.get(current_press, {}).get('rows', []),
            }
            continue

        if current_press is None:
            continue

        if is_data_header(row):
            in_data_section = True
            col_map = {}
            for cidx, cell_val in enumerate(row):
                if cell_val is None or cidx > 30:
                    continue
                s = str(cell_val).upper().strip()
                for kw, field in KEYWORD_MAP.items():
                    if kw in s:
                        if field not in col_map:
                            col_map[field] = cidx
                        elif kw == s and len(s) == len(kw):
                            pass 
            continue

        if not in_data_section:
            continue

        if is_summary_row(row):
            parsed = parse_data_row(row, row_no, col_map)
            if parsed:
                parsed['row_type'] = 'break' 
                press_sections[current_press]['rows'].append(parsed)
            continue

        parsed = parse_data_row(row, row_no, col_map)
        if parsed:
            press_sections[current_press]['rows'].append(parsed)
            if parsed['row_type'] == 'job':
                row_no += 1

    return press_sections

def extract_upload_date(filepath, original_name):
    """Extract date from filename."""
    name = os.path.basename(original_name or filepath).upper()
    for month_key, month_num in MONTHS_ID.items():
        pattern = rf'(\d{{1,2}})[-_\s]*{month_key}[-_\s]*(\d{{4}})'
        m = re.search(pattern, name)
        if m:
            day  = m.group(1).zfill(2)
            year = m.group(2)
            month_names = {
                '01':'JANUARI','02':'FEBRUARI','03':'MARET','04':'APRIL',
                '05':'MEI','06':'JUNI','07':'JULI','08':'AGUSTUS',
                '09':'SEPTEMBER','10':'OKTOBER','11':'NOVEMBER','12':'DESEMBER',
            }
            return f"{day} {month_names.get(month_num, month_key)} {year}"
    now = datetime.now()
    month_names = {1:'JANUARI',2:'FEBRUARI',3:'MARET',4:'APRIL',5:'MEI',6:'JUNI',
                   7:'JULI',8:'AGUSTUS',9:'SEPTEMBER',10:'OKTOBER',11:'NOVEMBER',12:'DESEMBER'}
    return f"{str(now.day).zfill(2)} {month_names[now.month]} {now.year}"

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    filepath  = sys.argv[1]
    orig_name = sys.argv[2] if len(sys.argv) >= 3 else filepath
    upload_date = extract_upload_date(filepath, orig_name)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames

        all_candidate_sheets = []
        for s in sheets:
            sl = s.lower().strip()
            if ('shift' in sl or 'press' in sl) and not any(k in sl for k in ['master', 'format', 'resume', 'template']):
                all_candidate_sheets.append(s)

        best_sheets = {}
        for s in all_candidate_sheets:
            clean = s.replace(' (Rev)', '').replace('(Rev)', '').replace('(rev)', '').strip()
            is_rev = 'rev' in s.lower()
            if clean not in best_sheets:
                best_sheets[clean] = s
            else:
                if is_rev:
                    best_sheets[clean] = s
        
        sheets_to_parse = list(best_sheets.values())
        parsed_result = {}

        for sname in sheets_to_parse:
            ws = wb[sname]
            press_sections = parse_sheet(ws, sname)
            for press_name, section_data in press_sections.items():
                if not section_data['rows']:
                    continue
                key = f"{sname}|||{press_name}"
                parsed_result[key] = {
                    'shift_name': sname,
                    'press_name': press_name,
                    'hari':   section_data['hari'],
                    'tgl':    section_data['tgl'],
                    'jam':    section_data['jam'],
                    'revisi': section_data['revisi'],
                    'rows':   section_data['rows'],
                }

        wb.close()

        if not parsed_result:
            print(json.dumps({'error': 'Tidak ada data schedule ditemukan.'}))
            sys.exit(1)

        print(json.dumps({
            'success': True,
            'upload_date': upload_date,
            'total_sheets': len(parsed_result),
            'sheets': parsed_result,
        }))

    except Exception as e:
        import traceback
        print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
        sys.exit(1)

if __name__ == '__main__':
    main()
