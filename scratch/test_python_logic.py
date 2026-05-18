import sys, json, os, warnings
from openpyxl import load_workbook
import importlib.util

spec = importlib.util.spec_from_file_location("rss", "dashboard-logistik/read_schedule_stamping.py")
rss = importlib.util.module_from_spec(spec)
spec.loader.exec_module(rss)

file_path = "07. Schedule Stamping 08 Mei 2026.xlsx"
wb = load_workbook(file_path, data_only=True, read_only=True)
ws = wb.active

print(f"Checking file: {file_path}")
all_rows = list(ws.iter_rows(min_row=1, max_row=400, values_only=True))

# Hardcoded col_map based on common stamping layout
col_map = {'job_master': 3, 'job_no': 6, 'keterangan': 28}

for ri, row in enumerate(all_rows):
    row_str = " ".join([str(c) for c in row if c is not None]).upper()
    if any(kw in row_str for kw in ['ISTIRAHAT', 'JUMAT', 'BREAK', 'CINGKORAK']):
        parsed = rss.parse_data_row(row, ri, col_map)
        if parsed and parsed['row_type'] == 'break':
            print(f"Line {ri+1} | JOB_NO: [{parsed.get('job_no')}] | MASTER: [{parsed.get('job_master')}] | ROW: {row[:10]}")
