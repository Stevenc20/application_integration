import openpyxl
import sys

def debug_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        if 'SHIFT PAGI' in sheet_name.upper():
            print(f"--- Checking Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True)):
                # Detect break-like rows
                full_text = " ".join([str(c).strip() for c in row if c is not None])
                if '—' in full_text or 'ISTIRAHAT' in full_text.upper() or 'CINGKORAK' in full_text.upper():
                    print(f"Row {ri+1}: {row}")
    wb.close()

if __name__ == "__main__":
    debug_excel(r'c:\MAMP\htdocs\application_integration\referensi\07. Schedule Stamping 08 Mei 2026.xlsx')
