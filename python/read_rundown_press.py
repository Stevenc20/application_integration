#!/usr/bin/env python3
"""
read_rundown_press.py
Reads Rundown Press Excel (.xlsm/.xlsx).
Fokus sheet bernama 'RUNDOWN'.
Setiap tanggal dideteksi dari baris 6 (row index 6).
Setiap tanggal punya group 7 kolom: MDFO, ORDER, PLAN DAY, PLAN NIGHT, ACT PROD, STOK AKHIR, STRENGTH

Usage: python read_rundown_press.py <excel_file> [original_filename]
"""
import sys, json, warnings, re, os
from datetime import datetime
import openpyxl

warnings.filterwarnings('ignore')

VALID_MONTHS = {
    'JANUARI','FEBRUARI','MARET','APRIL','MEI','JUNI',
    'JULI','AGUSTUS','SEPTEMBER','OKTOBER','NOVEMBER','DESEMBER'
}

MONTHS_ID = {
    1:'JANUARI', 2:'FEBRUARI', 3:'MARET', 4:'APRIL',
    5:'MEI', 6:'JUNI', 7:'JULI', 8:'AGUSTUS',
    9:'SEPTEMBER', 10:'OKTOBER', 11:'NOVEMBER', 12:'DESEMBER'
}

def safe_float(v):
    try:
        if v is None: return 0.0
        if isinstance(v, str):
            v = v.replace(',', '.').strip()
            if v in ('', '-', '#N/A', 'N/A'): return 0.0
        return float(v)
    except:
        return 0.0

def safe_int(v):
    try:
        if v is None: return None
        return int(float(str(v).strip()))
    except:
        return None

def extract_date_from_filename(filepath):
    name = os.path.basename(filepath).upper()
    for month in VALID_MONTHS:
        pattern = rf'(\d{{1,2}})\s*[-_]?\s*{month}'
        m = re.search(pattern, name)
        if m:
            day = m.group(1).zfill(2)
            return f"{day} {month}"
    return None

def date_to_sheet_label(dt):
    """Convert datetime to '04 MEI' format."""
    if isinstance(dt, datetime):
        return f"{str(dt.day).zfill(2)} {MONTHS_ID.get(dt.month, '')}"
    return None

def get_status(strength):
    if strength < 2:     return 'CRITICAL'
    elif strength < 5:   return 'STANDAR'
    else:                return 'OVER'

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No file path provided'}))
        sys.exit(1)

    filepath = sys.argv[1]
    arg_name = sys.argv[2] if len(sys.argv) >= 3 else filepath

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        # Find RUNDOWN sheet
        target_sheet = None
        for sn in wb.sheetnames:
            if 'RUNDOWN' in sn.strip().upper():
                target_sheet = sn
                break
        if not target_sheet:
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]

        # === STEP 1: Read all rows into memory ===
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not all_rows:
            print(json.dumps({'error': 'Sheet kosong'}))
            sys.exit(1)

        # === STEP 2: Find header row (contains 'JOB NO') ===
        header_row_idx = None
        for i, row in enumerate(all_rows[:20]):
            for cell in row:
                if cell and 'JOB NO' in str(cell).upper():
                    header_row_idx = i
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            print(json.dumps({'error': 'Header baris tidak ditemukan (tidak ada kolom JOB NO)'}))
            sys.exit(1)

        # === STEP 3: Detect date columns from row 6 (index 5) ===
        # Dates appear as datetime objects in row 6
        date_cols = {}  # {col_index_0based: 'DD BULAN'}
        date_row_idx = 5  # row 6 is index 5

        if date_row_idx < len(all_rows):
            date_row = all_rows[date_row_idx]
            for col_i, val in enumerate(date_row):
                if isinstance(val, datetime):
                    label = date_to_sheet_label(val)
                    if label:
                        date_cols[col_i] = label

        # If no dates detected from row 6, try to detect from filename
        if not date_cols:
            detected_date = extract_date_from_filename(arg_name) or extract_date_from_filename(filepath)
            if not detected_date:
                now = datetime.now()
                detected_date = f"{str(now.day).zfill(2)} {MONTHS_ID[now.month]}"
            # Fallback: use column 23 (0-based: 22) as stok_akhir for a single date
            # Based on the column structure: STOK AKHIR at col 23 (1-based), STRENGTH at 24
            date_cols = {23: detected_date}  # 0-based col 23 = 1-based col 24

        # === STEP 4: Map date columns to their data column offsets ===
        # Pattern from analysis:
        # Date is stored in col X (0-based). The 7-col group ends at X.
        # Group layout (0-based offsets from STOK AKHIR col):
        #   STOK AKHIR = X-1 (0-based), STRENGTH = X (0-based)
        # Wait - let's re-examine. From row 6:
        #   Dates are at 0-based: 23, 30, 37, 44, 51, 58, 65
        # From header row 7 (index 6), cols are:
        #   17(0-based)=STOK AWAL, 17+1=MDFO, 17+2=ORDER, 17+3=PLAN DAY, 17+4=PLAN NIGHT, 17+5=ACT PROD
        #   17+6=STOK AKHIR=23(0-based), 17+7=STRENGTH=24... but dates are at 23(0-based)
        # Actually from the data: dates at 0-based 23,30,37,44,51,58,65
        # And header: STOK AKHIR at 0-based 22, STRENGTH at 23
        # So date is at the STRENGTH col of each group!
        # Group starts at: date_col - 6 = MDFO, date_col-5=ORDER, ..., date_col-1=STOK AKHIR, date_col=STRENGTH
        
        # Build date group mapping
        date_groups = {}  # {'DD BULAN': {'mdfo':col, 'order':col, 'plan_day':col, 'plan_night':col, 'act_prod':col, 'stok_akhir':col, 'strength':col}}
        for col_i, label in date_cols.items():
            # col_i is 0-based STRENGTH column
            date_groups[label] = {
                'mdfo':       col_i - 6,
                'order':      col_i - 5,
                'plan_day':   col_i - 4,
                'plan_night': col_i - 3,
                'act_prod':   col_i - 2,
                'stok_akhir': col_i - 1,
                'strength':   col_i,
            }

        # === STEP 5: Parse fixed columns from header ===
        hdr = all_rows[header_row_idx]
        col_no         = None
        col_job_sched  = None
        col_job_deliv  = None
        col_job_stamp  = None
        col_keterangan = None
        col_variant    = None
        col_pcs_day    = None
        col_cust       = None
        col_qty_kbn    = None
        col_maker      = None
        col_stock_awal = None

        for ci, h in enumerate(hdr):
            if not h: continue
            hs = str(h).upper().replace('\n', ' ').strip()
            if hs == 'NO' and col_no is None:
                col_no = ci
            elif 'JOB NO' in hs and 'SCHEDULE' in hs:
                col_job_sched = ci
            elif 'JOB NO' in hs and 'DELIVERY' in hs:
                col_job_deliv = ci
            elif 'JOB NO' in hs and 'STAMPING' in hs:
                col_job_stamp = ci
            elif 'KETERANGAN' in hs and col_keterangan is None:
                col_keterangan = ci
            elif 'VARIANT' in hs and col_variant is None:
                col_variant = ci
            elif ('PCS/DAY' in hs or 'PCS / DAY' in hs) and col_pcs_day is None:
                col_pcs_day = ci
            elif hs in ('CUST', 'CUSTOMER') and col_cust is None:
                col_cust = ci
            elif 'QTY' in hs and 'KBN' in hs and col_qty_kbn is None:
                col_qty_kbn = ci
            elif 'MAKER' in hs and col_maker is None:
                col_maker = ci
            elif ('STOK AWAL' in hs or 'STOCK AWAL' in hs) and col_stock_awal is None:
                col_stock_awal = ci

        # Fallback to known positions (0-based) from analysis of user's screenshot
        if col_no is None:          col_no = 0
        if col_job_sched is None:   col_job_sched = 1
        if col_job_deliv is None:   col_job_deliv = 2
        if col_job_stamp is None:   col_job_stamp = 3
        if col_keterangan is None:  col_keterangan = 4
        if col_variant is None:     col_variant = 5
        if col_pcs_day is None:     col_pcs_day = 8
        if col_cust is None:        col_cust = 9
        if col_qty_kbn is None:     col_qty_kbn = 10
        if col_maker is None:       col_maker = 11
        if col_stock_awal is None:  col_stock_awal = 14

        # === STEP 6: Parse data rows per date ===
        all_results = {}
        data_start = header_row_idx + 1

        for label, cols in date_groups.items():
            sheet_data = []
            row_no = 1

            for row in all_rows[data_start:]:
                if not row or len(row) <= col_job_sched:
                    continue

                job_val = row[col_job_sched] if col_job_sched < len(row) else None
                if not job_val:
                    continue
                job_str = str(job_val).strip().upper()
                job_str = re.sub(r'\s+', ' ', job_str)  # normalize newlines/tabs to single space
                # Must look like a job number
                # More inclusive job number matching
                if len(job_str) < 3 or len(job_str) > 30:
                    continue
                if job_str in ('JOB NO', 'SCHEDULE', 'DELIVERY', 'STAMPING', 'JOB NO SCHEDULE', 'JOB NO DELIVERY', 'JOB NO STAMPING'):
                    continue
                # Skip summary/notes/total rows
                SKIP_KEYWORDS = ('TOTAL', 'SUBTOTAL', 'GRAND', 'JUMLAH', 'CATATAN', 'NOTE', 'NOTES', 'REMARK', 'SUMMARY', 'KETERANGAN', '*', 'RUNDOWN STOK', 'BLANKING')
                if any(kw in job_str for kw in SKIP_KEYWORDS):
                    continue

                def gcol_raw(c):
                    if c is None or c < 0 or c >= len(row): return None
                    return row[c]

                no_val        = gcol_raw(col_no)
                job_delivery  = gcol_raw(col_job_deliv)
                job_stamping  = gcol_raw(col_job_stamp)
                keterangan    = gcol_raw(col_keterangan)
                variant       = gcol_raw(col_variant)
                pcs_day       = safe_float(gcol_raw(col_pcs_day))
                cust          = gcol_raw(col_cust)
                qty_kbn       = safe_float(gcol_raw(col_qty_kbn))
                maker         = gcol_raw(col_maker)
                stock_awal    = safe_float(gcol_raw(col_stock_awal))

                mdfo_val      = gcol_raw(cols['mdfo'])
                order_val     = gcol_raw(cols['order'])
                plan_day_val  = gcol_raw(cols['plan_day'])
                plan_night_val = gcol_raw(cols['plan_night'])
                act_prod_val  = gcol_raw(cols['act_prod'])
                stok_akhir_val = gcol_raw(cols['stok_akhir'])
                strength_val  = gcol_raw(cols['strength'])

                mdfo          = safe_float(mdfo_val)
                order         = safe_float(order_val)
                plan_day      = safe_float(plan_day_val)
                plan_night    = safe_float(plan_night_val)
                act_prod      = safe_float(act_prod_val) if act_prod_val is not None else None
                stok_akhir    = safe_float(stok_akhir_val)
                strength      = safe_float(strength_val)

                # Validate col_no is numeric (filters summary/notes/hidden rows with text in NO column)
                if no_val is not None:
                    try:
                        int(float(str(no_val).strip()))
                    except (ValueError, TypeError):
                        continue  # col_no bukan angka → baris summary/notes

                # Skip rows where all key data columns are empty/zero (hidden or empty rows)
                if all(v in (None, 0, 0.0) for v in [mdfo, order, plan_day, plan_night, act_prod, stok_akhir]):
                    if not maker and not cust and not stock_awal:
                        continue

                if strength == 0 and pcs_day > 0 and stok_akhir != 0:
                    strength = round(stok_akhir / pcs_day, 4)

                status = get_status(strength)

                no_int = safe_int(no_val) if no_val else row_no

                sheet_data.append({
                    'no':           no_int or row_no,
                    'job_no':       job_str,
                    'job_delivery': str(job_delivery or '').strip().upper(),
                    'tipe':         str(job_stamping or '').strip().upper(),
                    'vendor':       str(maker or '').strip().upper(),
                    'keterangan':   str(keterangan or '').strip(),
                    'update_stock': str(variant or '').strip(),
                    'cust':         str(cust or '').strip().upper(),
                    'stock_awal':   stock_awal,
                    'price':        qty_kbn,
                    'pcs_day':      pcs_day,
                    # Date-specific columns mapped to DB fields:
                    'incoming':     mdfo,       # MDFO -> incoming
                    'spare_part':   order,      # ORDER -> spare_part
                    'plan_day':     plan_day,   # PLAN DAY -> plan_day
                    'plan_night':   plan_night, # PLAN NIGHT -> plan_night
                    'actual_prod':  act_prod,   # ACT PROD -> actual_prod
                    'iami':         0.0,
                    'gkd':          0.0,
                    'sap':          0.0,
                    'kap':          0.0,
                    'gmo':          0.0,
                    'stok_akhir':   stok_akhir,
                    'strength':     strength,
                    'status':       status,
                })
                row_no += 1

            if sheet_data:
                all_results[label] = sheet_data

        wb.close()

        if not all_results:
            print(json.dumps({'error': 'Tidak ada data valid ditemukan. Pastikan sheet bernama RUNDOWN dan format sesuai.'}))
            sys.exit(1)

        print(json.dumps({
            'success': True,
            'total_sheets': len(all_results),
            'sheets': all_results
        }))

    except Exception as e:
        import traceback
        print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
        sys.exit(1)

if __name__ == '__main__':
    main()